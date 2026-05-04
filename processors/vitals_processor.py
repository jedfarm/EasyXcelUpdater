#VERSION: 1.0.0

import os
import re
import pandas as pd
from pathlib import Path
import shutil
import numpy as np
import threading
import datetime
import sys
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from utils import data_migration_resources as dmr
from utils.exceptions import AbortedByUser
from utils.file_utils import resource_path
#root_abs_path = Path(__file__).resolve().parent.parent.parent


def process_vitals_file(
        facility_code,
        facility_name,
        facility_type,
        input_file_path,
        output_file_path,
        log_fn=None,
        abort_event=threading.Event()
):
    if log_fn is None:
       log_fn = lambda msg: None

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    if not isinstance(output_file_path, Path):
        output_file_path = Path(output_file_path)   

    # Read the allergies file
    try:
        col_names = ['Date Taken', 'Time', 'Vital', 'BMI', 'Value',
                     'Details', 'Taken By']
        
        # Load the excel file to get sheet names.
        excel_file = pd.ExcelFile(input_file_path)
        
        # Check the header on the first sheet to ensure it matches the expected column names.
        first_sheet_name = excel_file.sheet_names[0]
        # Read only the header (nrows=0 returns an empty DataFrame with columns from the first row).
        df_first_sheet_header = \
            pd.read_excel(input_file_path, sheet_name=first_sheet_name,
                          nrows=0)
        actual_header = list(df_first_sheet_header.columns)
        
        if actual_header != col_names:
            error_message = "The column names in the file do not match the expected column names."
            log_fn(error_message)
            log_fn(f"Expected: {col_names}, but got: {actual_header}")
            log_fn("Please check the file format and try again.")
            raise ValueError(error_message)
        
        # Optimize reading based on the number of sheets.
        if len(excel_file.sheet_names) == 1:
            # Single sheet: read directly with the expected column names.
            df_vitals_raw = pd.read_excel(input_file_path,
                                          sheet_name=first_sheet_name,
                                          names=col_names,
                                          usecols=[0, 1, 2, 3, 4, 5, 6])
            log_fn(f"✅ Single sheet found: {first_sheet_name}. Reading data directly.")   
        else:
            # Multiple sheets: read each sheet into its own DataFrame and then concatenate.
            dataframes = {
                sheet_name: pd.read_excel(input_file_path, sheet_name=sheet_name, 
                                          usecols=[0, 1, 2, 3, 4, 5, 6],
                                          names=col_names)
                for sheet_name in excel_file.sheet_names
            }
            df_vitals_raw = pd.concat(dataframes.values(), ignore_index=True)
            log_fn(f"✅ Multiple sheets found: {list(dataframes.keys())}. Concatenating data.")

    except Exception as e:
        error_message = f"Error reading vitals file: {str(e)}"
        log_fn(error_message)
        raise Exception(error_message)

    log_fn(f"✅ Successfully read the vitals file: {input_file_path}")  

    df_vitals_c1 = df_vitals_raw.dropna(how='all').reset_index(drop=True)
    df_vitals_c1 = df_vitals_c1.copy()
    if df_vitals_c1.empty:
        log_fn("⚠️ No data found in the vitals file.")
        raise ValueError("The vitals file is empty or does not contain valid data.")    

    resid_pattern = r'MR#: *([A-Z0-9#]+(?:-\d+)?)'
    resident_pattern = r"^\s*(?P<resident>(?P<lastname>[A-Za-z][A-Za-z\-'\.]*(?:\s+[A-Za-z\-'\.]+)*),\s+(?P<firstname>[A-Za-z][A-Za-z\-'\.]*)(?:\s+(?P<middlename>[A-Za-z\-'\.]+))?)"

    # Define a function to extract the pattern from each cell.
    def extract_from_cell(cell):
        match = re.findall(resid_pattern, str(cell))
        return match[0] if match else None

    # Define a function to extract the pattern from an entire row.
    def extract_from_row(row):
        for cell in row:
            result = extract_from_cell(cell)
            if result is not None:
                return result
        return None

    # Apply the function across rows. This will check every cell in the row until a match is found.
    df_vitals_c1['Client_ID_Number'] = df_vitals_c1.apply(extract_from_row, axis=1)

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    def extract_resident(text):
        """Extracts a resident name from the beginning of a string if it follows the pattern.
        
        Args:
            text: A string potentially containing a name in the format
                'last_name, first_name [middle_name]'.
                
        Returns:
            The resident name substring if found; otherwise, None.
        """
        text = str(text).strip()  # Ensure we work with a stripped string.
        match = re.match(resident_pattern, text)
        if match:
            # Return the complete matched resident name.
            return match.group("resident").strip()
        return None

    # Apply the function to the "Date Taken" column, thereby creating a new "Resident" column.
    df_vitals_c1['Resident'] = df_vitals_c1['Date Taken'].apply(extract_resident)

    df_vitals_c1['Client_ID_Number'] = df_vitals_c1['Client_ID_Number'].ffill()
    df_vitals_c1['Resident'] = df_vitals_c1['Resident'].ffill()

    vitals_to_del = ['fluids', 'bedtime snack', 'pain', 'urine', 'lunch', 'bowel movement',
                 'dinner', 'breakfast', 'bmi:', 'am snack', 'pm snack', 'supplements',
                 'emesis']
    
    # Purging the dataframe from certain vitals we do not need
    df_vitals_c2 = df_vitals_c1[~df_vitals_c1['Vital'].str.lower().str.strip().isin(vitals_to_del)]

    # Reset the index if needed
    df_vitals_c2.reset_index(drop=True, inplace=True)
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    # Purging the dataframe from certain rows we do not longer need
    df_vitals_c3 = df_vitals_c2[~df_vitals_c2['Date Taken'].str.contains("mr#:", case=False, na=False)]
    df_vitals_c3 = df_vitals_c3.copy()
    df_vitals_c3.reset_index(drop=True, inplace=True)
    df_vitals_c3['Unit'] = None

    def convert_to_feet(feet_inches_str):
        # Split the input string to extract feet and inches
        feet = 0
        inches = 0
        
        if 'ft' in feet_inches_str.lower():
            feet_part = feet_inches_str.split('ft')[0].strip()
            feet = float(feet_part)
        
        if 'in' in feet_inches_str.lower():
            inches_part = feet_inches_str.split('ft,')[-1].split('in')[0].strip()
            inches = float(inches_part)
        
        # Convert inches to feet
        total_feet = round(feet + (inches / 12), 1)
        
        return total_feet

    def convert_c_to_f(row):
        """
        Check if the 'Value' column in a row contains a Celsius value (e.g. '36.5°C').
        If found, convert it to Fahrenheit and update the 'Value' field.
        """
        # Convert the value to a string (in case it's numeric or NaN)
        value_str = str(row['Value'])
        
        # Regex to match an optional minus sign, digits, optional decimal point & digits followed by "°C"
        pattern = r'(-?\d+\.?\d*)\s*°C'
        match = re.search(pattern, value_str)
        
        if match:
            # Extract the numeric Celsius value
            celsius = float(match.group(1))
            # Convert Celsius to Fahrenheit
            fahrenheit = (celsius * 9/5) + 32
            # Update the 'Value' with the converted value formatted to one decimal place and the unit °F
            row['Value'] = f"{fahrenheit:.1f}°F"
        
        return row

    # Apply the function to your DataFrame
    df_vitals_c3 = df_vitals_c3.apply(convert_c_to_f, axis=1)

    # Moving meassurement units away from vitals values

    def move_value(row):
        vital = "" if pd.isna(row['Vital']) else str(row['Vital']).lower()
        value_str = str(row['Value'])
        value_lower = value_str.lower()
        
        if 'temperature' in vital and '°f' in value_lower:
            #print("Temperature condition met for row:", row.name)
            row['Unit'] = '°F'
            row['Value'] = row['Value'].replace('°F', '', 1).strip()
        elif 'pulse' in vital and '/per minute' in value_lower:
            #print("Pulse condition met for row:", row.name)
            row['Unit'] = '/per minute'
            row['Value'] = row['Value'].replace('/per minute', '', 1).strip()
        elif 'respirations' in vital and '/per minute' in value_lower:
            #print("Respirations condition met for row:", row.name)
            row['Unit'] = '/per minute'
            row['Value'] = row['Value'].replace('/per minute', '', 1).strip()
        elif 'o2 saturation' in vital and '%' in value_lower:
            #print("O2 saturation condition met for row:", row.name)
            row['Unit'] = '%'
            row['Value'] = row['Value'].replace('%', '', 1).strip()
        elif 'blood pressure' in vital and 'mmhg' in value_lower:
            #print("Blood pressure condition met for row:", row.name)
            row['Unit'] = 'mmHg'
            row['Value'] = row['Value'].replace('mmHg', '', 1).strip()
        elif 'weight' in vital and 'lbs' in value_lower:
            #print("Weight condition met for row:", row.name)
            row['Unit'] = 'lbs'
            row['Value'] = row['Value'].replace('lbs', '', 1).strip()
        elif 'blood sugar' in vital and 'mg/dl' in value_lower:
            #print("Blood sugar condition met for row:", row.name)
            row['Unit'] = 'mg/dL'
            row['Value'] = row['Value'].replace('mg/dL', '', 1).strip()
        elif 'height' in vital and 'ft' in value_lower:
            #print("Height condition met for row:", row.name)
            row['Unit'] = 'ft'
            row['Value'] = convert_to_feet(row['Value'])
        
        return row

    # Apply the function and reassign the DataFrame
    df_vitals_c3 = df_vitals_c3.apply(move_value, axis=1)

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    # We do not want to lose any specifics in the column Details
    def merge_details(df):
        # Keep track of indices to drop after merging
        rows_to_drop = []
        
        # Iterate over DataFrame using positional indexing
        for i in range(len(df)):
            row = df.iloc[i]
            # Check if Date Taken, Vital, and Value are NaN and Details is not NaN.
            if (pd.isna(row['Date Taken']) and 
                pd.isna(row['Vital']) and 
                pd.isna(row['Value']) and 
                pd.notna(row['Details'])):
                
                # Find the previous row (searching backwards) that is "complete"
                j = i - 1
                while j >= 0:
                    prev_row = df.iloc[j]
                    if (pd.notna(prev_row['Date Taken']) and 
                        pd.notna(prev_row['Vital']) and 
                        pd.notna(prev_row['Value'])):
                        # Append the current row's Details to the previous row's Details.
                        # Using a space as a separator; you can change this if needed.
                        updated_details = str(prev_row['Details']) + " " + str(row['Details'])
                        # Update the previous row's Details value
                        df.iloc[j, df.columns.get_loc('Details')] = updated_details
                        break
                    j -= 1
                    
                # Mark the current row for deletion.
                rows_to_drop.append(df.index[i])
        
        # Drop all rows that only contained extra Details.
        df.drop(index=rows_to_drop, inplace=True)
        return df

    # Use the function on your DataFrame
    df_vitals_c3 = merge_details(df_vitals_c3)

    def remove_left_of_last_colon(details):
        # If the value is NaN, just return it.
        if pd.isna(details):
            return details
        # Find the rightmost colon.
        idx = details.rfind(":")
        if idx != -1:
            # Return the substring after the rightmost colon, stripped of extra whitespace.
            return details[idx+1:].strip()
        # If no colon exists, return the original value.
        return details

    # Apply the function to the "Details" column
    df_vitals_c3['Details'] = \
        df_vitals_c3['Details'].apply(remove_left_of_last_colon)

    df_vitals_c4 = df_vitals_c3.copy()

    vitals = ['temperature', 'pulse', 'respirations', 'o2 saturation',
              'blood sugar', 'blood pressure', 'height', 'weight']

    # Convert the first column to datetime format
    df_vitals_c4['WV_Date'] = pd.to_datetime(df_vitals_c4['Date Taken'], 
                                             errors='coerce')
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    # Rearranging the rows
    # Convert the 'Vital' column to a Categorical type with the desired order
    df_vitals_c4['Vital'] = \
        pd.Categorical(df_vitals_c4['Vital'].str.lower(), categories=vitals, ordered=True)

    # Sort by Client_ID_Number, Col2 (following the categorical order), and WV_Date (most recent first)
    df_vitals_c5 = df_vitals_c4.sort_values(by=['Resident', 'Vital', 'WV_Date'], ascending=[True, True, False])

    df_vitals_c5 = df_vitals_c5[~pd.isna(df_vitals_c5['Value'])]

    df_vitals_c5 = df_vitals_c5.drop_duplicates(subset=['Client_ID_Number', 
                                                        'Vital', 'Value'])
    df_vitals_c5.reset_index(drop=True, inplace=True)
    df_vitals_c5 = df_vitals_c5.copy()

    # Dictionary for assigning Std_Vitals_ID
    vital_type_to_id = {'weight': 1, 'temperature': 5, 'blood sugar': 7,
                        'pulse': 6, 'height': 8, 'respirations': 2,
                        'o2 saturation': 9}

    df_vitals_c5['Std_Vitals_ID'] = np.nan

    # Function to split blood pressure values and create new rows
    def split_blood_pressure(df):
        new_rows = []
        for idx, row in df.iterrows():
            if 'blood pressure' in str(row['Vital'])  and '/' in row['Value']:
                systolic, diastolic = row['Value'].split('/')
                new_row_systolic = row.copy()
                new_row_diastolic = row.copy()
                new_row_systolic['Value'] = systolic
                new_row_systolic['Std_Vitals_ID'] = 3
                new_row_diastolic['Value'] = diastolic
                new_row_diastolic['Std_Vitals_ID'] = 4
                new_rows.append(new_row_systolic)
                new_rows.append(new_row_diastolic)
                
            else:
                new_rows.append(row)
        return pd.DataFrame(new_rows)

    # Apply the function to split blood pressure values
    df_vitals_c6 = split_blood_pressure(df_vitals_c5)

    # Reset the index to ensure it's sequential
    df_vitals_c6.reset_index(drop=True, inplace=True)

    # Function to assign Std_Vitals_ID based on VITAL_TYPE
    def assign_std_vitals_id(df, vital_type_to_id):
        def assign_id(row):
            if 'blood pressure' in str(row['Vital']):
                return row['Std_Vitals_ID']
            return vital_type_to_id.get(row['Vital'], row['Std_Vitals_ID'])
        
        df['Std_Vitals_ID'] = df.apply(assign_id, axis=1)
        return df

    # Apply the function to assign Std_Vitals_ID
    df_vitals_c6 = assign_std_vitals_id(df_vitals_c6, vital_type_to_id)

    df_vitals_c6['Facility_Code'] = facility_code

    df_vitals_c6.rename(columns={'Details': 'VITAL_DESCRIPTION'}, inplace=True)
    df_vitals_c6['VITAL_DESCRIPTION'] = \
        df_vitals_c6['VITAL_DESCRIPTION'].fillna('')
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    # Remove duplicate rows from df_vitals_c6
    df_vitals_c7 = df_vitals_c6.drop_duplicates()

    # Reset the index after removing duplicates (optional)
    df_vitals_c7 = \
        df_vitals_c7.reset_index(drop=True)
    
    # Ensure non-numeric values are converted to NaN
    df_vitals_c7['Col11'] = pd.to_numeric(df_vitals_c7['Value'],
                                          errors='coerce')

    # Drop rows where 'Value' is NaN (non-numeric) or 0
    df_vitals_c8 = df_vitals_c7[(df_vitals_c7['Col11'].notna()) & (df_vitals_c7['Col11'] != 0)]
    df_vitals_c8 = df_vitals_c8.reset_index(drop=True)

    df_vitals_c8['Date Taken'] = \
        df_vitals_c8['Date Taken'].apply(dmr.format_date_or_datetime)

    columns_to_keep = ['Facility_Code', 'Client_ID_Number', 'Std_Vitals_ID',
                       'Date Taken', 'Value', 'VITAL_DESCRIPTION', 'Resident']

    df_vitals = df_vitals_c8[columns_to_keep].copy()

    def adjust_value(row):
        # Only adjust rows where Std_Vitals_ID equals 9.
        if row['Std_Vitals_ID'] == 9:
            try:
                # Try to convert the value to float.
                numeric_val = float(row['Value'])
                # If the value is between 0 and 1, multiply by 100.
                if 0 < numeric_val < 1:
                    return numeric_val * 100
                else:
                    return row['Value']
            except (ValueError, TypeError):
                # If conversion fails, return the original value.
                return row['Value']
        else:
            return row['Value']

    # Apply the function to each row in df_vitals_c4.
    df_vitals['Value'] = df_vitals.apply(adjust_value, axis=1)

    # Adjust the path as necessary.
    template_path = resource_path(
        "resources", "templates", "WEIGHTS_AND_VITALS.xlsx")
    if not template_path.exists():
        error_message = f"Template file not found at {template_path}"
        log_fn("⚠️ " + error_message)
        raise FileNotFoundError(error_message)
    output_file_path.parent.mkdir(parents=True, exist_ok=True)
    # Copy the template file to the output_file_path (this creates the new file based on the template)
    shutil.copy(template_path, output_file_path)

    # Load the workbook from the newly created file
    wb = load_workbook(output_file_path)

    # Select the "Data" sheet; ensure the sheet name matches exactly
    ws = wb['Data']

    # Assume that the first row in the template is the header.
    # We will write the data starting at row 2.
    start_row = 2

    # Get the dataframe containing your final allergy data.
    # Make sure that the dataframe's columns match the header in the template.
    for r_idx, row in enumerate(df_vitals.itertuples(index=False),
                                start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for dv in ws.data_validations.dataValidation:
        # dv.ranges is a list of range strings (e.g. ["A2:A100"])
        new_ranges = []
        for rng in dv.ranges:
            # Parse the original range boundaries (start_col, start_row, end_col, end_row)
            rng_str = str(rng)  # ensure we pass a string, e.g. "A2:F100"
            min_col, min_row, max_col, max_row = range_boundaries(rng_str)

            # Update the end row to the current maximum row with data (ws.max_row)
            new_range = (
                f"{ws.cell(row=min_row, column=min_col).coordinate}:"
                f"{ws.cell(row=ws.max_row, column=max_col).coordinate}"
            )
            new_ranges.append(new_range)
        # Replace the validation's range list with the updated ranges
        dv.ranges = new_ranges
        # Save the workbook to preserve changes
    wb.save(output_file_path)

    wb.close()
