#VERSION: 1.0.1

import os
import pandas as pd
from pathlib import Path
import sys
import numpy as np
import re
import threading
from utils import data_migration_resources as dmr
from dateutil import parser
import shutil
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from utils.exceptions import AbortedByUser
from utils.file_utils import resource_path
from utils.file_utils import get_template_path

root_abs_path = Path(__file__).resolve().parent.parent.parent

def process_diagnosis_file(
        facility_code,
        facility_name,
        facility_type,
        input_file_path,
        output_file_path,
        log_fn=None,
        abort_event=threading.Event()
):
    """
    Processes a diagnosis file, performs data cleaning, filtering, and formatting,
    and writes the results to an Excel file.

    Args:
        facility_code (str): The code of the facility.
        facility_name (str): The name of the facility.
        diagnosis_file_path (str): The path to the input diagnosis file.
        output_file_path (str): The path to the output Excel file.
        is_filtering_residents_needed (bool, optional): Whether to filter residents. Defaults to False.
        resdem_file_path (str, optional): The path to the resident demographics file. Defaults to None.
        debug_messages (list, optional): A list to accumulate debug messages. Defaults to None.

    Raises:
        Exception: If there is an error reading or writing files, or during data processing.

    Returns:
        str: The path to the output Excel file.
    """
    if log_fn is None:
       log_fn = lambda msg: None

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    if not isinstance(output_file_path, Path):
        output_file_path = Path(output_file_path)

    # Read the diagnosis file
    try:
        df_diagnosis_raw = pd.read_excel(input_file_path, header=0)
        log_fn(f"Successfully read diagnosis file: {input_file_path}")
    except Exception as e:
        error_message = f"Error reading diagnosis file: {str(e)}"
        log_fn(error_message)
        raise Exception(error_message)

    all_cols_in = \
        dmr.import_columns_checker('Diagnosis',
                                   df_diagnosis_raw.columns.tolist())
    if not all_cols_in[0]:
        log_fn("⚠️ Not all expected columns are present in the Diagnosis file.")
        log_fn(f"Missing: {all_cols_in[1]}")
        log_fn("⚠️ Cannot process this file at this time.")
        raise ValueError("Diagnosis file does not have the expected columns.")

    log_fn("✅ Diagnosis file successfully loaded.")

    # PROCESSING DATA
    df_diagnosis_c1 = dmr.remove_row_chunks(df_diagnosis_raw, 'Name', 
                            'Obsolete diagnosis', 'Name', 'Name', 1)
    df_diagnosis_c1 = df_diagnosis_c1.reset_index(drop=True).copy()

    # Fixing resident names split over two rows in column Name
    for i in range(len(df_diagnosis_c1) - 1):
        # Grab the current and next rows
        current_row = df_diagnosis_c1.iloc[i]
        next_row = df_diagnosis_c1.iloc[i+1]
        
        # Check the conditions:
        # 1. Both rows have a non-NaN value in 'Name'
        # 2. The current row has non-NaN in 'Census' and 'MR#'
        # 3. The next row has NaN in 'Census' and 'MR#'
        if (pd.notna(current_row['Name']) and pd.notna(next_row['Name']) and
            pd.notna(current_row['Census']) and pd.notna(current_row['MR#']) and
            pd.isna(next_row['Census']) and pd.isna(next_row['MR#'])):
            
            # Append the next row's 'Name' to the current row's 'Name'
            appended_name = str(current_row['Name']) + " " + str(next_row['Name'])
            df_diagnosis_c1.at[df_diagnosis_c1.index[i], 'Name'] = appended_name
            
            df_diagnosis_c1.at[df_diagnosis_c1.index[i+1], 'Name'] = np.nan


    # Create a column with boolean values based whether Code contains a valid ICD10 or not
    df_diagnosis_c1['Is_ICD10'] = df_diagnosis_c1['Code'].apply(dmr.is_icd10)

    def find_icd10_in_row(row):
        # If Is_ICD10 is True, set WhereIsCode to "Code"
        if row['Is_ICD10']:
            return "Code"
        
        # Otherwise, check the rest of the columns
        matching_columns = []
        for col in row.index:
            if col != 'Is_ICD10' and col != 'Code' and dmr.is_icd10(row[col]):
                matching_columns.append(col)
        
        return ', '.join(matching_columns) if matching_columns else None

    # For rows where Is_Code is False, let's find where the ICD10 is, if any
    df_diagnosis_c1['WhereIsCode'] = df_diagnosis_c1.apply(find_icd10_in_row, axis=1)

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    # Define a function to extract the ICD code based on the first column referenced in WhereIsCode
    def extract_icd_code(row):
        # Check if WhereIsCode has a reference
        if pd.notna(row['WhereIsCode']):
            # Get the first column name from WhereIsCode
            first_column = row['WhereIsCode'].split(',')[0].strip()
            # Return the value from the referenced column
            return row[first_column]
        return None

    # Create a new column with ICD10 vales found in multiple columns
    df_diagnosis_c1['ICD_Code'] = df_diagnosis_c1.apply(extract_icd_code, axis=1)

    # Now that a thorough search for ICD10 is done, NaN rows based on ICD_Code makes sense.
    df_diagnosis_c2 = \
        df_diagnosis_c1.dropna(subset=['ICD_Code']).reset_index(drop=True)
    df_diagnosis_c2 =df_diagnosis_c2.copy()

    #log_fn(f"df_diagnosis_c2: Rows: {len(df_diagnosis_c2)}, Columns: {df_diagnosis_c2.columns}")

    def contains_date(s):
        try:
            s = str(s)
            # Search for date patterns in the string
            date_patterns = [
                r'(?<!\d)(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})(?!\d)',  # Matches dates like 11/21/2021, 21-11-2021
                r'(?<!\d)(\d{4}[/-]\d{1,2}[/-]\d{1,2})(?!\d)'     # Matches dates like 2021-11-21
            ]
            for pattern in date_patterns:
                if re.search(pattern, s):
                    return True
            return False
        except:
            return False

    df_diagnosis_c2['Is_Date'] = \
        df_diagnosis_c2['Diagnosed'].apply(contains_date)

    def extract_date(s):
        try:
            s = str(s)
            # Search for date patterns and return the first match
            date_patterns = [
                r'(?<!\d)(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})(?!\d)',  # Matches dates like 11/21/2021, 21-11-2021
                r'(?<!\d)(\d{4}[/-]\d{1,2}[/-]\d{1,2})(?!\d)'     # Matches dates like 2021-11-21
            ]
            for pattern in date_patterns:
                match = re.search(pattern, s)
                if match:
                    # Try to parse the found date to validate it
                    date_str = match.group(1)  # Use group(1) to get the date portion
                    parsed_date = parser.parse(date_str)
                    return parsed_date.strftime('%Y-%m-%d')  # Return the date in standard format
            return None
        except Exception as e:
            # Optionally, you can print or log the exception for debugging
            log_fn(f"Error parsing date: {e}")
            return None

    # Define a function to populate the WhereIsDate column
    def find_date_in_row(row):
        # If Is_Date is True, set WhereIsDate to "Diagnosed"
        if row['Is_Date']:
            return "Diagnosed"
        
        # Otherwise, check the rest of the columns
        matching_columns = []
        for col in row.index:
            if col != 'Is_Date' and col != 'Diagnosed' and contains_date(row[col]):
                matching_columns.append(col)
        
        return ', '.join(matching_columns) if matching_columns else None

    # Apply the function to each row to create the WhereIsDate column
    df_diagnosis_c2['WhereIsDate'] = df_diagnosis_c2.apply(find_date_in_row, axis=1)
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    # Slice the dataframe to find rows where WhereIsDate references more than one column
    # This MUST be addressed case by case. Check before go on
    filtered_df = df_diagnosis_c2[df_diagnosis_c2['WhereIsDate'].notna()]
    multiple_references_df = \
        filtered_df[filtered_df['WhereIsDate'].str.contains(',')]
    if multiple_references_df.shape[0]:
        message = f"Please check, there are {multiple_references_df.shape[0]} rows in multiple_references_df worth checking"
        log_fn(message)
               
    def extract_onset_date(row):
        # Check if WhereIsDate has a reference
        if pd.notna(row['WhereIsDate']):
            # Get the first column name from WhereIsDate
            first_column = row['WhereIsDate'].split(',')[0].strip()
            # Return the value from the referenced column
            return extract_date(row[first_column])
        return None

    df_diagnosis_c2['Onset_Date'] = \
        df_diagnosis_c2.apply(extract_onset_date, axis=1)

    # Filling up the gaps in MR#
    df_diagnosis_c2['MR#'] = df_diagnosis_c2['MR#'].ffill()
    df_diagnosis_c2['Name'] = df_diagnosis_c2['Name'].ffill()

    # Initialize the new column with default value 1
    df_diagnosis_c2['Diag_Classification_ID'] = 'Admission'

    # Iterate over unique MR# values
    for mr in df_diagnosis_c2['MR#'].unique():      
        # Get a boolean mask for the current MR# value
        mask = df_diagnosis_c2['MR#'] == mr      
        # Check for 'Primary' in 'Category'
        primary_mask = \
            df_diagnosis_c2.loc[mask, 
                                'Category'].str.strip().str.lower() =='primary'
        # Check for '(Admission)' in 'Description'
        admission_mask = df_diagnosis_c2.loc[mask, 'Description'].str.contains(
            'Admission', case=False, na=False)      
        if primary_mask.sum() > 0:
            df_diagnosis_c2.loc[mask & primary_mask, 
                                'Diag_Classification_ID'] =\
                  'Admitting Dx (#69)'
            continue  # Move to the next MR# value
        elif admission_mask.sum() > 0:
            df_diagnosis_c2.loc[mask & admission_mask, 
                                'Diag_Classification_ID'] = \
                                    'Admitting Dx (#69)'
            continue  # Move to the next MR# value
        # If neither condition was met, set the Diag_Classification_ID for the first row of this MR# 
        if not (primary_mask.any() or admission_mask.any()):
            first_index = df_diagnosis_c2.loc[mask].index[0]  # Get the index of the first row for this MR#
            df_diagnosis_c2.at[first_index, 'Diag_Classification_ID'] = 'Admitting Dx (#69)'

    # Create the "Rank_ID" column
    df_diagnosis_c2['Rank_ID'] = \
        np.where(df_diagnosis_c2['Diag_Classification_ID'] ==
                 'Admitting Dx (#69)',
                 'Primary Diagnosis (#67)',
                 'Other diagnosis')

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    rank_id_map = {
                    '1st Secondary':'Diagnosis A', 
                    '2nd Secondary': 'Diagnosis B',
                    '3rd Secondary': 'Diagnosis C',
                    '4th Secondary': 'Diagnosis D', 
                    '5th Secondary': 'Diagnosis E',
                    '6th Secondary': 'Diagnosis F',
                    '7th Secondary': 'Diagnosis G',
                    '8th Secondary': 'Diagnosis H',
                    '9th Secondary': 'Diagnosis I',
                    '10th Secondary': 'Diagnosis J',
                    '11th Secondary': 'Diagnosis K',
                    '12th Secondary': 'Diagnosis L',
                    '13th Secondary': 'Diagnosis M',
                    '14th Secondary': 'Diagnosis N',
                    '15th Secondary': 'Diagnosis O',
                    '16th Secondary': 'Diagnosis P',
                    '17th Secondary': 'Diagnosis Q'
    }

    def update_rank_id(row):
        if row['Rank_ID'] == 'Other diagnosis' and isinstance(row['Category'],
                                                              str):
            for key in rank_id_map:
                # Ensure exact word match using regex
                if re.search(rf'\b{re.escape(key)}\b', row['Category']):
                    return rank_id_map[key]
        return row['Rank_ID']

    # Apply the function to the dataframe
    df_diagnosis_c2['Rank_ID'] = df_diagnosis_c2.apply(update_rank_id, axis=1)

    # Ensure the column is treated as a string, then split on '-' and take the first part.
    df_diagnosis_c2['Client_ID_Number'] = \
        df_diagnosis_c2['MR#'].astype(str).str.split('-', expand=True)[0]
          
    df_diagnosis_c2['Facility_Code'] = facility_code
    df_diagnosis_c2['Resolved_Date'] = ''
    df_diagnosis_c2['Comments'] = ''
    df_diagnosis_c2['Confidential'] = 'N'
    df_diagnosis_c2['Therapy'] = 'N'

    df_diagnosis_c2['Name'] = df_diagnosis_c2['Name'].ffill()

    ordered_columns = ['Facility_Code', 'Client_ID_Number', 'ICD_Code', 'Onset_Date',
                    'Resolved_Date', 'Diag_Classification_ID', 'Rank_ID',
                    'Comments', 'Confidential', 'Therapy', 'Name']
    # Subset the DataFrame using the list ordered_columns
    df_diagnosis_c3 = df_diagnosis_c2[ordered_columns].copy()

    df_diagnosis_c3['Client_ID_Number'] = \
        df_diagnosis_c3['Client_ID_Number'].str.strip()
   
    df_diagnosis = df_diagnosis_c3.copy()

    df_diagnosis['Onset_Date'] = pd.to_datetime(df_diagnosis['Onset_Date'])
    df_diagnosis['Onset_Date'] = \
        df_diagnosis['Onset_Date'].dt.strftime('%m/%d/%Y')

    df_diagnosis = df_diagnosis.dropna(subset=['Facility_Code', 'Onset_Date',
                                               'Diag_Classification_ID', 'Rank_ID'],
                                       how='all')

    df_diagnosis.drop_duplicates(subset=['Client_ID_Number', 'Onset_Date',
                                         'ICD_Code'], inplace=True)
    df_diagnosis = df_diagnosis.reset_index(drop=True)

    log_fn("✅ Diagnosis file processed successfully.")
    log_fn(f"There are {df_diagnosis['Client_ID_Number'].nunique()} unique residents in the file.")
     
    # Adjust the path as necessary.
    template_path = get_template_path("CLIENT_DIAGNOSIS.xlsx")
    #template_path = resource_path("resources", "templates", "CLIENT_DIAGNOSIS.xlsx")

    if not template_path.exists():
        error_message = f"Template file not found at {template_path}"
        log_fn("⚠️ " + error_message)
        raise FileNotFoundError(error_message)
    output_file_path.parent.mkdir(parents=True, exist_ok=True)
    # Copy the template file to the output_file_path (this creates the new file based on the template)
    shutil.copy(template_path, output_file_path)

    # Load the workbook from the newly created file
    wb = load_workbook(output_file_path)

    # Select the "Data" sheet; ensure the sheet name matches exactly
    ws = wb['Data']

    # Assume that the first row in the template is the header.
    # We will write the data starting at row 2.
    start_row = 2

    # Get the dataframe containing your final diagnosis data.
    # In your code, this is likely stored in a variable such as df_diagnosis.
    # Make sure that the dataframe's columns match the header in the template.
    for r_idx, row in enumerate(df_diagnosis.itertuples(index=False), 
                                start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for dv in ws.data_validations.dataValidation:
        # dv.ranges is a list of range strings (e.g. ["A2:A100"])
        new_ranges = []
        for rng in dv.ranges:
            # Parse the original range boundaries (start_col, start_row, end_col, end_row)
            min_col, min_row, max_col, max_row = range_boundaries(rng)
            # Update the end row to the current maximum row with data (ws.max_row)
            new_range = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=ws.max_row, column=max_col).coordinate}"
            new_ranges.append(new_range)
        # Replace the validation's range list with the updated ranges
        dv.ranges = new_ranges
        # Save the workbook to preserve changes
    wb.save(output_file_path)
    wb.close()
