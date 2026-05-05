#VERSION: 1.0.0

import os
import pandas as pd
from pathlib import Path
import sys
import numpy as np
import re
import threading
from utils import data_migration_resources as dmr
from dateutil import parser
import shutil
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from fuzzywuzzy import fuzz
from openpyxl.utils import range_boundaries
from utils.exceptions import AbortedByUser
from utils.file_utils import resource_path
from utils.file_utils import get_template_path

root_abs_path = Path(__file__).resolve().parent.parent.parent



def process_contacts_file(
        facility_code,
        facility_name,
        facility_type,
        input_file_path,
        output_file_path,
        log_fn=None,
        abort_event=threading.Event()
):
    if log_fn is None:
       log_fn = lambda msg: None

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    if not isinstance(output_file_path, Path):
        output_file_path = Path(output_file_path)

    # DATA LOADING
    try:
        df_contacts_raw = pd.read_excel(input_file_path)
        #log_fn(f"Successfully read contacts file: {input_file_path}")
    except Exception as e:
        raise Exception(f"Error reading contacts file: {str(e)}")
    
    all_cols_in = \
        dmr.import_columns_checker('Contacts',
                                   df_contacts_raw.columns.tolist())
    if not all_cols_in[0]:
        log_fn("⚠️ Not all expected columns are present in the contacts file.")
        log_fn(f"Missing: {all_cols_in[1]}")
        log_fn("⚠️ Cannot process this file at this time. Please fix it before proceeding.")
        return
    log_fn("✅ Contacts file successfully loaded.")

    try:
        zip_to_county_db = resource_path("resources", "databases", 
                                         "zip_code_database.xlsx")
        df_zip_to_county = pd.read_excel(
            zip_to_county_db,
            usecols=['zip', 'primary_city', 'acceptable_cities', 'county',
                     'state']
        )
        df_zip_to_county['zip'] = df_zip_to_county['zip'].apply(
            lambda x: f"{int(x):05d}" if pd.notnull(x) else x
        )
    except Exception as e:
        log_fn(f"⚠️ Error: {e}")
        raise ValueError("The zip_county database cannot be found.")

    # DATA PROCESSING
    df_contacts_c1 = df_contacts_raw.copy()

    def extract_mr_info(row):
        # Retrieve the values for the three columns and ensure we have string representations.
        name_str = str(row.get("Name", "")) if pd.notnull(row.get("Name", "")) else ""
        rel_str = str(row.get("Relationship", "")) if pd.notnull(row.get("Relationship", "")) else ""
        resp_str = str(row.get("Responsibilities", "")) if pd.notnull(row.get("Responsibilities", "")) else ""
        
        client_id = None
        mr_found = False  # flag to indicate that we found a valid MR# token

        # Helper function: returns True if s is empty after stripping spaces.
        def is_empty(s):
            return s.strip() == ""

        # ----- Case 1: MR# at the start of the string -----
        # Priority is given in this order: check Relationship, then Name, then Responsibilities.
        if rel_str.strip().startswith("MR#"):
            # Example: "MR# 123456" becomes client id "123456"
            client_id = rel_str.strip()[len("MR#"):].strip()
            mr_found = True

        elif name_str.strip().startswith("MR#"):
            client_id = name_str.strip()[len("MR#"):].strip()
            mr_found = True

        elif resp_str.strip().startswith("MR#") and is_empty(rel_str):
            # We extract the token from Responsibilities using regex:
            # The pattern matches: "MR#" then optional spaces and one non-space token.
            match = re.match(r"^(MR#\s*\S+)", resp_str.strip())
            if match:
                token = match.group(1)
                # Copy the token to the Relationship column.
                row["Relationship"] = token
                # Remove the token from Responsibilities. Any text following is retained.
                row["Responsibilities"] = resp_str.strip()[len(token):].lstrip()
                client_id = token[len("MR#"):].strip()
                mr_found = True

        # ----- Case 2: MR# in the middle of Name and Relationship is empty -----
        if not mr_found and ("MR#" in name_str) and is_empty(rel_str):
            # Look for a token of the form "MR# <non-space characters>"
            match = re.search(r"(MR#\s*\S+)", name_str)
            if match:
                token = match.group(1)
                # Remove the token from Name (only the first occurrence).
                new_name = name_str.replace(token, "", 1).strip()
                row["Name"] = new_name
                client_id = token[len("MR#"):].strip()
                mr_found = True

        # ----- Write to the new columns -----
        if mr_found and client_id:
            row["Client_ID_Number"] = client_id
            
            # For Resident:
            # We assume that in rows where MR# was found, the Name column contains the resident name.
            # Remove any trailing substring that begins with "MR#".
            name_updated = str(row.get("Name", ""))
            resident_name = re.sub(r'\s*MR#.*$', '', name_updated).strip()
            row["Resident"] = resident_name
        else:
            # For rows with no MR# found, we can choose to leave Client_ID_Number blank.
            row["Client_ID_Number"] = None
            row["Resident"] = None  
        
        return row

    # Apply the extraction function to every row in df_contacts_c1.
    df_contacts_c1 = df_contacts_c1.apply(extract_mr_info, axis=1)

    indexes_to_delete = df_contacts_c1[df_contacts_c1['Client_ID_Number'].notna()].index.tolist()
    df_contacts_c1['Client_ID_Number'] = df_contacts_c1['Client_ID_Number'].ffill()
    df_contacts_c1['Resident'] = df_contacts_c1['Resident'].ffill()

    df_contacts_c2 = df_contacts_c1.drop(indexes_to_delete).copy()
    if df_contacts_c2.shape[0] == 0:
        log_fn("⚠️ No contacts data found after processing. Please check the input file.")
        raise Exception("No contacts data found after processing.")
    df_contacts_c2 = df_contacts_c2.reset_index(drop=True)

    # Removing rows that do not contain info.
    df_contacts_c3 = \
        df_contacts_c2[~df_contacts_c2['Name'].str.strip().isin(['Contacts', 
                                        'Name', 'MatrixCare Report'])].copy()   
    if df_contacts_c3.shape[0] == 0:
        log_fn("⚠️ No contacts data found after the 2nd step of processing. Please, check the input file.")
        raise Exception("No contacts data found after processing.")
    df_contacts_c3 = df_contacts_c3.reset_index(drop=True)   
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    # Removing Notes from the column Name and place them in a dedicated column
    if 'Notes' not in df_contacts_c3.columns:
        df_contacts_c3['Notes'] = None

    # Iterate over the DataFrame by index.
    # We assume the DataFrame index is integer-based. If not, you might want to reset it.
    for idx in df_contacts_c3.index:
        current_name = df_contacts_c3.loc[idx, 'Name']
        
        # Check if current cell is not NaN and its text (after stripping) starts with "Notes:"
        if pd.notna(current_name) and str(current_name).strip().startswith("Notes:"):
            # Extract note text: all the text after "Notes:".
            note_text = str(current_name).strip()[len("Notes:"):].strip()
            
            # Look upward from idx for the first row where the Name is not NaN.
            found_previous = False
            for prev_idx in range(idx - 1, -1, -1):
                if pd.notna(df_contacts_c3.loc[prev_idx, 'Name']):
                    # If the target row already has a note, we append the new note text.
                    if pd.notna(df_contacts_c3.loc[prev_idx, 'Notes']):
                        df_contacts_c3.loc[prev_idx, 'Notes'] = (
                            str(df_contacts_c3.loc[prev_idx, 'Notes']).strip() + " " + note_text
                        ).strip()
                    else:
                        df_contacts_c3.loc[prev_idx, 'Notes'] = note_text
                    found_previous = True
                    break  # Stop after finding the first valid row.
            
            # Replace the original "Name" cell containing the note with NaN.
            df_contacts_c3.loc[idx, 'Name'] = np.nan

    # Fixing names split over two rows
    for i in range(len(df_contacts_c3) - 1):
        # Check that both the current row and the next row have non-NaN values in the Name column.
        if pd.notna(df_contacts_c3.loc[i, 'Name']) and pd.notna(df_contacts_c3.loc[i+1, 'Name']):
            # For the current row, check that Relationship is not NaN
            # AND that at least one of Call Order or Phone/Email is not NaN.
            if (pd.notna(df_contacts_c3.loc[i, 'Relationship']) and 
                (pd.notna(df_contacts_c3.loc[i, 'Call Order']) or pd.notna(df_contacts_c3.loc[i, 'Phone/Email']))):
                # Now, check that the next row (the second row of the pair) has NaN in the extra info fields.
                if  (pd.isna(df_contacts_c3.loc[i+1, 'Call Order'])):
                    
                    # Append the text of the next row's Name to the current row's Name.
                    df_contacts_c3.loc[i, 'Name'] = (
                        str(df_contacts_c3.loc[i, 'Name']).strip() + " " + 
                        str(df_contacts_c3.loc[i+1, 'Name']).strip()
                    ).strip()
                    # Replace the second row's Name text with NaN, indicating it has been merged.
                    df_contacts_c3.loc[i+1, 'Name'] = np.nan
    
    # Certain info (such as Resposibilities, Address, etc.) is spread over multiple rows
    def content_condenser(df, target_column):
        for i in range(len(df)):
            # Check if the current row qualifies: Name and Responsibilities are not NaN.
            if pd.notna(df_contacts_c3.loc[i, 'Name']):
                # Store the existing text from the current row's target column.
                if pd.isna(df.loc[i, target_column]):
                    consolidated_text = ''
                else: 
                    consolidated_text = str(df.loc[i, target_column])
                j = i + 1
                # Loop to process consecutive rows below where Name is NaN and the target column is not NaN.
                while j < len(df) and pd.isna(df.loc[j, 'Name']) \
                    and pd.notna(df.loc[j, target_column]):
                    additional_text = str(df_contacts_c3.loc[j, target_column])
                    # Append the additional text (with a space separator) to the consolidated text.
                    consolidated_text += " " + additional_text
                    # Replace the text in this row with NaN since it has been moved.
                    df_contacts_c3.loc[j, target_column] = np.nan
                    j += 1
                # Update the original row's Responsibilities cell with the consolidated text.
                df.loc[i, target_column] = consolidated_text.strip()
        return df.copy()
    
    df_contacts_c4 = content_condenser(df_contacts_c3, 'Responsibilities')
    df_contacts_c4 = content_condenser(df_contacts_c3, 'Phone/Email')
    df_contacts_c4 = content_condenser(df_contacts_c3, 'Address')
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    # Making some cleaning and removing the rows with the cells emptied out from last step
    df_contacts_c4 = df_contacts_c4.replace(r"^\s*nan\s*$", np.nan, regex=True)
    df_contacts_c5 = df_contacts_c4.dropna(subset=['Relationship', 'Responsibilities', 
                    'Call Order', 'Phone/Email', 'Address'], how='all')
    df_contacts_c5 = df_contacts_c5.reset_index(drop=True)

    #Fixing minor issues in Name and Address columns
    df_contacts_c5 = df_contacts_c5.fillna('')

    remove_from_string = ['n/a', 'unknown', 'no-last-name']
    # re.escape is used to safely escape any special characters in the strings.
    pattern = '|'.join(map(re.escape, remove_from_string))
    # Apply the replacement on the Address column, ignoring case.
    df_contacts_c5['Address'] = df_contacts_c5['Address'].str.replace(pattern, '', 
                                                    flags=re.IGNORECASE, regex=True)
    df_contacts_c5['Address'] = df_contacts_c5['Address'].str.strip()
    df_contacts_c5['Name'] = df_contacts_c5['Name'].str.replace(pattern, '', 
                                                    flags=re.IGNORECASE, regex=True)
    df_contacts_c5['Name'] = df_contacts_c5['Name'].str.strip()

    # More cleaning for the Name column
    pattern = r'^(?P<title>(?:Mrs\s*\.?|Miss\s*\.?|Ms\s*\.?|Mr\s*\.?|Dr\s*.?))\s*'

    # Create the new column 'Title' by extracting the title portion from the 'Name' column.
    df_contacts_c5['Title'] = df_contacts_c5['Name'].str.extract(pattern, expand=False).str.strip()
    # Remove the title from the 'Name' column and strip leading whitespace.
    df_contacts_c5['Name'] = df_contacts_c5['Name'].str.replace(pattern, '', regex=True).str.lstrip()
    #df_contacts_c5['Name'] = df_contacts_c5['Name'].str.replace(',', '').str.title()

    # Columns that need to be checked for empty strings.
    cols_to_check = ['Relationship', 'Responsibilities', 'Phone/Email', 'Address']
    # Create a mask: True for rows where, after stripping, all the values in the selected columns are empty strings.
    mask_empty = \
        df_contacts_c5[cols_to_check].apply(lambda col: col.str.strip()).eq('').all(axis=1)

    # Filter out these rows by keeping only rows where the condition is False.
    df_contacts_c6 = df_contacts_c5[~mask_empty].copy()
    if df_contacts_c6.shape[0] == 0:
        log_fn("⚠️ No contacts data remains after the 5th step of processing. Please check the input file.")
        raise Exception("No contacts data found after processing.")
    df_contacts_c6 = df_contacts_c6.reset_index(drop=True)

    # Function to find the relation code
    def find_relation_code(relation):
        relation = str(relation).lower()  # Convert to lowercase
        for key, values in dmr.new_relation_codes.items():
            if 'other-' in relation:
                part_rel = relation[6:]
                if part_rel in values:
                    return key
                elif 'guardian' in part_rel:
                    return 'Guardian'
            if relation in values:
                return key
            
        return 'Other'  # Default value if no match is found

    # Apply the function to the 'Relationship' column
    df_contacts_c6['Relation_Code'] = \
        df_contacts_c6['Relationship'].apply(find_relation_code)
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    # Dealing with postal addresses involves several steps 
    # 1. Extract zip codes
    def extract_postal_code(address):
        """
        Extracts a US postal code from an address string if present at the end and removes it.

        This function searches the given address string for a US postal code at the very end.
        A valid postal code can be either a 5-digit code (e.g., "12345") or a 9-digit code in the 
        format "12345-6789". If a postal code is found, the function returns a tuple containing the 
        postal code and the remaining address string with the postal code removed. Leading and trailing 
        whitespace from the remaining address is also removed. If the resulting address becomes empty, 
        it is replaced with np.nan. If the address parameter is empty (or evaluates to False), the function 
        returns np.nan. If no postal code is found at the end of the address, the function returns 
        (None, address).

        Parameters:
            address (str): A string representing an address that may end with a postal code.

        Returns:
            tuple:
                - postal_code (str or None): The extracted postal code if found, otherwise None.
                - address (str or np.nan): The remaining address with the postal code removed, or np.nan 
                if the address becomes empty, or np.nan if the original address is empty.
        """
        if not address:
            return ''
        else:
            address = str(address)
            match = re.search(r'\b\d{5}(?:-\d{4})?\b$', address)
            if match:
                # Extract postal code
                postal_code = match.group(0)
                if '-' in str(postal_code):
                    postal_code = postal_code.split('-')[0]
                # Remove postal code from address
                address = re.sub(r'\b\d{5}(?:-\d{4})?\b$', '', address).strip()
                if not address:
                    address = ''
                return postal_code, address
        return '', address


    # Apply the extract_postal_code function to each value in the Address column.
    # It returns a tuple (postal_code, remaining_address).
    # We then convert the tuple into a Series to assign to new columns.
    df_contacts_c6[['Postal_Zip_Code', 'Address1']] = df_contacts_c6['Address'].apply(
        lambda addr: pd.Series(extract_postal_code(addr))
    )

    # 2. Extract State
    us_state_codes = dmr.us_state_codes

    state_pattern = r",\s*([A-Z]{2})$"
    df_contacts_c6['Prov_State'] = df_contacts_c6['Address1'].str.extract(state_pattern, expand=False)
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.replace(state_pattern, "", regex=True)
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.strip()

    # The pattern ^\s*na\s*$ ensures that the entire string is just 'na' with possible surrounding whitespace.
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.replace(r'(?i)^\s*na\s*$', '', regex=True)

    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.replace('unknown', '', case=False)
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.replace('unkown', '', case=False)
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.replace(r'(?i)\buknown\b', '', regex=True, case=False)
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.replace(r'(?i)\bukn\b', '', regex=True, case=False)
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.replace(r'(?i)\bunk\b', '', regex=True, case=False)
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.replace(r'(?i)\bna\b', '', regex=True)
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.replace(r'(?i)xxx+', '', regex=True, case=False)
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.replace('not available', '', case=False)
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.replace(
        r'(?i)\bn\\?a\b', '', regex=True).str.strip()
    df_contacts_c6.loc[
        ~df_contacts_c6['Address1'].str.contains(r'[A-Za-z]', na=False), 'Address1'
    ] = ""
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].str.strip()
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    def remove_single_letter_if_only(address):
        """
        Removes the entire address if, after removing non-letter characters, only a single letter remains.
        Otherwise, returns the original address.
        """
        if not address or address == '':
            return ''
        else:
            # Remove non-letter characters.
            letters_only = re.sub(r'[^a-zA-Z]', '', str(address))
            # If after cleanup there's exactly one letter, return an empty string.
            if len(letters_only) == 1:
                return ""
            return address

    # Apply the function to the Address1 column.
    df_contacts_c6['Address1'] = df_contacts_c6['Address1'].apply(remove_single_letter_if_only)

    def extract_city_county(row):
        """
        Uses Postal_Zip_Code to narrow the set of possible city names from df_zip_to_county,
        applies fuzzy matching on Address1 to identify if one of the candidate cities is present,
        and if so extracts the official city name and county. It then returns the city, county,
        and a new Address1 with the matched city name removed.
        """
        address1 = row['Address1']
        postal_zip = row['Postal_Zip_Code']
        
        # If there is no zip code, we do not change the address.
        if pd.isna(postal_zip) or postal_zip.strip() == "":
            return pd.Series(['', '', address1])
        
        # Here we assume exact match on zip codes between the address and the reference table.
        # You may wish to adjust for leading zeros or different formats if needed.
        matches_zip = df_zip_to_county[df_zip_to_county['zip'] == postal_zip.strip()]
        
        # If no matching zip is found, leave Address1 as is.
        if matches_zip.empty:
            return pd.Series(['', '', address1])
        
        # Prepare to compare text in lower case. Note: zip codes are typically numeric so we may not need to alter them.
        address1_lower = str(address1).lower()
        
        best_match = None  # will hold the candidate city with best score
        best_score = 0
        best_row = None  # will eventually hold the row from df_zip_to_county corresponding to best match
        
        # For each matching zip code record, consider the primary city and its acceptable alternatives.
        for _, zip_row in matches_zip.iterrows():
            # Build the list of candidate city names from the reference.
            candidates = [zip_row['primary_city']]
            # Split acceptable_cities by comma and strip any whitespace.
            if pd.notna(zip_row['acceptable_cities']):
                candidates.extend([c.strip() for c in zip_row['acceptable_cities'].split(',') if c.strip() != ""])
            
            # Iterate over the candidate cities and look for a fuzzy match in Address1.
            for city in candidates:
                # Compute the fuzzy match score between the candidate city (in lower case) and the Address1 text.
                score = fuzz.partial_ratio(str(city).lower(), address1_lower)
                if score > best_score and score >= 90:  # Only consider matches with score over or equal to 90
                    best_score = score
                    best_match = city  # store the official city name from the reference
                    best_row = zip_row
                    
        # If no candidate is found, leave the columns empty.
        if best_match is None:
            return pd.Series(['', '', address1])
        
        # Remove the found city substring from Address1.
        # We use re.compile with re.IGNORECASE to make a case-insensitive substitution.
        pattern = re.compile(re.escape(best_match), re.IGNORECASE)
        # Remove the matched city from Address1. Also, remove any leftover trailing commas.
        new_address1 = pattern.sub('', address1).strip().rstrip(',')
        
        # Return the official city name, county (from best_row), and cleaned Address1.
        return pd.Series([best_match, best_row['county'], new_address1])

    # Apply the function row-wise and assign results to new columns.
    df_contacts_c6[['City', 'County', 'Address1']] = \
        df_contacts_c6.apply(extract_city_county, axis=1)

    # Using existing good addresses to complete ones with small errors
    # Filter the rows where Address1 is not empty/NaN.
    mask_valid_address = df_contacts_c6['Address1'].notna() & (df_contacts_c6['Address1'] != "")

    # Filter out rows where Postal_Zip_Code is missing (empty string or NaN).
    mask_missing_zip = df_contacts_c6['Postal_Zip_Code'].isna() | (df_contacts_c6['Postal_Zip_Code'] == "")

    # Sub-dataframe for rows missing zipcode and county information.
    df_missing = df_contacts_c6[mask_valid_address & mask_missing_zip]

    # Sub-dataframe for rows that have a Postal_Zip_Code.
    df_with_zip = df_contacts_c6[mask_valid_address & ~(mask_missing_zip)]

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    # Loop over each row that is missing a zip code.
    for idx, missing_row in df_missing.iterrows():
        best_score = 0
        best_zip = None
        best_county = None
        best_city = None
        best_state = None

        # Compare the current Address1 with all Address1 values that have a zip code.
        for jdx, zip_row in df_with_zip.iterrows():
            # Use case-insensitive comparison by converting both strings to lower case.
            score = fuzz.partial_ratio(str(missing_row['Address']).lower(), zip_row['Address'].lower())
            # If this candidate is better than previous ones, remember its details.
            if score > best_score:
                best_score = score
                best_zip = zip_row['Postal_Zip_Code']
                best_county = zip_row['County']
                best_city = zip_row['City']
                best_state = zip_row['Prov_State']

        # If the best match score meets the threshold, update the missing row.
        if best_score >= 90:
            df_contacts_c6.loc[idx, 'Postal_Zip_Code'] = best_zip
            df_contacts_c6.loc[idx, 'County'] = best_county
            df_contacts_c6.loc[idx, 'City'] = best_city
            df_contacts_c6.loc[idx, 'Prov_State'] = best_state

    df_contacts_c6['County'] = df_contacts_c6['County'].str.replace('County', '', 
                                                                    case=False)

    # Spliting up names
    df_contacts_c7 = df_contacts_c6.copy()
    org_names = ['aetna', 'aarp', 'medicare', 'health', 'healthcare', 'hospice',
                'trust', 'blue cross', 'blue shield', 'care', 'hospital',
                'memorial', 'group', 'foundation', 'united', 'ppo', 'assisted',
                'living', 'services', 'agency', 'support', 'center', 'institute',
                'organization', 'home', 'facility', 'wellness', 'rehab', 
                'partnership', 'network', 'nursing', 'signa']

    def parse_name(full_name):
        """
        Splits a full name into Last_Name, First_Name, and Middle_Name components.
        
        Special Handling:
        - If any substring from org_names exists in the name (case-insensitive),
            then returns the entire string (title cased) as First_Name, with Last_Name and Middle_Name empty.
        - If the name contains a comma, assumes the format is:
            last_name, first_name middle_name
        - Otherwise, assumes the regular order:
            first_name middle_name last_name
        - If only one token is present, it is assumed to be First_Name.
        
        Returns:
        pandas.Series with keys: 'Last_Name', 'First_Name', 'Middle_Name'
        """
        # Handle missing values gracefully.
        if pd.isnull(full_name) or str(full_name).strip() == "":
            return pd.Series({"Last_Name": "", "First_Name": "", "Middle_Name": ""})
        
        # Ensure full_name is a string.
        full_name = str(full_name).strip()
        lower_name = str(full_name).lower()
        
        # Check for organization name tokens.
        if any(org in lower_name for org in org_names):
            # Return the whole string as First_Name (title-cased).
            return pd.Series({"Last_Name": "", "First_Name": full_name.title(), "Middle_Name": ""})
        
        # Check if the name contains a comma.
        if ',' in full_name:
            # Format assumed: last_name, first_name middle_name
            parts = full_name.split(',', 1)
            last = parts[0].strip()
            remaining = parts[1].strip() if len(parts) > 1 else ""
            remaining_parts = remaining.split()
            if len(remaining_parts) == 0:
                first = ""
                middle = ""
            elif len(remaining_parts) == 1:
                first = remaining_parts[0]
                middle = ""
            else:
                first = remaining_parts[0]
                # All tokens after the first are considered the middle name.
                middle = " ".join(remaining_parts[1:])
        else:
            # Format assumed: first_name middle_name last_name.
            parts = full_name.split()
            if len(parts) == 0:
                first = ""
                middle = ""
                last = ""
            elif len(parts) == 1:
                first = parts[0]
                middle = ""
                last = ""
            elif len(parts) == 2:
                first = parts[0]
                last = parts[1]
                middle = ""
            else:
                first = parts[0]
                last = parts[-1]
                middle = " ".join(parts[1:-1])
        
        # Capitalize each component using title case.
        first = first.title()
        middle = middle.title()
        last = last.title()
        
        return pd.Series({"Last_Name": last, "First_Name": first, "Middle_Name": middle})

    df_contacts_c7[['Last_Name', 'First_Name', 'Middle_Name']] = \
        df_contacts_c7['Name'].apply(parse_name)
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    def extract_contact_info(text):
        """
        Extract phone numbers and email from a text string.
        
        The function finds the following labels (case-insensitive):
        - Primary
        - Cell
        - Home
        - Work
        - Email
        
        It applies the following logic:
        - If both Home and Primary exist:
                Phone_Home = Home number,
                Phone_Cell = Primary number.
        - Else if Primary and Cell exist:
                Phone_Home = Primary number,
                Phone_Cell = Cell number.
        - Else if only Primary exists:
                Phone_Home = Primary number.
        - Also, assign Work to Phone_Office and Email to Email.
        
        Returns a pandas Series with keys:
            'Email', 'Phone_Office', 'Phone_Cell', 'Phone_Home'
        """
        # Initialize a dictionary to hold the extracted values.
        parsed = {}
        
        # Regex pattern:
        # This pattern looks for one of the labels followed (optionally after some whitespace)
        # by either a phone number in the format (NNN) NNN-NNNN or by an email address.
        pattern = re.compile(
            r'(?i)(Primary|Cell|Home|Work|Email)\s*'
            r'(\([0-9]{3}\)\s*[0-9]{3}-[0-9]{4}|[\w\.\-]+@[\w\.\-]+\.\w+)'  # phone or email pattern
        )
        
        # Find all matches (each match is a tuple (label, value)).
        matches = pattern.findall(text)
        
        for label, value in matches:
            # Save the value in a dictionary using lowercase for the key.
            parsed[str(label).lower()] = value.strip()
        
        # Initialize new columns.
        phone_office = parsed.get('work', None)
        email = parsed.get('email', None)
        phone_cell = None
        phone_home = None
        
        # Decide how to assign Primary, Cell, Home per the rules:
        
        # Rule 1: If both Home and Primary exist:
        if 'home' in parsed and 'primary' in parsed:
            phone_home = parsed['home']
            phone_cell = parsed['primary']
        # Rule 2: If no Home but both Primary and Cell exist:
        elif 'primary' in parsed and 'cell' in parsed:
            phone_home = parsed['primary']
            phone_cell = parsed['cell']
        # Rule 3: If only Primary exists:
        elif 'primary' in parsed:
            phone_home = parsed['primary']
        # Otherwise, if only one of Home or Cell exists, assign it.
        elif 'home' in parsed:
            phone_home = parsed['home']
        elif 'cell' in parsed:
            phone_cell = parsed['cell']
        
        return pd.Series({
            'Email': email,
            'Phone_Office': phone_office,
            'Phone_Cell': phone_cell,
            'Phone_Home': phone_home
        })

    # Apply the function to the 'Phone/Email' column and create the new columns.
    df_contacts_c7[['Email', 'Phone_Office', 'Phone_Cell', 'Phone_Home']] = \
        df_contacts_c7['Phone/Email'].apply(extract_contact_info)

    # Assigning Contacts Type
    # Create a new column with default values
    df_contacts_c7['ContactType01'] = None

    # Initialize variables
    current_client_id = None
    contact_count = 0
    # Iterate through the DataFrame
    for index, row in df_contacts_c7.iterrows():
        client_id = row['Client_ID_Number']

        # Check if the client ID has changed
        if client_id != current_client_id:
            # Reset the counter and update the current client ID
            contact_count = 1
            current_client_id = client_id
        else:
            # Increment the counter for the same client ID
            contact_count += 1

        # Assign the value to ContactType01
        df_contacts_c7.at[index, 'ContactType01'] = \
            f"Emergency Contact # {contact_count}"

    # Initialize ContactType columns
    for i in range(2, 6):
        df_contacts_c7[f'ContactType0{i}'] = None

    responsible_party_terms = ['responsible party', 'receive a/r', 'receive a/r statement',
                            'primary financial contact', 'legal guardian', 
                            'guarantor']

    # Populate ContactType02 based on Responsibilities
    df_contacts_c7['ContactType02'] = df_contacts_c7['Responsibilities'].str.lower().apply(
        lambda x: 'Responsible Party' if any(term in x for term in responsible_party_terms) else None)

    # Ensure each MRN group has at least one "Responsible Party"
    for mrn, group in df_contacts_c7.groupby('Client_ID_Number'):
        if not group['ContactType02'].eq('Responsible Party').any():
            idx = group.index.min()
            df_contacts_c7.at[idx, 'ContactType02'] = 'Responsible Party'
            
    # Term lists for various POA and guardian types.
    poaf_terms = ['power of attorney/financial', 'power of attorney financial', 'poa financial']
    poac_terms = ['power of attorney/health care', 'power of attorney/healthcare',
                'power of attorney/health', 'power of attorney health', 'poa health', 'poa care']
    poa_terms = ['power of attorney', 'poa']
    guardian_terms = ['guardian', 'legal guardian']

    # List of ContactType columns to populate
    contact_type_columns = [f'ContactType0{i}' for i in range(2, 6)]

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    def populate_contact_types(row):
        responsibilities_lower = str(row['Responsibilities']).lower()
        
        # Flags to track whether a type was assigned.
        poaf_assigned = False
        poac_assigned = False
        guardian_assigned = False

        # Determine if explicit keywords for POA Financial or Care were found.
        explicit_poaf = any(term in responsibilities_lower for term in poaf_terms)
        explicit_poac = any(term in responsibilities_lower for term in poac_terms)
        
        # --- Specific Assignments ---

        # Specific scan for POA - Financial terms
        if explicit_poaf:
            for col in contact_type_columns:
                if pd.isna(row[col]) or row[col] is None:
                    row[col] = 'POA - Financial'
                    poaf_assigned = True
                    break

        # Specific scan for POA - Care terms
        if explicit_poac:
            for col in contact_type_columns:
                if pd.isna(row[col]) or row[col] is None:
                    row[col] = 'POA - Care'
                    poac_assigned = True
                    break

        # Scan for Guardian terms
        if not guardian_assigned:
            for term in guardian_terms:
                if term in responsibilities_lower:
                    for col in contact_type_columns:
                        if pd.isna(row[col]) or row[col] is None:
                            row[col] = 'Guardian'
                            guardian_assigned = True
                            break
                    break

        # --- General POA Handling ---
        # If the responsibilities mention "power of attorney" (or "poa")
        # and there was no explicit indicator (i.e. neither explicit_poaf nor explicit_poac is True),
        # then assign both POA - Financial and POA - Care.
        if (("power of attorney" in responsibilities_lower or "poa" in responsibilities_lower)
            and not (explicit_poaf or explicit_poac)):
            # Assign POA - Financial if not assigned.
            if not poaf_assigned:
                for col in contact_type_columns:
                    if pd.isna(row[col]) or row[col] is None:
                        row[col] = 'POA - Financial'
                        poaf_assigned = True
                        break
            # Assign POA - Care if not assigned.
            if not poac_assigned:
                for col in contact_type_columns:
                    if pd.isna(row[col]) or row[col] is None:
                        row[col] = 'POA - Care'
                        poac_assigned = True
                        break

        return row

    # Apply the function on each row of the DataFrame.
    df_contacts_c7 = df_contacts_c7.apply(populate_contact_types, axis=1)

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    df_contacts_c7['Facility_Id'] = facility_code
    df_contacts_c7['Phone_Office_Ext'] = ''
    df_contacts_c7['Address2'] = ''
    df_contacts_c7['Country'] = 'United States'

    columns_to_keep = ['Facility_Id', 'Client_ID_Number', 'Last_Name', 'First_Name',
                    'Middle_Name', 'Address1', 'Address2', 'County', 
                    'Postal_Zip_Code', 'City', 'Prov_State', 'Country', 'Email',
                    'Phone_Office', 'Phone_Office_Ext', 'Phone_Cell', 'Phone_Home',
                    'Relation_Code', 'ContactType01', 'ContactType02', 
                    'ContactType03', 'ContactType04', 'ContactType05', 'Resident']

    df_contacts = df_contacts_c7[columns_to_keep].copy()

    df_contacts['Last_Name'] = df_contacts['Last_Name'].apply(dmr.name_proper)
    df_contacts['First_Name'] = df_contacts['First_Name'].apply(dmr.name_proper)
    df_contacts['Middle_Name'] = df_contacts['Middle_Name'].apply(dmr.name_proper)
    df_contacts['City'] = df_contacts['City'].apply(dmr.name_proper)
    df_contacts['Address1'] = df_contacts['Address1'].apply(dmr.address_proper)

    log_fn(f"✅ {df_contacts['Client_ID_Number'].nunique()} unique residents processed.")

    # Write to an Excel Template
    template_path = get_template_path("CLIENT_CONTACT.xlsx")
    #template_path = resource_path("resources", "templates", "CLIENT_CONTACT.xlsx")

    if not template_path.exists():
        error_message = f"Template file not found at {template_path}"
        log_fn("⚠️ " + error_message)
        raise FileNotFoundError(error_message)

    output_file_path.parent.mkdir(parents=True, exist_ok=True)

    # Copy the template file to the output_file_path (this creates the new file based on the template)
    shutil.copy(template_path, output_file_path)

    # Load the workbook from the newly created file
    wb = load_workbook(output_file_path)

    # Select the "Data" sheet; ensure the sheet name matches exactly
    ws = wb['Data']

    # Assume that the first row in the template is the header.
    # We will write the data starting at row 2.
    start_row = 2

    # Get the dataframe containing your final contacts data.
    # Make sure that the dataframe's columns match the header in the template.
    for r_idx, row in enumerate(df_contacts.itertuples(index=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for dv in ws.data_validations.dataValidation:
        new_ranges = []
        for rng in dv.ranges:
            # ensure we pass a plain "A2:F100" string, not the CellRange object
            rng_str = str(rng)
            min_col, min_row, max_col, max_row = range_boundaries(rng_str)
            new_ranges.append(
                f"{ws.cell(row=min_row, column=min_col).coordinate}:"
                f"{ws.cell(row=ws.max_row, column=max_col).coordinate}"
            )
        dv.ranges = new_ranges
    # Save the workbook to preserve changes
    wb.save(output_file_path)
    wb.close()
