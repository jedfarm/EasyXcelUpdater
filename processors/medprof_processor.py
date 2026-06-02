#VERSION: 1.0.4

import os
import pandas as pd
from pathlib import Path
import numpy as np
import sys
import re
import requests
from datetime import datetime
from functools import lru_cache
import string
import threading
from fuzzywuzzy import fuzz
import shutil
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from requests.exceptions import RequestException
from utils import data_migration_resources as dmr
from utils.exceptions import AbortedByUser
from utils.file_utils import resource_path
from utils.file_utils import get_template_path

root_abs_path = Path(__file__).resolve().parent.parent.parent

def process_medprof_file(
        facility_code,
        facility_name,
        facility_type,
        input_file_path,
        output_file_path,
        log_fn=None,
        abort_event=threading.Event()
):
    if log_fn is None:
       log_fn = lambda msg: None

    if abort_event.is_set():
       raise AbortedByUser("Process aborted by user.")
    
    if not isinstance(output_file_path, Path):
        output_file_path = Path(output_file_path)

    # Read the medical professionals file
    try:
        df_medprof_raw = pd.read_excel(input_file_path, header=None)
        # Rename columns to Excel letters: A, B, C, …
        df_medprof_raw.columns = [chr(65 + i) for i in range(df_medprof_raw.shape[1])]
        log_fn("✅ Medprof file successfully loaded")
    except Exception as e:
        error_message = f"Error reading medprof file: {str(e)}"
        log_fn(error_message)
        raise Exception(error_message)

    try:
        zip_to_county_db = resource_path(
            "resources", "databases", "zip_code_database.xlsx")
        df_zip_to_city = pd.read_excel(zip_to_county_db,
                                       usecols=['zip', 'primary_city',
                                                'acceptable_cities',
                                                'state'])
        df_zip_to_city['zip'] = \
            df_zip_to_city['zip'].apply(lambda x: f"{int(x):05d}"
                                        if pd.notnull(x) else x)
    except Exception as e:
        error_message = f"Error reading zip code database: {str(e)}"
        log_fn(error_message)
        raise Exception(error_message) 
      
    country_id = 100

    # Creating a new column related to physician names
    df_medprof_c1 = df_medprof_raw.copy()
    # mask = True for rows *i* where row i+1's A starts with 'Address:'
    mask_next_address = (
        df_medprof_c1['A']
        .astype(str)
        .shift(-1)                           # look one row *down*
        .str.startswith('Address:', na=False)
    )

    # vectorized assignment: if True, copy df['A'], else None
    df_medprof_c1['Physician'] = np.where(
        mask_next_address,
        df_medprof_c1['A'],
        None
    )

    df_medprof_c1['Physician'] = df_medprof_c1['Physician'].ffill()

    # Removing initial rows that do not contain any relevant data
    df_medprof_c1 = df_medprof_c1.dropna(subset=['Physician'])
    df_medprof_c1 = df_medprof_c1.reset_index(drop=True)
    
    if df_medprof_c1.empty:
        log_fn("⚠️ No data found in the uploaded file.")
        raise ValueError("No data found in the uploaded file.")

    target_fields = {
        'Address:':'Address1', 
        'DEA No:': 'Other_ID', 
        'License No:':'State_License_Number', 
        'Primary Phone:': 'Phone_Office',
        'Medicaid No:': 'Medicaid_Provider_No',
        'Pager:':'Phone_Pager',
        'Medicare No:': 'Medicare_Provider_No',
        'Fax:': 'Fax',
        'Tax ID No:': 'Taxonomy_Code',
        'System Email:':'Email_Address',
        'UPIN No:': 'Medicare_Provider_No',
        'National Provider ID:':'National_Provider_ID'
    }

    df = df_medprof_c1.copy()
    letter_cols = [c for c in df.columns if re.fullmatch(r'[A-Z]+', c)]
    physicians_list = []

    for phys_name, group in df.groupby('Physician', sort=False):
        # init fields as empty strings
        physician = { v: '' for v in target_fields.values() }
        physician['Address2'] = ''
        physician['Name']     = str(phys_name)

        # make a list of this group’s original df indices, in order
        group_idx = list(group.index)

        for idx, row in group.iterrows():
            for col in letter_cols:
                cell = row[col]
                if not isinstance(cell, str):
                    continue

                for key, field_name in target_fields.items():
                    if not cell.startswith(key):
                        continue

                    # skip UPIN if Medicare already filled
                    if key == 'UPIN No:' and physician['Medicare_Provider_No']:
                        continue

                    extra = cell[len(key):].strip()
                    if extra:
                        physician[field_name] = extra
                    else:
                        # fall back to the right-hand column
                        i = letter_cols.index(col)
                        if i+1 < len(letter_cols):
                            neigh = row[letter_cols[i+1]]
                            if pd.notna(neigh):
                                physician[field_name] = str(neigh)

                    # ─── Address2: scan down until 'A' is non-blank ───
                    if key == 'Address:' and physician['Address1']:
                        i = letter_cols.index(col)
                        if i+1 < len(letter_cols):
                            next_col = letter_cols[i+1]
                            # find our position in group_idx
                            pos = group_idx.index(idx)
                            # scan subsequent rows
                            for j in group_idx[pos+1:]:
                                # stop if column A becomes non-empty
                                if pd.notna(df.at[j, 'A']):
                                    break
                                val = df.at[j, next_col]
                                if isinstance(val, str) and val.strip():
                                    physician['Address2'] = val.strip()
                                    break

                    break  # done with this cell → next column

        physicians_list.append(physician)

    df_medprof_c2 = pd.DataFrame(physicians_list).astype("string")
    if df_medprof_c2.empty:
        log_fn("⚠️ No valid medical professionals found in the file.")
        raise ValueError("No valid medical professionals found in the file.")   
    
    log_fn(f"Collecting data ... there are roughly: {df_medprof_c2.shape[0]} professionals")


    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    # Working with addresses
    # 1. Zip Code
    df_medprof_c2['Postal_Zip_Code'] = ''

    postal_and_addr = df_medprof_c2['Address2'] \
        .apply(dmr.extract_postal_code) \
        .tolist()

    # make a DataFrame from the list of (zip, addr) tuples
    postal_df = pd.DataFrame(
        postal_and_addr,
        columns=['Postal_Zip_Code','Address2'],
        index=df_medprof_c2.index
    )

    # assign back
    df_medprof_c2[['Postal_Zip_Code','Address2']] = postal_df


    VALID_REPEATING_ZIPS = {"22222", "44444", "55555", "88888"}

    def clean_zip(z):
        """
        Normalize a zip code field:
        - NaN → ""
        - ""  → ""
        - "12345-6789", "123456789", "12345 6789" → "12345"
        - Remove second part of ZIP
        - If the 5-digit code consists of the same digit:
            - Keep only if it's in VALID_REPEATING_ZIPS
            - Otherwise → ""
        - Anything else (too short / no digits) → ""
        """
        # 1) NaN → empty
        if pd.isna(z):
            return ""
        s = str(z).strip()
        # 2) leave existing blank
        if s == "":
            return ""
        # 3) regex: capture leading five digits
        m = re.match(r"^(\d{5})", s)
        if not m:
            return ""
        zip5 = m.group(1)
        # 4) check for repeating digits invalid codes
        if len(set(zip5)) == 1 and zip5 not in VALID_REPEATING_ZIPS:
            return ""
        return zip5

    df_medprof_c2["Postal_Zip_Code"] = df_medprof_c2["Postal_Zip_Code"].apply(clean_zip)

    #2. State
    us_state_codes = dmr.us_state_codes

    state_pattern = r",\s*([A-Z]{2})$"
    codes = df_medprof_c2['Address2'].str.extract(state_pattern, expand=False)
    valid = set(us_state_codes.keys())
    df_medprof_c2['Prov_State'] = codes.where(codes.isin(valid), '')

    mask = df_medprof_c2['Prov_State'] != ''
    df_medprof_c2.loc[mask, 'Address2'] = (
        df_medprof_c2.loc[mask, 'Address2']
        .str.replace(state_pattern, '', regex=True)
        .str.strip()
    )
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    # 3. City

    def extract_city(row):
                """
                Uses Postal_Zip_Code to narrow the set of possible city names from df_zip_to_county,
                applies fuzzy matching on Address1 to identify if one of the candidate cities is present,
                and if so extracts the official city name and county. It then returns the city, county,
                and a new Address1 with the matched city name removed.
                """
                address1 = row['Address2']
                postal_zip = row['Postal_Zip_Code']
                
                # If there is no zip code, we do not change the address.
                if pd.isna(postal_zip) or postal_zip.strip() == "":
                    return pd.Series(['', address1])
                
                # Here we assume exact match on zip codes between the address and the reference table.
                # You may wish to adjust for leading zeros or different formats if needed.
                matches_zip = df_zip_to_city[df_zip_to_city['zip'] == postal_zip.strip()]
                
                # If no matching zip is found, leave Address1 as is.
                if matches_zip.empty:
                    return pd.Series(['', address1])
                
                # Prepare to compare text in lower case. Note: zip codes are typically numeric so we may not need to alter them.
                address1_lower = str(address1).lower()
                
                best_match = None  # will hold the candidate city with best score
                best_score = 0
                best_row = None  # will eventually hold the row from df_zip_to_county corresponding to best match
                
                # For each matching zip code record, consider the primary city and its acceptable alternatives.
                for _, zip_row in matches_zip.iterrows():
                    # Build the list of candidate city names from the reference.
                    candidates = [zip_row['primary_city']]
                    # Split acceptable_cities by comma and strip any whitespace.
                    if pd.notna(zip_row['acceptable_cities']):
                        candidates.extend([c.strip() for c in zip_row['acceptable_cities'].split(',') if c.strip() != ""])
                    
                    # Iterate over the candidate cities and look for a fuzzy match in Address1.
                    for city in candidates:
                        # Compute the fuzzy match score between the candidate city (in lower case) and the Address1 text.
                        score = fuzz.partial_ratio(str(city).lower(), address1_lower)
                        if score > best_score and score >= 90:  # Only consider matches with score over or equal to 90
                            best_score = score
                            best_match = city  # store the official city name from the reference
                            best_row = zip_row
                            
                # If no candidate is found, leave the columns empty.
                if best_match is None:
                    return pd.Series(['', address1])
                
                # Remove the found city substring from Address1.
                # We use re.compile with re.IGNORECASE to make a case-insensitive substitution.
                pattern = re.compile(re.escape(best_match), re.IGNORECASE)
                # Remove the matched city from Address1. Also, remove any leftover trailing commas.
                new_address1 = pattern.sub('', address1).strip().rstrip(',')
                
                # Return the official city name, county (from best_row), and cleaned Address1.
                return pd.Series([best_match, new_address1])

    df_medprof_c2[['City', 'Address2']] = \
                df_medprof_c2.apply(extract_city, axis=1)

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    # Putting the phone numbers in the right format
    def contain_phone(s):
            output = ''
            if s:
                s = str(s).strip()
                digits = [c for c in s if c.isdigit()]
                if len(digits) == 10:
                    phone_number = ''.join(digits)
                    if phone_number != phone_number[0] * 10:
                        output = f'({phone_number[:3]}) {phone_number[3:6]}-{phone_number[6:]}'
            return output

    df_medprof_c2['Phone_Office'] = df_medprof_c2['Phone_Office'].apply(contain_phone)
    df_medprof_c2['Phone_Pager'] = df_medprof_c2['Phone_Pager'].apply(contain_phone)
    df_medprof_c2['Fax'] = df_medprof_c2['Fax'].apply(contain_phone)

    # Splitting up names

    # 1) apply name_splitter to each Name and expand into a DataFrame
    name_parts = df_medprof_c2['Name'].apply(dmr.name_splitter).apply(pd.Series)

    # 2) pull out just the three columns you need
    df_medprof_c2[['First_Name','Middle_Name','Last_Name']] = \
        name_parts[['First_Name','Middle_Name','Last_Name']]

    df_medprof_c2 = df_medprof_c2.fillna('')
    df_medprof_c2['Facility_Id'] = facility_code
    df_medprof_c2['Staff_ID_Number'] = ''
    df_medprof_c2['Title'] = ''
    df_medprof_c2['Address3'] = ''
    df_medprof_c2['Country_ID'] = country_id
    df_medprof_c2['Phone_Office_Ext'] = ''
    df_medprof_c2['Phone_Cell'] = ''
    df_medprof_c2['Phone_Other'] = ''
    df_medprof_c2['Phone_Home'] = ''
    df_medprof_c2['Medical_Professional_Type'] = ''
    df_medprof_c2['Taxonomy_Code'] = ''
    df_medprof_c2['EIN'] = ''
    df_medprof_c2['Credentialed'] = ''
    df_medprof_c2['Sanctioned'] = ''
    df_medprof_c2['Comments'] = ''
    df_medprof_c2['Login_Name'] = ''

    ordered_columns = ['Facility_Id', 'Staff_ID_Number',
    'First_Name','Last_Name','Middle_Name', 'Medicaid_Provider_No', 'Medicare_Provider_No', 
    'Address1', 'Address2', 'Address3', 'City', 'Prov_State', 'Country_ID', 'Postal_Zip_Code', 
    'Phone_Office', 'Phone_Office_Ext', 'Phone_Cell','Phone_Pager', 'Phone_Other', 
    'Phone_Home', 'Fax','Email_Address', 'Medical_Professional_Type', 'Taxonomy_Code',
    'State_License_Number', 'EIN', 'Other_ID', 'National_Provider_ID', 'Credentialed',
    'Sanctioned', 'Comments', 'Login_Name']

    df_medprof_c3 = df_medprof_c2[ordered_columns].copy()

    # Data Integrity check on NPI

    def clean_npi_value(value):
        if pd.isna(value):
            return ""

        s = str(value).strip()

        if s == "":
            return ""

        # Fix Excel/pandas float-looking NPIs: 1538197884.0 -> 1538197884
        if re.fullmatch(r"\d+\.0", s):
            s = s[:-2]

        # Remove everything except digits
        digits = re.sub(r"\D", "", s)

        # Keep valid 10-digit NPIs
        if re.fullmatch(r"\d{10}", digits):
            return digits

        # Otherwise return original cleaned digits so we can log it as bad
        return digits

    original_npi = (
        df_medprof_c3["National_Provider_ID"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df_medprof_c3["National_Provider_ID"] = (
        df_medprof_c3["National_Provider_ID"]
        .apply(clean_npi_value)
    )

    valid_npi_mask = (
        df_medprof_c3["National_Provider_ID"].eq("")
        |
        df_medprof_c3["National_Provider_ID"].str.fullmatch(r"\d{10}", na=False)
    )

    bad_npi_rows = df_medprof_c3.loc[~valid_npi_mask].copy()

    if not bad_npi_rows.empty:
        log_fn("⚠️ Some National Provider IDs were not exactly 10 digits and were cleared:")

        for idx, row in bad_npi_rows.iterrows():
            bad_original = original_npi.loc[idx]

            first = str(row.get("First_Name", "")).strip()
            middle = str(row.get("Middle_Name", "")).strip()
            last = str(row.get("Last_Name", "")).strip()

            provider_name = " ".join(
                part for part in [first, middle, last]
                if part
            )

            if not provider_name:
                provider_name = f"Row {idx + 2}"

            log_fn(f"- {provider_name}: {bad_original}")

        df_medprof_c3.loc[~valid_npi_mask, "National_Provider_ID"] = ""


#     df_medprof_c3['National_Provider_ID'] = (
#         df_medprof_c3['National_Provider_ID']
#         .fillna('')
#         .astype(str)
#         .str.strip()
#     )
#
#     # Keep original value for logging
#     original_npi = df_medprof_c3['National_Provider_ID'].copy()
#
#
#
#     # Remove non-digits
#     df_medprof_c3['National_Provider_ID'] = (
#         df_medprof_c3['National_Provider_ID']
#         .str.replace(r'\D', '', regex=True)
#     )
#
#     # Valid if blank or exactly 10 digits
#     valid_npi_mask = (
#         df_medprof_c3['National_Provider_ID'].eq('')
#         |
#         df_medprof_c3['National_Provider_ID'].str.fullmatch(r'\d{10}', na=False)
#     )
#
#     bad_npi_rows = df_medprof_c3.loc[~valid_npi_mask].copy()
#
#     if not bad_npi_rows.empty:
#         log_fn("⚠️ Some National Provider IDs were not exactly 10 digits and were cleared:")
#
#         for idx, row in bad_npi_rows.iterrows():
#             bad_original = original_npi.loc[idx]
#
#             first = str(row.get("First_Name", "")).strip()
#             middle = str(row.get("Middle_Name", "")).strip()
#             last = str(row.get("Last_Name", "")).strip()
#
#             provider_name = " ".join(
#                 part for part in [first, middle, last]
#                 if part
#             )
#
#             if not provider_name:
#                 provider_name = f"Row {idx + 2}"
#
#             log_fn(f"- {provider_name}: {bad_original}")
#
#         df_medprof_c3.loc[~valid_npi_mask, "National_Provider_ID"] = ""
#
#         s = df_medprof_c3['National_Provider_ID'].fillna('').astype(str)
#
#         mask_non_empty = s.str.strip().ne('')                    # not blank
#         mask_not_10 = ~s.str.fullmatch(r'\d{10}')            # not exactly 10 digits
#
#         # 3) Select the offending IDs
#         #non_valid_npi = s[mask_non_empty & mask_not_10].tolist()
#
#         #non_valid_idx = s[mask_non_empty & mask_not_10].index.tolist()


    NON_EMAILS = {
        'na@na.com', 'johndoe@na.com', 'noemail@email.com', 'email@email.com',
        'test@test.com', '111@111.com', 'none@none.com', 'xxx@yyy.zzz',
        'noemail@zzz.com', 'doesnothave@email.com', 'x@x.com'
    }

    # your existing (or tightened) regex
    EMAIL_RE = re.compile(
        r"^([-!#-'*+/-9=?A-Z^-~]+(\.[-!#-'*+/-9=?A-Z^-~]+)*|"
        r"\"([]!#-[^-~ \t]|(\\[\t -~]))+\")@"
        r"([-!#-'*+/-9=?A-Z^-~]+(\.[-!#-'*+/-9=?A-Z^-~]+)*|\[[\t -Z^-~]*\])$"
    )

    def email_checker(s):
        s = (s or '').strip()            # NaN/None → '' and strip spaces
        lower = s.lower()

        # 1) obvious “dummy” list
        if not s or lower in NON_EMAILS:
            return ''

        # 2) detect “only one repeated character” placeholders
        alnum = [c for c in lower if c.isalnum()]
        if alnum and len(set(alnum)) == 1:
            return ''

        # 3) final pattern check
        return s if EMAIL_RE.match(s) else ''

    df_medprof_c3['Email_Address'] = \
        df_medprof_c3['Email_Address'].apply(email_checker)

    df_medprof_c3['Address1'] = \
        df_medprof_c3['Address1'].apply(dmr.address_proper)
    df_medprof_c3['Address2'] = \
        df_medprof_c3['Address2'].apply(dmr.address_proper)
    df_medprof_c3['First_Name'] = \
        df_medprof_c3['First_Name'].apply(dmr.name_proper)
    df_medprof_c3['Last_Name'] = \
        df_medprof_c3['Last_Name'].apply(dmr.name_proper)
    df_medprof_c3['Middle_Name']=\
        df_medprof_c3['Middle_Name'].apply(dmr.name_proper)
    df_medprof_c3['City'] = df_medprof_c3['City'].apply(dmr.name_proper)

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    df_medprof_c4 = df_medprof_c3.drop_duplicates()  
    df_medprof_c4 = df_medprof_c4.reset_index(drop=True)
    df_medprof_c4['First_Name'] = (
        df_medprof_c4['First_Name']
        .fillna('')       # NaN → ''
        .str.upper()
    )
    df_medprof_c4['Last_Name'] = (
        df_medprof_c4['Last_Name']
        .fillna('')       # NaN → ''
        .str.upper()
    )
    df_medprof_c4['Middle_Name'] = (
        df_medprof_c4['Middle_Name']
        .fillna('')       # NaN → ''
        .str.upper()
    )

    df_medprof_c4['National_Provider_ID'] = (
        df_medprof_c4['National_Provider_ID']
        .fillna('')       # NaN → ''
        .str.replace('0000000000', '')
    )

    # ——— 1) Cached API fetcher ———
    @lru_cache(maxsize=1024)
    def fetch_npi_record(npi: str, version: str = "2.1") -> dict:
        try:
            url    = "https://npiregistry.cms.hhs.gov/api/"
            params = {"number": npi, "version": version, "pretty": "off"}
            r      = requests.get(url, params=params, timeout=5)
            r.raise_for_status()
            results = r.json().get("results", [])
            return results[0] if results else {}
        except RequestException as e:
            #log_fn(f"Error fetching NPI record for {npi}: {str(e)}")
            log_fn(f"⚠️ Could not reach NPI registry ({e.__class__.__name__}): {e}")
            return {}


    def get_npi_details(npi: str) -> dict:
        rec   = fetch_npi_record(npi)
        basic = rec.get("basic", {})

        # split out the LOCATION address bits
        loc = next((a for a in rec.get("addresses", [])
                    if a.get("address_purpose") == "LOCATION"), {})
        # taxonomy
        tax = next((t for t in rec.get("taxonomies", [])
                    if t.get("primary") in (True, "true")), None) \
            or (rec.get("taxonomies") or [None])[0]
        # medicaid
        med = next((o for o in rec.get("other_identifiers", [])
                    if o.get("identifier_type","").upper() == "MEDICAID"), {})

        return {
            "first_name":       basic.get("first_name"),
            "middle_name":      basic.get("middle_name"),
            "last_name":        basic.get("last_name"),
            #"credential":       basic.get("credential"),
            "addr_1":           loc.get("address_1"),
            "city":             loc.get("city"),
            "state":            loc.get("state"),
            "postal_code":      loc.get("postal_code"),
            "phone":            loc.get("telephone_number"),
            "fax":              loc.get("fax_number"),
            "taxonomy_code":    tax.get("code")    if tax else None,
            "license_number":   tax.get("license") if tax else None,
            "license_state":    tax.get("state")   if tax else None,
            "medicaid_id":      med.get("identifier")
        }

    # ——— 2) Row‐wise merger function ———
    API_TO_DF = {
        "addr_1":         "Address1",
        "city":           "City",
        "state":          "Prov_State",
        "postal_code":    "Postal_Zip_Code",
        "phone":          "Phone_Office",
        "fax":            "Fax",
        "medicaid_id":    "Medicaid_Provider_No",
        "taxonomy_code":  "Taxonomy_Code",
        "license_number": "State_License_Number",
        "other_identifier": "Other_ID",
        "credential":     "Title" 
        
    }

    def merge_api_data(row, threshold=90):
        npi = str(row["National_Provider_ID"])
        if not npi:
            return row
        record = fetch_npi_record(npi)
        if not record:
            row["National_Provider_ID"] = "INVALID"
            return row

        api = get_npi_details(npi)

        # … (name‐matching logic unchanged) …

        # —— unified API→DF mapping —— 
        for api_key, df_col in API_TO_DF.items():
            new_val = api.get(api_key)
            if new_val and (pd.isna(row[df_col]) or str(row[df_col]).strip() == ""):
                row[df_col] = new_val

        return row

    df_medprof_c5 = df_medprof_c4.apply(merge_api_data, axis=1).copy()

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    def npi_full_lookup(first_name: str, last_name: str, limit: int = 25) -> pd.DataFrame:
        """Returns a DataFrame with practice address, phone, fax, primary taxonomy code,
        license number & state, plus Medicaid ID (if any)."""
        base_url = "https://npiregistry.cms.hhs.gov/api/"
        params = {
            "first_name": first_name,
            "last_name": last_name,
            "version": "2.1",
            "limit": limit,
            "pretty": "off"
        }
        resp = requests.get(base_url, params=params)
        resp.raise_for_status()
        results = resp.json().get("results", [])
        
        rows = []
        for rec in results:

            # 0) Provider credential
            #credential = rec.get("basic", {}).get("credential", "")
            # 1) Practice address & contact
            loc = next((a for a in rec.get("addresses", [])
                        if a.get("address_purpose") == "LOCATION"), {})
            street = " ".join(filter(None, [loc.get("address_1"), loc.get("address_2")]))
            city, st, zipc = loc.get("city"), loc.get("state"), loc.get("postal_code")
            practice_address = f"{street}, {city}, {st} {zipc}"
            phone = loc.get("telephone_number")
            fax   = loc.get("fax_number")
            
            # 2) Primary taxonomy + its license/state
            tax = next((t for t in rec.get("taxonomies", []) if t.get("primary") in (True, "true")), None)
            if not tax:
                tax = rec.get("taxonomies", [{}])[0]
            tax_code      = tax.get("code")
            license_num   = tax.get("license")
            license_state = tax.get("state")
            
            # 3) Medicaid ID (from other_identifiers)
            med = next((oi for oi in rec.get("other_identifiers", [])
                        if oi.get("identifier_type","").upper() == "MEDICAID"), {})
            medicaid_id = med.get("identifier")
            
            # 4) Collect
            rows.append({
                "npi": rec.get("number"),
                "name": f"{rec['basic'].get('first_name')} {rec['basic'].get('last_name')}",
                #"credential": credential,
                "practice_address": practice_address,
                "phone": phone,
                "fax": fax,
                "primary_taxonomy": tax_code,
                "license_number": license_num,
                "license_state": license_state,
                "medicaid_id": medicaid_id
            })
        
        return pd.DataFrame(rows)


    def normalize_phone(p: str) -> str:
        return re.sub(r'\D', '', str(p) or '')

    def normalize_address(a: str) -> str:
        return re.sub(r'[^\w\s]', '', str(a) or '').lower()

    def normalize_other_id(x: str) -> str:
        # strip whitespace, drop any non-alphanumeric, lowercase
        return re.sub(r'\W+', '', str(x) or '').lower()


    def lookup_and_merge_by_name(row):
        # Only run if National_Provider_ID is empty or NaN/blank
        npi = str(row.get("National_Provider_ID") or "").strip()
        if npi:
            return row

        # 1) name‐based lookup
        candidates = npi_full_lookup(row["First_Name"], row["Last_Name"])
        count = len(candidates)
        if count == 0:
            row["National_Provider_ID"]="NOT FOUND"
            return row

        if count == 1:
            rec = candidates.iloc[0]
        else:
            # 2) multiple hits → try to narrow by phone, fax, or address
            cand = candidates.copy()
            # normalize our row values
            rp = normalize_phone(row.get("Phone_Office"))
            rf = normalize_phone(row.get("Fax"))
            ra = normalize_address(row.get("Address1"))
            ro = normalize_other_id(row.get("Other_ID"))
            
            

            cand["phone_norm"] = cand["phone"].apply(normalize_phone)
            cand["fax_norm"]   = cand["fax"].apply(normalize_phone)
            cand["addr_norm"]  = cand["practice_address"].apply(normalize_address)
            cand["other_norm"] = cand["medicaid_id"].apply(normalize_other_id)

            matched = cand[
                (cand["phone_norm"] == rp) |
                (cand["fax_norm"]   == rf) |
                (cand["addr_norm"].str.contains(ra))|
                (cand["other_norm"] == ro)
            ]
            if len(matched) == 1:
                rec = matched.iloc[0]
            else:
                row["National_Provider_ID"] = "MULTIPLE VALUES"
                return row

        # 3) single “rec” → unpack into row
        row["National_Provider_ID"] = rec["npi"]

        # split practice_address → street / city / state+zip
        addr_parts = [p.strip() for p in rec["practice_address"].split(",")]
        if addr_parts:
            row["Address1"] = addr_parts[0]
        if len(addr_parts) > 1:
            row["City"] = addr_parts[1]
        if len(addr_parts) > 2:
            st_zip = addr_parts[2].split()
            if st_zip:
                row["Prov_State"] = st_zip[0]
            if len(st_zip) > 1:
                row["Postal_Zip_Code"] = st_zip[1]

        # fill contact & licensure
        row["Phone_Office"]        = rec["phone"]
        row["Fax"]                 = rec["fax"]
        row["Taxonomy_Code"]       = rec["primary_taxonomy"]
        row["State_License_Number"]= rec["license_number"]
        row["Medicaid_Provider_No"]= rec["medicaid_id"]
        #row["Title"]               = rec["credential"]


        return row

    # then apply it:
    df_medprof_c6 = df_medprof_c5.apply(lookup_and_merge_by_name, axis=1).copy()

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    # Properly formatting data on all fields
    df_medprof_c6['First_Name'] = df_medprof_c6['First_Name'].apply(dmr.name_proper)
    df_medprof_c6['Last_Name'] = df_medprof_c6['Last_Name'].apply(dmr.name_proper)
    df_medprof_c6['Middle_Name'] = df_medprof_c6['Middle_Name'].apply(dmr.name_proper)
    df_medprof_c6['Address1'] = df_medprof_c6['Address1'].apply(dmr.address_proper)
    df_medprof_c6['City'] = df_medprof_c6['City'].apply(dmr.name_proper)
    df_medprof_c6["Postal_Zip_Code"] = df_medprof_c6["Postal_Zip_Code"].apply(clean_zip)
    df_medprof_c6['Prov_State'] = \
        df_medprof_c6['Prov_State'].apply(lambda x: str(x).upper() 
                                        if str(x).upper().strip() in dmr.us_state_codes
                                        else "")
    df_medprof_c6['Phone_Office'] = df_medprof_c6['Phone_Office'].apply(contain_phone)
    df_medprof_c6['Phone_Pager'] = df_medprof_c6['Phone_Pager'].apply(contain_phone)
    df_medprof_c6['Fax'] = df_medprof_c6['Fax'].apply(contain_phone)
    #df_medprof_c6['Title'] = df_medprof_c6['Title'].str.upper().str.replace(".", "")

    # Removing content from Address 2 that already exists in City
    a2 = df_medprof_c6['Address2'].fillna('').str.strip().str.lower()
    city = df_medprof_c6['City'].fillna('').str.strip().str.lower()

    # Build mask where they match
    mask = a2 == city

    # Clear Address2 where it duplicates City
    df_medprof_c6.loc[mask, 'Address2'] = ""

    city_mask = df_medprof_c6['City'].fillna('').str.contains(r'\d')

    # Wherever there's a digit, blank it out
    df_medprof_c6.loc[city_mask, 'City'] = ""

    # Searching for duplicates and removing them
    keys = ['First_Name','Last_Name','Middle_Name']

    # 1) copy and only replace ''→NaN in the *other* columns
    temp = df_medprof_c6.copy()
    non_key_cols = [c for c in temp.columns if c not in keys]
    temp[non_key_cols] = temp[non_key_cols].replace('', np.nan)

    # 2) pull out duplicates
    dup_mask = temp.duplicated(subset=keys, keep=False)
    df_dups = temp.loc[dup_mask]

    # 3) helper
    def first_non_na(x):
        non_na = x.dropna()
        return non_na.iloc[0] if not non_na.empty else np.nan

    # 4) group & merge
    df_non_dup = (
        df_dups
        .groupby(keys, as_index=False)
        .agg(first_non_na)
        .fillna('')      # back to empty‐string convention
    )
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    # 5) Build df_c6_temp by dropping *all* the original duplicate‐rows
    df_c6_temp = df_medprof_c6.drop(df_dups.index)

    # 5) Finally, recombine uniques + merged duplicates into your new df_medprof_c7
    df_medprof_c7 = pd.concat([df_c6_temp, df_non_dup], ignore_index=True)[
        df_medprof_c6.columns  # preserve original column order
    ]

    # 1) Mask only the 10‐digit NPIs
    mask_valid = df_medprof_c6['National_Provider_ID'].str.fullmatch(r'\d{10}', na=False)

    # 2) Among those, find all values that occur more than once
    mask_dup = df_medprof_c6['National_Provider_ID'].duplicated(keep=False)

    # 3) Combined mask
    mask = mask_valid & mask_dup

    # # 4) Apply strikethrough markdown only where needed, avoiding double‐wrapping
    # def strike(s):
    #     s = str(s)
    #     if s.startswith("~~") and s.endswith("~~"):
    #         return s
    #     return f"~~{s}~~"

    # df_medprof_c7.loc[mask, 'National_Provider_ID'] = (
    #     df_medprof_c7.loc[mask, 'National_Provider_ID']
    #     .map(strike)
    # )

    professional_type_mapping = {
       'Advanced Practice Nurse':['367A00000X', '367H00000X','364S00000X',
                    '364SA2100X','364SA2200X','364SC2300X','364SC1501X',
                    '364SC0200X','364SE0003X','364SE1400X','364SF0001X',
                    '364SG0600X','364SH1100X','364SH0200X','364SI0800X',
                    '364SL0600X','364SM0705X','364SN0000X','364SN0800X',
                    '364SX0106X','364SX0200X','364SX0204X','364SP0200X',
                    '364SP1700X','364SP2800X','364SP0808X','364SP0809X',
                    '364SP0807X','364SP0810X','364SP0811X','364SP0812X',
                    '364SP0813X','364SR0400X','364SS0200X','364ST0500X',
                    '364SW0102X','367500000X'],
       'Alternate Physician':['PENDING'],
       'Anesthesiologist':['207L00000X','207LA0401X','207LC0200X','207LH0002X',
                           '207LP2900X', '207LP3000X', '207LP4000X'],
       'Audiologist':['231H00000X', '231HA2400X',  '231HA2500X', '237600000X'],
       'Cardiologist': ['207RA0001X', '207RI0011X', '207RC0000X', '207UN0901X',
                        '2080P0202X'],
       'Chiropractor': ['111N00000X','111NI0013X', '111NI0900X', '111NN0400X', 
                        '111NN1001X', '111NX0100X', '111NX0800X', '111NP0017X',
                        '111NR0200X', '111NR0400X', '111NS0005X', '111NT0100X'],
       'Dental Hygienist':['124Q00000X'],
       'Dentist':['122300000X', '1223D0004X', '1223D0001X', '1223E0200X',
                  '1223G0001X', '1223P0106X','1223X0008X','1223S0112X',
                  '125Q00000X','1223X2210X','1223X0400X','1223P0221X',
                  '1223P0300X', '1223P0700X'],
       'Dermatologist':['207NI0002X', '207ND0900X', '207ND0101X','207NP0225X',
                        '207NS0135X'],
       'Endocrinologist':['207RE0101X', '2080P0205X', '207VE0102X'],
       'Gastroenterologist':['207RG0100X','2080P0206X', ],
       'Geriatrician':['207QG0300X', '207RG0300X'],
       'Medical Director':['PENDING'],
       'Medical Specialist':['207RH0005X', '207RA0201X', '207RH0000X', 
                    '207RI0008X', '207RM1200X', '207RI0200X', '207RP1002X',
                    '207RS0012X', '207RS0010X', '207RA0401X', '207VG0400X',
                    '193200000X','193400000X', '103K00000X', '174400000X',
                    '208000000X', '2081P0301X','2081H0002X','2081N0008X',
                    '2081P2900X','2081P0010X','2081P0004X','2081S0010X',
                    '173000000X'],
       'Medical Student':['390200000X'],
       'Nephrologist':['207RN0300X'],
       'Neurologist':['2084N0400X', '2084N0402X', '2084N0008X', '2081N0008X'],
       'Neuropsychologist':['2251N0400X'],
       'Neurosurgeon':['207T00000X'],
       'Nurse Practitioner':['363L00000X','363LA2100X','363LA2200X','363LC1500X',
                             '363LC0200X','363LF0000X','363LG0600X','363LN0000X',
                             '363LN0005X','363LX0001X','363LX0106X','363LP0200X',
                             '363LP0222X','363LP1700X','363LP2300X','363LP0808X',
                             '363LS0200X','363LW0102X'],
       'Oncologist':['207RH0003X'],
       'Ophthalmologist':['207W00000X','207WX0120X','207WX0009X','207WX0109X',
                          '207WX0200X','207WX0110X','207WX0107X','207WX0108X'],
       'Optometrist':['152W00000X','152WC0802X','152WL0500X','152WX0102X',
                      '152WP0200X','152WS0006X','152WV0400X','156F00000X',
                      '156FC0800X','156FC0801X','156FX1700X','156FX1100X',
                      '156FX1101X','156FX1800X','156FX1201X','156FX1202X',
                      '156FX1900X'],
       'Orthopedist':['207X00000X','207XS0114X','207XX0004X','207XS0106X',
                      '207XS0117X','207XX0801X','207XP3100X','207XX0005X'],
       'Otolaryngologist':['207Y00000X','207YS0123X','207YX0602X','207YX0905X',
                           '207YX0901X','207YP0228X','207YX0007X','207YS0012X'],
       'Physician':['208D00000X', '207P00000X'],
       'Physicians Assistant':['363A00000X', '363AM0700X','363AS0400X'],
       'Podiatrist':['213E00000X','213ES0103X','213ES0131X','213EG0000X',
                     '213EP1101X','213EP0504X','213ER0200X','213ES0000X'],
       'Primary Physician':['207Q00000X', '207QA0505X', '207QH0002X',
                            '207R00000X', '208M00000X', '207RH0002X', 
                            '208100000X'],
       'Psychiatrist':['2084A0401X','2084P0802X','2084B0040X','2084P0301X',
                       '2084P0804X','2084N0600X','2084D0003X','2084E0001X',
                       '2084F0202X','2084P0805X','2084H0002X','2084A2900X',
                       '2084P0005X','2084N0008X','2084B0002X','2084P2900X',
                       '2084P0800X','2084P0015X','2084S0012X','2084S0010X',
                       '2084V0102X'],
       'Psychologist':['103G00000X','103GC0700X', '103T00000X','103TA0400X',
                       '103TA0700X','103TC0700X','103TC2200X','103TB0200X',
                       '103TC1900X','103TE1000X','103TE1100X','103TF0000X',
                       '103TF0200X','103TP2701X','103TH0004X','103TH0100X',
                       '103TM1700X','103TM1800X','103TP0016X','103TP0814X',
                       '103TP2700X','103TR0400X','103TS0200X','103TW0100X'],
       'Pulmonologist':['207RP1001X'],
       'Rheumatologist':['207RR0500X'],
       'Surgeon':['208200000X','2082S0099X','2082S0105X', '208600000X',
                  '2086H0002X','2086S0120X','2086P0122X','2086S0122X',
                  '2086S0105X','2086S0102X','2086X0206X','2086S0127X',
                  '2086S0129X','208G00000X','204F00000X'],
       'Urologist':['208800000X','2088F0040X','2088P0231X'],
       'Wound Specialist':['163WW0000X']
    }

    code_to_prof = {
        code: prof
        for prof, codes in professional_type_mapping.items()
        for code in codes
    }

    # 3) now overwrite in your df_import
    #    only for rows where Taxonomy_Code is non-empty
    mask = (
        df_medprof_c7['Taxonomy_Code'].notna() &
        (df_medprof_c7['Taxonomy_Code'].astype(str).str.strip() != "")
    )

    # map each code → prof, defaulting to "NOT FOUND"
    df_medprof_c7.loc[mask, 'Medical_Professional_Type'] = (
        df_medprof_c7.loc[mask, 'Taxonomy_Code']
                .map(code_to_prof)
                .fillna('PRIMARY PHYSICIAN')
    )

    df_medprof = df_medprof_c7.copy()
    
    # Adjust the path as necessary.
    template_path = get_template_path("MEDICAL_PROFESSIONAL.xlsx")

    if not template_path.exists():
        error_message = f"Template file not found at {template_path}"
        log_fn("⚠️ " + error_message)
        raise FileNotFoundError(error_message)
    output_file_path.parent.mkdir(parents=True, exist_ok=True)
    # Copy the template file to the output_file_path (this creates the new file based on the template)
    shutil.copy(template_path, output_file_path)

    # Load the workbook from the newly created file
    wb = load_workbook(output_file_path)

    # Select the "Data" sheet; ensure the sheet name matches exactly
    ws = wb['Medical Prof Data']

    # Assume that the first row in the template is the header.
    # We will write the data starting at row 2.
    start_row = 2

    # Get the dataframe containing your final allergy data.
    # Make sure that the dataframe's columns match the header in the template.
    for r_idx, row in enumerate(df_medprof.itertuples(index=False),
                                start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for dv in ws.data_validations.dataValidation:
        # dv.ranges is a list of range strings (e.g. ["A2:A100"])
        new_ranges = []
        for rng in dv.ranges:
            rng_str = str(rng)
            # Parse the original range boundaries (start_col, start_row, end_col, end_row)
            min_col, min_row, max_col, max_row = range_boundaries(rng_str)
            # Update the end row to the current maximum row with data (ws.max_row)
            new_range = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=ws.max_row, column=max_col).coordinate}"
            new_ranges.append(new_range)
        # Replace the validation's range list with the updated ranges
        dv.ranges = new_ranges
        # Save the workbook to preserve changes
    wb.save(output_file_path)
    wb.close()
