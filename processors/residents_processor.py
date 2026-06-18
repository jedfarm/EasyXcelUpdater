#VERSION: 1.0.2

import os
import re
import shutil
from pathlib import Path
import pandas as pd
import numpy as np
from fuzzywuzzy import fuzz
import threading
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
#from config.definitions import ROOT_DIR
from utils import data_migration_resources as dmr
from utils.exceptions import AbortedByUser
from utils.file_utils import resource_path
from utils.file_utils import get_template_path

def process_residents_file(
        facility_code,
        facility_name,
        facility_type,
        input_file_path,
        output_file_path,
        log_fn=None,
        abort_event=threading.Event()
):
    if log_fn is None:
       log_fn = lambda msg: None

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    if not isinstance(output_file_path, Path):
        output_file_path = Path(output_file_path)

    # --------------------- DATA LOADING  --------------------
    try:
        df_fs_raw = pd.read_excel(input_file_path, header=None)
        df_fs_raw .columns = [chr(65 + i) for i in range(df_fs_raw .shape[1])]
        log_fn(f"✅ Successfully read the facesheets file: {input_file_path}")
        
        # Create a boolean mask where each row is checked for the condition
        mask = df_fs_raw['A'].str.startswith("Resident Face Sheet:")

        # Check if any row satisfies the condition to avoid errors
        if mask.any():
            # Find the index of the first occurrence
            first_index = mask.idxmax()  # idxmax returns the index of the first True value
                    
            # Slice the DataFrame from the first matching row onward.
            df_fs_raw = df_fs_raw.loc[first_index:]
            
            df_fs_raw = df_fs_raw.copy()

    except Exception as e:
        error_message = f"Error reading facesheets file: {str(e)}"
        log_fn(error_message)
        log_fn("Please ensure the file is in the correct format.")
        raise Exception(error_message)
    
    try:
        zip_to_county_db = resource_path(
            "resources", "databases", "zip_code_database.xlsx")
        df_zip_to_county = pd.read_excel(
            zip_to_county_db,
            usecols=['zip', 'primary_city', 'acceptable_cities', 'county',
                     'state']
        )
        df_zip_to_county['zip'] = df_zip_to_county['zip'].apply(
            lambda x: f"{int(x):05d}" if pd.notnull(x) else x
        )
    except Exception as e:
        error_message = f"Error reading teh zipcode database file: {str(e)}"
        log_fn(error_message)
        raise Exception(error_message)   
    # Make a copy of the DataFrame to work with
    df_fs_c1 = df_fs_raw.copy()

    # Create a boolean mask to select rows starting with "Resident Face Sheet:" while handling NaN values.
    mask = df_fs_raw['A'].fillna('').str.startswith("Resident Face Sheet:")

    def extract_info(text):
        # Extract text between "Resident Face Sheet:" and "MRN:" (non-greedy and allow extra spaces)
        res_matching = re.search(r"Resident Face Sheet:\s*(.*?)\s*MRN:", text, re.DOTALL | re.IGNORECASE)
        if res_matching:
            resident_raw = res_matching.group(1)
            # Remove any text within parentheses (non-greedy)
            resident = re.sub(r"\(.*?\)", "", resident_raw).strip()
        else:
            resident = None

        # Extract the Client ID Number after "MRN:" matching one or more digits with an optional hyphen and digits
        client_matching = re.search(r"MRN:\s*(\d+(?:-\d+)?)", text, re.IGNORECASE)
        if client_matching:
            client_id = client_matching.group(1)
        else:
            client_id = None

        # Return a Series with keys matching the DataFrame columns
        return pd.Series({"Resident": resident, "Client_ID_Number": client_id})

    # Apply the extraction function to the cells that match the mask and assign the output to the respective new columns.
    df_fs_c1.loc[mask, ["Resident", "Client_ID_Number"]] = df_fs_c1.loc[mask, "A"].apply(extract_info)

    df_fs_c1.reset_index(drop=True, inplace=True)
    df_fs_c1['Resident'] = df_fs_c1['Resident'].ffill()
    df_fs_c1['Client_ID_Number'] = df_fs_c1['Client_ID_Number'].ffill()
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    def extract_resident_data(df_fs_c1):
        """
        Extracts resident information from a DataFrame based on specific key locations.

        Args:
            df_fs_c1 (pd.DataFrame): The input DataFrame containing resident data.

        Returns:
            list of dict: One dict per Client_ID_Number with extracted fields.
        """
        if not isinstance(df_fs_c1, pd.DataFrame) or df_fs_c1.empty:
            return []

        # --- Configuration ---
        resident_template = {
            'Client_ID_Number': '', 'Resident': '', 'SSN:': '', 'Medicare A #:': '',
            'Medicaid #:': '', 'Birth Date:': '', 'Sex:': '', 'Religion:': '',
            'Attending:':'', 'Marital Status:': '', 'Preferred Language:': '',
            'Prev Occupation:': '', 'Race:': '', 'Ethnicity:': '', 'Phone:': '',
            'Previous Address:': '', 'Insurance': '', "Insured's ID #": '',
            'Primary Payer:': ''
        }

        h_fields = [
            'SSN:', 'Medicare A #:','Medicaid #:','Birth Date:','Sex:','Attending:',
            'Religion:','Marital Status:','Preferred Language:','Primary Payer:',
            'Prev Occupation:','Race:','Ethnicity:','Phone:'
        ]

        # Find columns labeled by single letters
        single_letter_cols = sorted(
            col for col in df_fs_c1.columns
            if re.match(r'^[A-Z]$', str(col))
        )
        if not single_letter_cols:
            print("Warning: No single-letter columns found.")
            return []

        residents_all = []

        # Group rows by Client_ID_Number
        for client_id, client_df in df_fs_c1.groupby('Client_ID_Number'):
            resident = resident_template.copy()
            resident['Client_ID_Number'] = client_id
            client_df = client_df.reset_index(drop=True)

            # 1) Resident name (first non-null)
            try:
                resident['Resident'] = client_df['Resident'].dropna().iloc[0]
            except:
                resident['Resident'] = ''

            # Prepare holders for discovered values
            found_values = {k: None for k in resident_template}
            potential_insurances   = []
            potential_insured_ids  = []
            address_street         = None
            address_county         = None
            insurance_stop_by_col  = {}

            # Scan every cell
            for r_idx in range(len(client_df)):
                for col in single_letter_cols:
                    cell = client_df.loc[r_idx, col]
                    if pd.isna(cell) or str(cell).strip()=='':
                        continue
                    txt = str(cell).strip()

                    # --- Horizontal fields ---
                    for key in h_fields:
                        if txt.startswith(key) and found_values.get(key) is None:
                            val = txt[len(key):].strip()
                            if val:
                                found_values[key] = val
                            else:
                                # try next column
                                next_i = single_letter_cols.index(col) + 1
                                if next_i < len(single_letter_cols):
                                    nxt = client_df.loc[r_idx, single_letter_cols[next_i]]
                                    if pd.notna(nxt):
                                        found_values[key] = str(nxt).strip()
                            break

                    # Special: Previous Address (street + next row)
                    if txt.startswith('Previous Address:') and address_street is None:
                        addr = txt[len('Previous Address:'):].strip()
                        # maybe next row in same col
                        below = client_df.loc[r_idx+1, col] if r_idx+1 < len(client_df) else None
                        if pd.notna(below):
                            addr += ' ' + str(below).strip()
                        address_street = addr

                    # --- Vertical: Insurance block ---
                    if txt == 'Insurance':
                        stop_row = None
                        for rr in range(r_idx+1, len(client_df)):
                            v = client_df.loc[rr, col]
                            if pd.isna(v): continue
                            vs = str(v).strip()
                            if vs == 'Additional Fields:':
                                stop_row = rr
                                break
                            potential_insurances.append({'name': vs, 'row': rr})
                        insurance_stop_by_col[col] = stop_row or len(client_df)
                        continue

                    # --- Vertical: Insured's ID # block ---
                    if txt == "Insured's ID #":
                        stop_at = insurance_stop_by_col.get(col, len(client_df))
                        for rr in range(r_idx+1, stop_at):
                            v = client_df.loc[rr, col]
                            if pd.isna(v): continue
                            potential_insured_ids.append({'id': str(v).strip(), 'row': rr})
                        continue

            # --- Assign horizontal findings ---
            for key, val in found_values.items():
                if val is not None and key not in ('Insurance', "Insured's ID #"):
                    resident[key] = val

            if address_street:
                resident['Previous Address:'] = address_street

            # --- Merge split insurance names ending with '-' ---
            merged_ins = []
            i = 0
            while i < len(potential_insurances):
                entry = potential_insurances[i]
                name, row = entry['name'], entry['row']
                if name.endswith('-') and i+1 < len(potential_insurances):
                    # append next line
                    nxt = potential_insurances[i+1]['name']
                    name = name.rstrip('-').rstrip() + ' ' + nxt
                    i += 2
                else:
                    i += 1
                merged_ins.append({'name': name, 'row': row})
            potential_insurances = merged_ins

            # --- Choose Insurance ---
            primary = resident.get('Primary Payer:', '')
            names = [e['name'] for e in potential_insurances]
            selected_insurance = None

            if primary in names:
                selected_insurance = primary
            elif 'Private' in primary:
                filt = [e for e in potential_insurances
                        if 'Medicare' not in e['name'] and 'Medicaid' not in e['name']]
                if filt:
                    selected_insurance = filt[0]['name']
            if not selected_insurance and names:
                selected_insurance = names[0]

            resident['Insurance'] = selected_insurance or ''

            # --- Match Insured's ID by row ---
            selected_id = ''
            if selected_insurance:
                ins_row = next((e['row'] for e in potential_insurances
                                if e['name'] == selected_insurance), None)
                if ins_row is not None:
                    id_match = next((e for e in potential_insured_ids
                                    if e['row'] == ins_row), None)
                    if id_match:
                        selected_id = id_match['id']

            resident["Insured's ID #"] = selected_id or ''

            # Add to results
            residents_all.append(resident)

        return residents_all

    residents_data = extract_resident_data(df_fs_c1)

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    df_residents_c1 = pd.DataFrame(residents_data)
    if df_residents_c1.empty:
        log_fn("⚠️ No residents found in the file.")
        log_fn("Please ensure this is the right file or that the file is in the correct format.")
        log_fn("No output file will be created.")
        return
    df_residents_c1 = df_residents_c1.copy()
    log_fn(f"Thre are roughly {df_residents_c1.shape[0]} residents in the file.")

    # apply name_splitter to each Resident, turn each dict into a Series,
    # then grab only the three name‐fields
    split_df = df_residents_c1['Resident'] \
        .apply(dmr.name_splitter) \
        .apply(pd.Series)[['First_Name','Middle_Name','Last_Name']]

    df_residents_c2 = pd.concat([df_residents_c1, split_df], axis=1)
    df_residents_c2 = df_residents_c2.copy()

    #Removing extra info from physician names
    df_residents_c2['Attending:'] = df_residents_c2['Attending:'].fillna('')
    df_residents_c2['Primary_Physician_Code'] = df_residents_c2['Attending:'].str.split('-').str[0]
    df_residents_c2['Primary_Physician_Code'] = df_residents_c2['Primary_Physician_Code'].str.replace('Dr.', '')
    df_residents_c2['Primary_Physician_Code'] = df_residents_c2['Primary_Physician_Code'].str.strip()

    def clean_ssn(ssn):
        """
        Takes any input, strips out non‐digits, and if exactly 9 digits remain,
        returns them in ###-##-#### form. Otherwise returns None.
        """
        ssn_str = str(ssn)
        # strip out everything but digits
        digits = re.sub(r'\D', '', ssn_str)
        if len(digits) != 9:
            return None
        return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"


    df_residents_c2['Social_Security_Number'] = df_residents_c2['SSN:'].apply(clean_ssn)

    def clean_birthdate(bd):
        """
        Takes any input, tries to parse it as %m/%d/%Y.
        If successful, returns the string in that exact format.
        Otherwise returns None.
        """
        try:
            dt = pd.to_datetime(bd, format='%m/%d/%Y', errors='raise')
            return dt.strftime('%m/%d/%Y')
        except Exception:
            return None

    # Apply the cleaner
    df_residents_c2['Date_Of_Birth'] = \
        df_residents_c2['Birth Date:'].apply(clean_birthdate)

    def extract_postal_code(address):
        """Extract US ZIP (5 or 9 digits) at the end of the string; return (zip, cleaned_addr)."""
        if pd.isna(address) or str(address).strip() == '':
            return pd.NA, pd.NA

        s = str(address)
        m = re.search(r'\b\d{5}(?:-\d{4})?\b$', s)
        if m:
            zip_code = m.group(0)
            new_addr = re.sub(r'\b\d{5}(?:-\d{4})?\b$', '', s).strip()
            if not new_addr:
                new_addr = pd.NA
            return zip_code, new_addr

        return pd.NA, s

    df_residents_c2[['Postal_Zip_Code', 'Previous Address:']] = (
        df_residents_c2['Previous Address:']
        .apply(extract_postal_code)
        .apply(pd.Series)
    )

    us_state_codes = dmr.us_state_codes

    state_pattern = r",\s*([A-Z]{2})$"
    codes = df_residents_c2['Previous Address:'].str.extract(state_pattern, expand=False)
    valid = set(us_state_codes.keys())
    df_residents_c2['Prov_State'] = codes.where(codes.isin(valid), '')

    mask = df_residents_c2['Prov_State'] != ''
    df_residents_c2.loc[mask, 'Previous Address:'] = (
        df_residents_c2.loc[mask, 'Previous Address:']
        .str.replace(state_pattern, '', regex=True)
        .str.strip()
    )

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    def extract_city(row):
            """
            Uses Postal_Zip_Code to narrow the set of possible city names from df_zip_to_county,
            applies fuzzy matching on Address1 to identify if one of the candidate cities is present,
            and if so extracts the official city name and county. It then returns the city, county,
            and a new Address1 with the matched city name removed.
            """
            address1 = row['Previous Address:']
            postal_zip = row['Postal_Zip_Code']
            
            # If there is no zip code, we do not change the address.
            if pd.isna(postal_zip) or postal_zip.strip() == "":
                return pd.Series(['', address1])
            
            # Here we assume exact match on zip codes between the address and the reference table.
            # You may wish to adjust for leading zeros or different formats if needed.
            matches_zip = df_zip_to_county[df_zip_to_county['zip'] == postal_zip.strip()]
            
            # If no matching zip is found, leave Address1 as is.
            if matches_zip.empty:
                return pd.Series(['', address1])
            
            # Prepare to compare text in lower case. Note: zip codes are typically numeric so we may not need to alter them.
            address1_lower = str(address1).lower()
            
            best_match = None  # will hold the candidate city with best score
            best_score = 0
            best_row = None  # will eventually hold the row from df_zip_to_county corresponding to best match
            
            # For each matching zip code record, consider the primary city and its acceptable alternatives.
            for _, zip_row in matches_zip.iterrows():
                # Build the list of candidate city names from the reference.
                candidates = [zip_row['primary_city']]
                # Split acceptable_cities by comma and strip any whitespace.
                if pd.notna(zip_row['acceptable_cities']):
                    candidates.extend([c.strip() for c in zip_row['acceptable_cities'].split(',') if c.strip() != ""])
                
                # Iterate over the candidate cities and look for a fuzzy match in Address1.
                for city in candidates:
                    # Compute the fuzzy match score between the candidate city (in lower case) and the Address1 text.
                    score = fuzz.partial_ratio(str(city).lower(), address1_lower)
                    if score > best_score and score >= 90:  # Only consider matches with score over or equal to 90
                        best_score = score
                        best_match = city  # store the official city name from the reference
                        best_row = zip_row
                        
            # If no candidate is found, leave the columns empty.
            if best_match is None:
                return pd.Series(['', address1])
            
            # Remove the found city substring from Address1.
            # We use re.compile with re.IGNORECASE to make a case-insensitive substitution.
            pattern = re.compile(re.escape(best_match), re.IGNORECASE)
            # Remove the matched city from Address1. Also, remove any leftover trailing commas.
            new_address1 = pattern.sub('', address1).strip().rstrip(',')
            
            # Return the official city name, county (from best_row), and cleaned Address1.
            return pd.Series([best_match, new_address1])

    df_residents_c2[['City', 'Previous Address:']] = \
            df_residents_c2.apply(extract_city, axis=1)

    def remove_single_char_if_only(s):
        """
          If, after stripping non‑alphanumerics, only 1 char remains, return NA.
          Otherwise return the original value.
        """
        if pd.isna(s):                      # handles pd.NA, np.nan, None
            return pd.NA

        s_str = str(s)
        cleaned = re.sub(r'[^A-Za-z0-9]', '', s_str)
        if len(cleaned) == 1:
            return pd.NA                    # or '' if you prefer empty strings
        return s_str
            
    df_residents_c2['Previous Address:'] = \
        df_residents_c2['Previous Address:'].apply(remove_single_char_if_only)

    df_residents_c2['Medicare A #:'] = \
        df_residents_c2['Medicare A #:'].apply(remove_single_char_if_only)

    df_residents_c2['Medicaid #:'] = \
        df_residents_c2['Medicaid #:'].apply(remove_single_char_if_only)

    #Checking the integrity of the Sex column

    sex_first = df_residents_c2['Sex:'] \
        .astype(str) \
        .str.strip() \
        .str.upper() \
        .str[:1]

    df_residents_c2['Sex:'] = sex_first.where(sex_first.isin(['M','F']), '')

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    def contain_phone(s):
        output = ''
        if s:
            s = str(s).strip()
            digits = [c for c in s if c.isdigit()]
            if len(digits) == 10:
                phone_number = ''.join(digits)
                if phone_number != phone_number[0] * 10:
                    output = f'({phone_number[:3]}) {phone_number[3:6]}-{phone_number[6:]}'
        return output

    df_residents_c2['Phone_Home'] = df_residents_c2['Phone:'].apply(contain_phone)

    marital_map = {
        'single': 'Never married',
        'never married': 'Never married',
        'divorced': 'Divorced',
        'widowed': 'Widowed',
        'married': 'Married',
        'separated': 'Separated'
    }

    df_residents_c2['Marital_Status_Code'] = df_residents_c2['Marital Status:'] \
        .apply(lambda x: marital_map[x.lower()] 
                        if isinstance(x, str) and x.lower() in marital_map 
                        else '-UTD')

    mask = df_residents_c2['Preferred Language:'] == ''
    df_residents_c2.loc[mask, 'Preferred Language:'] = '- Declined to Specify'

    primary_language = [
        'chinese', 'dutch', 'english', 'french', 'german', 'greek', 'hebrew',
        'italian', 'japanese', 'polish', 'portuguese', 'russian', 'spanish',
        'yugoslavian', '- declined to specify'
        ]

    df_residents_c2['Primary_Language'] = df_residents_c2['Preferred Language:'] \
        .apply(lambda x: x.title() 
                        if isinstance(x, str) and x.lower() in primary_language 
                        else 'Other')

    # Religion
    mask = df_residents_c2['Religion:'] == ''
    df_residents_c2.loc[mask, 'Religion:'] = 'Unknown'

    # Race
    mask = df_residents_c2['Race:'] == ''
    df_residents_c2.loc[mask, 'Race:'] = '- Declined to Specify'

    # Ethnicity
    mask = df_residents_c2['Ethnicity:'] == ''
    df_residents_c2.loc[mask, 'Ethnicity:'] = '- Declined to Specify'

    df_residents_c2['Org_Code'] = ''
    df_residents_c2['Facility_Id'] = facility_code
    df_residents_c2['Title'] = ''
    df_residents_c2['Address2'] = ''
    df_residents_c2['Medicare_Number'] = ''
    df_residents_c2['Citizenship_Code'] = 'U.S.'
    df_residents_c2['Education'] = ''

    columns_to_keep = ['Org_Code','Facility_Id', 'Client_ID_Number', 'Title',
                       'First_Name', 'Last_Name', 'Middle_Name',
                       'Date_Of_Birth', 'Marital_Status_Code', 'Sex:',
                       'Previous Address:', 'Address2', 'City', 'Prov_State',
                       'Postal_Zip_Code', 'Phone_Home', 'Religion:', 'Race:',
                       'Ethnicity:', 'Primary_Physician_Code', 
                       'Medicare_Number', 'Medicaid #:', 
                       'Social_Security_Number', "Insured's ID #",
                       'Insurance', 'Medicare A #:', 'Citizenship_Code',
                       'Primary_Language', 'Education', 'Prev Occupation:']

    df_residents = df_residents_c2[columns_to_keep].copy()

    # Applying some formatting to the columns
    df_residents['Last_Name'] = \
        df_residents['Last_Name'].apply(dmr.name_proper)
    df_residents['First_Name'] = df_residents['First_Name'].apply(dmr.name_proper)
    df_residents['Middle_Name'] = df_residents['Middle_Name'].apply(dmr.name_proper)
    df_residents['City'] = df_residents['City'].apply(dmr.name_proper)
    df_residents['Prev Occupation:'] = df_residents['Prev Occupation:'].apply(dmr.name_proper)
    df_residents['Insurance'] = df_residents['Insurance'].apply(dmr.name_proper)
    df_residents['Primary_Physician_Code'] = \
        df_residents['Primary_Physician_Code'].apply(dmr.name_proper)
    df_residents['Previous Address:'] = \
        df_residents['Previous Address:'].apply(dmr.address_proper)
    df_residents = df_residents.fillna('')

    def normalize_person_text(val):
        val = '' if pd.isna(val) else str(val).strip().upper()
        val = re.sub(r'["“”]', '', val)
        val = re.sub(r'[^A-Z0-9\s\-]', ' ', val)
        val = re.sub(r'\s+', ' ', val).strip()
        return val

    def field_match(a, b, threshold=95):
        a = normalize_person_text(a)
        b = normalize_person_text(b)

        if not a or not b:
            return True

        if a == b:
            return True

        if a in b or b in a:
            return True

        return max(
            fuzz.ratio(a, b),
            fuzz.partial_ratio(a, b),
            fuzz.token_set_ratio(a, b)
        ) >= threshold

    def same_person_duplicate_group(group):
        ssns = (
            group['Social_Security_Number']
            .fillna('')
            .astype(str)
            .str.strip()
        )
        valid_ssns = sorted(set(ssns[ssns != '']))

        # Priority 1: SSN
        if len(valid_ssns) == 1:
            return True, ''

        if len(valid_ssns) > 1:
            return False, f"conflicting SSNs: {', '.join(valid_ssns)}"

        birthdates = (
            group['Date_Of_Birth']
            .fillna('')
            .astype(str)
            .str.strip()
        )
        valid_birthdates = sorted(set(birthdates[birthdates != '']))

        # Priority 2: Birthdate
        if len(valid_birthdates) == 1:
            return True, ''

        if len(valid_birthdates) > 1:
            return False, f"conflicting birthdates: {', '.join(valid_birthdates)}"

        # Priority 3: Name
        base = group.iloc[0]

        for _, row in group.iloc[1:].iterrows():
            last_match = field_match(base['Last_Name'], row['Last_Name'], threshold=95)
            first_match = field_match(base['First_Name'], row['First_Name'], threshold=95)

            if not last_match or not first_match:
                return (
                    False,
                    f"name mismatch: '{base['Last_Name']}, {base['First_Name']}' "
                    f"vs '{row['Last_Name']}, {row['First_Name']}'"
                )

            base_middle = normalize_person_text(base['Middle_Name'])
            row_middle = normalize_person_text(row['Middle_Name'])

            if base_middle and row_middle:
                if base_middle[0] != row_middle[0]:
                    return (
                        False,
                        f"middle name mismatch: '{base['Middle_Name']}' vs '{row['Middle_Name']}'"
                    )

                if len(base_middle) > 1 and len(row_middle) > 1:
                    if not field_match(base_middle, row_middle, threshold=95):
                        return (
                            False,
                            f"middle name mismatch: '{base['Middle_Name']}' vs '{row['Middle_Name']}'"
                        )

        return True, ''

    def more_complete_row(group):
        non_blank_counts = group.apply(
            lambda row: sum(str(v).strip() != '' for v in row),
            axis=1
        )
        return non_blank_counts.idxmax()

    def consolidate_same_person_group(group):
        survivor_idx = more_complete_row(group)
        survivor = group.loc[survivor_idx].copy()

        for col in group.columns:
            current = str(survivor[col]).strip()
            if current:
                continue

            values = (
                group[col]
                .fillna('')
                .astype(str)
                .map(str.strip)
            )
            values = values[values != '']

            if not values.empty:
                survivor[col] = values.iloc[0]

        return survivor

    def resolve_duplicate_client_ids(df):
        output_rows = []
        warning_count = 0

        for client_id, group in df.groupby('Client_ID_Number', sort=False):
            if len(group) == 1:
                output_rows.append(group.iloc[0])
                continue

            same_person, reason = same_person_duplicate_group(group)

            if same_person:
                output_rows.append(consolidate_same_person_group(group))
            else:
                warning_count += 1
                names = (
                    group[['Last_Name', 'First_Name', 'Middle_Name', 'Social_Security_Number']]
                    .fillna('')
                    .astype(str)
                    .agg(lambda r: f"{r['Last_Name']}, {r['First_Name']} {r['Middle_Name']} | SSN: {r['Social_Security_Number']}", axis=1)
                    .tolist()
                )

                log_fn(
                    f"⚠️ DUPLICATED Client_ID_Number appears to belong to different people: "
                    f"{client_id}. Reason: {reason}. Rows kept for manual review: {' || '.join(names)}"
                )

                # Keep all rows because merging would be unsafe.
                for _, row in group.iterrows():
                    output_rows.append(row)

        result = pd.DataFrame(output_rows, columns=df.columns).reset_index(drop=True)

        if warning_count == 0:
            log_fn("✅ Duplicate Client_ID_Number check completed: no conflicting residents found.")

        return result

    def warn_duplicate_ssn_different_client_ids(df):
        ssn_col = 'Social_Security_Number'
        id_col = 'Client_ID_Number'

        ssn_series = (
            df[ssn_col]
            .fillna('')
            .astype(str)
            .str.strip()
        )

        valid_mask = ssn_series.ne('')

        df_valid_ssn = df.loc[valid_mask, [id_col, ssn_col]].copy()

        duplicated_ssns = df_valid_ssn[
            df_valid_ssn.duplicated(subset=[ssn_col], keep=False)
        ]

        if duplicated_ssns.empty:
            return

        for ssn, group in duplicated_ssns.groupby(ssn_col):
            client_ids = (
                group[id_col]
                .fillna('')
                .astype(str)
                .str.strip()
                .drop_duplicates()
                .tolist()
            )

            if len(client_ids) > 1:
                log_fn(
                    f"⚠️ DUPLICATED SSN across different Client_ID_Number values: "
                    f"SSN {ssn} belongs to Client_ID_Number(s): "
                    f"{', '.join(client_ids)}"
                )


    df_residents = resolve_duplicate_client_ids(df_residents)
    warn_duplicate_ssn_different_client_ids(df_residents)

    log_fn("✅ Resident data processing completed successfully.")
    log_fn(f"Total residents processed: {df_residents.shape[0]}")

    # Save the DataFrame to an Excel file
    # Adjust the path as necessary.
    template_path = get_template_path("CLIENT.xlsx")
    #template_path = resource_path("resources", "templates", "CLIENT.xlsx")
    if not template_path.exists():
        error_message = f"Template file not found at {template_path}"
        log_fn("⚠️ " + error_message)
        raise FileNotFoundError(error_message)
    output_file_path.parent.mkdir(parents=True, exist_ok=True)

    # Copy the template file to the output_file_path 
    # #(this creates the new file based on the template)
    shutil.copy(template_path, output_file_path)

    # Load the workbook from the newly created file
    wb = load_workbook(output_file_path)

    # Select the "Data" sheet; ensure the sheet name matches exactly
    ws = wb['Data']

    # Assume that the first row in the template is the header.
    # We will write the data starting at row 2.
    start_row = 2

    # Get the dataframe containing the final residents data.
    # Make sure that the dataframe's columns match the header in the template.
    for r_idx, row in enumerate(df_residents.itertuples(index=False),
                                start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for dv in ws.data_validations.dataValidation:
        # dv.ranges is a list of range strings (e.g. ["A2:A100"])
        new_ranges = []
        for rng in dv.ranges:
            # Parse the original range boundaries (start_col, start_row, end_col, end_row)
            rng_str = str(rng)  # ensure we pass a string, e.g. "A2:F100"
            min_col, min_row, max_col, max_row = range_boundaries(rng_str)

            # Update the end row to the current maximum row with data (ws.max_row)
            new_range = (
                f"{ws.cell(row=min_row, column=min_col).coordinate}:" 
                f"{ws.cell(row=ws.max_row, column=max_col).coordinate}"
            )
            new_ranges.append(new_range)
        # Replace the validation's range list with the updated ranges
        dv.ranges = new_ranges
        # Save the workbook to preserve changes
    wb.save(output_file_path)

    wb.close()
