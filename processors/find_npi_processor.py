#VERSION: 1.0.0

import os
import re
import shutil
import sys
from openpyxl.utils import range_boundaries
from openpyxl import load_workbook
import pandas as pd
from itertools import combinations
import requests
import json
import threading
from functools import lru_cache
from fuzzywuzzy import fuzz
from pathlib import Path
from utils import data_migration_resources as dmr
from utils.exceptions import AbortedByUser
from utils.file_utils import resource_path
#root_abs_path = Path(__file__).resolve().parent.parent.parent

def process_find_npi_file(
                    import_file_path: str,
                    output_file_path: str,
                    log_fn = None,
                    abort_event=threading.Event()
):

    if log_fn is None:
       log_fn = lambda msg: None

    if abort_event is not None and abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    if not isinstance(output_file_path, Path):
        output_file_path = Path(output_file_path)

    with pd.ExcelFile(import_file_path) as xls:
        df_import_raw = xls.parse(dtype=str)

    medprof_columns = ['Facility_Id', 'Staff_ID_Number', 'First_Name',
                       'Last_Name', 'Middle_Name', 'Medicaid_Provider_No',
                       'Medicare_Provider_No', 'Medicare_Provider_No', 
                       'Address1', 'Address2', 'Address3', 'City', 'Prov_State',
                       'Country_ID', 'Postal_Zip_Code', 'Phone_Office', 'Fax',
                       'Email_Address', 'Medical_Professional_Type',
                       'Taxonomy_Code', 'State_License_Number', 'EIN',
                       'Other_ID']

    existing = set(df_import_raw.columns)

    missing = [col for col in medprof_columns if col not in existing]

    if missing:
        log_fn("⚠️ This tool works exclusively with Medical Professionals imports.")
        log_fn(f"Missing required columns: {', '.join(missing)}")
        raise ValueError(
            "This file does not have the expected columns. "
            "Please check the import file format."
        )
    log_fn("✅ MedProf file successfully loaded.")
    df_import_c1 = df_import_raw.fillna('').copy()

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    # ——— 1) Your cached API fetcher ———
    @lru_cache(maxsize=1024)
    def fetch_npi_record(npi: str, version: str = "2.1") -> dict:
        url    = "https://npiregistry.cms.hhs.gov/api/"
        params = {"number": npi, "version": version, "pretty": "off"}
        r      = requests.get(url, params=params, timeout=5)
        r.raise_for_status()
        results = r.json().get("results", [])
        return results[0] if results else {}

    def get_npi_details(npi: str) -> dict:
        rec   = fetch_npi_record(npi)
        basic = rec.get("basic", {})

        # split out the LOCATION address bits
        loc = next((a for a in rec.get("addresses", [])
                    if a.get("address_purpose") == "LOCATION"), {})
        # taxonomy
        tax = next((t for t in rec.get("taxonomies", [])
                    if t.get("primary") in (True, "true")), None) \
            or (rec.get("taxonomies") or [None])[0]
        # medicaid
        med = next((o for o in rec.get("other_identifiers", [])
                    if o.get("identifier_type","").upper() == "MEDICAID"), {})

        return {
            "first_name":       basic.get("first_name"),
            "middle_name":      basic.get("middle_name"),
            "last_name":        basic.get("last_name"),
            "addr_1":           loc.get("address_1"),
            "city":             loc.get("city"),
            "state":            loc.get("state"),
            "postal_code":      loc.get("postal_code"),
            "phone":            loc.get("telephone_number"),
            "fax":              loc.get("fax_number"),
            "taxonomy_code":    tax.get("code")    if tax else None,
            "license_number":   tax.get("license") if tax else None,
            "license_state":    tax.get("state")   if tax else None,
            "medicaid_id":      med.get("identifier")
        }

    # ——— 2) Row‐wise merger function ———
    API_TO_DF = {
        "addr_1":         "Address1",
        "city":           "City",
        "state":          "Prov_State",
        "postal_code":    "Postal_Zip_Code",
        "phone":          "Phone_Office",
        "fax":            "Fax",
        "medicaid_id":    "Medicaid_Provider_No",
        "taxonomy_code":  "Taxonomy_Code",
        "license_number": "State_License_Number",
        "other_identifier": "Other_ID"
        
    }

    def merge_api_data(row, threshold=90):
        npi = str(row.get("National_Provider_ID", "")).strip()
        if not npi:
            return row

        record = fetch_npi_record(npi)
        if not record:
            row["National_Provider_ID"] = "INVALID"
            return row

        api = get_npi_details(npi)

        # pull & normalize for case-insensitive comparison
        existing_first = str(row.get("First_Name", "")).strip()
        existing_last  = str(row.get("Last_Name",  "")).strip()
        ef_low, el_low = existing_first.lower(), existing_last.lower()

        api_first = api.get("first_name", "") or ""
        api_last  = api.get("last_name",  "") or ""
        af_low, al_low = api_first.lower(), api_last.lower()

        # 1) swapped-names detection
        swap_first = fuzz.ratio(ef_low, al_low) if ef_low and al_low else 0
        swap_last  = fuzz.ratio(el_low, af_low) if el_low and af_low else 0
        if swap_first >= threshold and swap_last >= threshold:
            log_fn(f"NPI {npi}: detected swapped names "
                f"(imported '{existing_first} {existing_last}', API '{api_first} {api_last}')")
            row["First_Name"], row["Last_Name"] = api_first, api_last

        else:
            # 2) former-name lookup
            matched_former = False
            for other in record.get("other_names", []):
                if other.get("other_name_type", "").lower() == "former name":
                    f_first = (other.get("first_name") or "").strip()
                    f_last  = (other.get("last_name")  or "").strip()
                    if f_first and f_last:
                        full_imp = f"{ef_low} {el_low}"
                        full_for = f"{f_first.lower()} {f_last.lower()}"
                        score_for = fuzz.ratio(full_imp, full_for)
                        if score_for >= threshold:
                            log_fn(f"NPI {npi}: matched former name "
                                f"'{f_first} {f_last}', updating to current "
                                f"'{api_first} {api_last}'")
                            row["FirstName"], row["LastName"] = api_first, api_last
                            matched_former = True
                            break

            if not matched_former:
                # 3) normal name-match logic
                score_first = fuzz.ratio(ef_low, af_low) if ef_low and af_low else 0
                score_last = fuzz.ratio(el_low, al_low) if el_low and al_low else 0

                # full mismatch → skip row
                if score_first < threshold and score_last < threshold:
                    log_fn(f"NPI {npi}: full name mismatch "
                        f"(imported '{existing_first} {existing_last}', "
                        f"API '{api_first} {api_last}')")
                    return row

                # partial mismatches overwrite individually
                if score_first < threshold <= score_last:
                    log_fn(f"NPI {npi}: first name mismatch "
                        f"(imported '{existing_first}', API '{api_first}')")
                    row["First_Name"] = api_first

                if score_last < threshold <= score_first:
                    log_fn(f"NPI {npi}: last name mismatch "
                        f"(imported '{existing_last}', API '{api_last}')")
                    row["Last_Name"] = api_last

        # 4) merge other blank fields
        for api_key, df_col in API_TO_DF.items():
            new_val = api.get(api_key)
            if new_val and (pd.isna(row[df_col]) or str(row[df_col]).strip() == ""):
                row[df_col] = new_val

        return row

    df_import_c2 = df_import_c1.apply(merge_api_data, axis=1).copy()
    
    # ——— 3) Organization vs. Individual detection ———
    org_names = ['care', 'health', 'healthcare', 'hospital', 'group', 'clinic', 'dental']

    def is_organization_row(row, org_tokens=org_names):
        """
        Returns True if any of the tokens in `org_tokens` appears (case-insensitive)
        in First_Name, Middle_Name, or Last_Name.
        """
        name_str = " ".join(
            str(row.get(col, "")).lower()
            for col in ("First_Name", "Middle_Name", "Last_Name")
        )
        return any(tok in name_str for tok in org_tokens)

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    def npi_full_lookup(
        first_name: str = "",
        last_name: str = "",
        organization_name: str = "",
        limit: int = 50,
        enumeration_type: str = None
    ) -> pd.DataFrame:
        """
        Returns a DataFrame with these columns (for every matching record):
        - npi
        - legal_name
        - other_names
        - suffix
        - practice_address
        - phone
        - fax
        - primary_taxonomy
        - license_number
        - license_state
        - medicaid_id

        Pass either (first_name + last_name + enumeration_type='NPI-1') for individuals,
        or (organization_name + enumeration_type='NPI-2') for orgs. Even if there are
        no hits, we’ll return an EMPTY DataFrame with the correct column names.
        """
        # 1) Define the exact list of column names we want, regardless of hits:
        cols = [
            "npi",
            "legal_name",
            "other_names",
            "suffix",
            "practice_address",
            "phone",
            "fax",
            "primary_taxonomy",
            "license_number",
            "license_state",
            "medicaid_id"
        ]

        base_url = "https://npiregistry.cms.hhs.gov/api/"
        params = {"limit": limit, "version": "2.1", "pretty": "off"}

        if organization_name:
            params["organization_name"] = organization_name
            params["enumeration_type"]  = enumeration_type or "NPI-2"
        else:
            params["first_name"]       = first_name
            params["last_name"]        = last_name
            params["enumeration_type"] = enumeration_type or "NPI-1"

        resp = requests.get(base_url, params=params)
        resp.raise_for_status()
        results = resp.json().get("results", [])

        rows = []
        for rec in results:
            basic = rec.get("basic", {})

            # a) Determine legal_name
            if organization_name:
                legal_name = basic.get("organization_name", "") or ""
            else:
                fname = basic.get("first_name", "") or ""
                lname = basic.get("last_name", "") or ""
                legal_name = f"{fname} {lname}".strip()

            # b) Collect 'other_names'
            other_list = []
            for on in rec.get("other_names", []):
                name = on.get("other_name")
                if name and isinstance(name, str):
                    other_list.append(name.strip())

            # c) Suffix
            suffix = basic.get("name_suffix") or ""

            # d) Address & contact
            loc = next(
                (a for a in rec.get("addresses", [])
                if a.get("address_purpose") == "LOCATION"),
                {}
            )
            street = " ".join(filter(None, [loc.get("address_1"), loc.get("address_2")]))
            city, st, zipc = loc.get("city"), loc.get("state"), loc.get("postal_code")
            practice_address = f"{street}, {city}, {st} {zipc}".strip().strip(",")
            phone = loc.get("telephone_number", "") or ""
            fax   = loc.get("fax_number", "")       or ""

            # e) Taxonomy + license
            tax = next(
                (t for t in rec.get("taxonomies", [])
                if t.get("primary") in (True, "true")),
                None
            )
            if not tax:
                tax = rec.get("taxonomies", [{}])[0]

            tax_code = tax.get("code")    or ""
            license_num = tax.get("license") or ""
            license_state = tax.get("state")   or ""

            # f) Medicaid ID
            med = next(
                (oi for oi in rec.get("other_identifiers", [])
                if oi.get("identifier_type", "").upper() == "MEDICAID"),
                {}
            )
            medicaid_id = med.get("identifier") or ""

            rows.append({
                "npi":               rec.get("number", ""),
                "legal_name":        legal_name,
                "other_names":       other_list,
                "suffix":            suffix,
                "practice_address":  practice_address,
                "phone":             phone,
                "fax":               fax,
                "primary_taxonomy":  tax_code,
                "license_number":    license_num,
                "license_state":     license_state,
                "medicaid_id":       medicaid_id
            })

        # 2) If we never appended any rows, return an empty DataFrame with the column names
        if not rows:
            return pd.DataFrame([], columns=cols)

        # 3) Otherwise, build the DataFrame normally
        return pd.DataFrame(rows)

    def normalize_phone(p: str) -> str:
        return re.sub(r'\D', '', str(p) or '')

    def normalize_address(a: str) -> str:
        return re.sub(r'[^\w\s]', '', str(a) or '').lower()

    def normalize_other_id(x: str) -> str:
        # strip whitespace, drop any non-alphanumeric, lowercase
        return re.sub(r'\W+', '', str(x) or '').lower()

    def extract_address_components(practice_address: str):
        """
        Given a practice_address string in the form "street, city, state zip",
        returns a dict with keys: 'street', 'city', 'state', 'zip'.
        If any component is missing, value will be an empty string.
        """
        parts = [p.strip() for p in practice_address.split(",")]
        
        street = parts[0] if len(parts) > 0 else ""
        city = parts[1] if len(parts) > 1 else ""
        
        state = ""
        zip_code = ""
        if len(parts) > 2:
            state_zip = parts[2].strip()
            match = re.match(r"(\S+)\s+(\S+)", state_zip)
            if match:
                state, zip_code = match.groups()
            else:
                # If no zip present, treat entire string as state (or zip)
                state = state_zip
        
        return {
            "street": street.lower(),
            "city": city.lower(),
            "state": state.lower(),
            "zip": zip_code
        }

    def normalize_street_for_matching(street: str) -> str:
        """
        Lowercase, remove punctuation, strip out suite or apt numbers.
        E.g. "123 Main St Suite 200" -> "123 main st"
        """
        # Remove suite/apt identifiers
        street_no_suite = re.sub(r"\b(suite|ste|apt|#)\s*\S+", "", street, flags=re.IGNORECASE)
        # Remove any remaining punctuation
        street_clean = re.sub(r"[^\w\s]", "", street_no_suite)
        return street_clean.lower().strip()
    
    def is_blank(val):
        """
        Return True if val is None, pd.NA, NaN, or a string that's blank/whitespace.
        """
        if pd.isna(val):
            return True
        if isinstance(val, str) and val.strip() == "":
            return True
        return False

    def lookup_and_merge_by_name(row):
        """
        0) If National_Provider_ID already exists → skip.
        1) Fetch candidates via npi_full_lookup (org vs. individual).
        2) For individuals: enforce exact match on row.First_Name & row.Last_Name (using legal_name tokens).
        3) If zero → set "NOT FOUND" and return.
        4) If exactly one → unpack into blank columns.
        5) If multiple → refine by phone/fax → state → city → zip → street.
        If still not exactly one → set "MULTIPLE VALUES" and return.
        """

        # ─── 0) Skip if NPI already exists ──────────────────────────────────────
        if str(row.get("National_Provider_ID") or "").strip():
            return row

        # ─── 1) Determine org vs. individual, fetch all candidates ──────────────
        if is_organization_row(row):
            # Organization branch (no exact‐name enforcement here)
            org_input = row.get("Last_Name", "").strip()
            lower_org = org_input.lower()

            all_orgs = npi_full_lookup(
                organization_name=org_input,
                enumeration_type="NPI-2"
            )
            # Keep rows where legal_name or any other_names contain org_input substring
            mask_org = (
                all_orgs["legal_name"].str.lower().str.contains(lower_org, na=False)
                |
                all_orgs["other_names"].apply(
                    lambda lst: any(lower_org in (on or "").lower() for on in lst or [])
                )
            )
            candidates = all_orgs[mask_org] if mask_org.any() else all_orgs.copy()

            # (We do NOT attempt first/last matching for organizations—legal_name check above suffices.)

        else:
            # ── Individual branch: strip suffix from Last_Name if present ─
            last_raw = str(row.get("Last_Name", "")).strip()
            suffix_match = re.search(r"\b(Jr\.?|Sr\.?)$", last_raw, flags=re.IGNORECASE)
            if suffix_match:
                my_suffix = suffix_match.group(1).lower().replace(".", "")
                last_base = last_raw[:suffix_match.start()].strip()
            else:
                my_suffix = None
                last_base = last_raw

            # Call NPI-1 lookup with stripped last name
            candidates = npi_full_lookup(
                first_name=row.get("First_Name", "").strip(),
                last_name=last_base,
                enumeration_type="NPI-1"
            )

            # If there was a “Jr./Sr.” in your row, filter by suffix column
            if my_suffix:
                candidates = candidates[
                    candidates["suffix"]
                    .fillna("")
                    .str.lower()
                    .str.replace(".", "", regex=False)
                    == my_suffix
                ]

            # ── 2) ENFORCE EXACT MATCH on First_Name & last_base via legal_name tokens ──
            # Extract lowercase first & last tokens from the row
            fname_in_row = row.get("First_Name", "").strip().lower()
            lname_in_row = last_base.lower()

            if fname_in_row and lname_in_row:
                def tokens_of(name_str):
                    # Lowercase & grab all alphabetic words
                    return re.findall(r"[a-z]+", name_str.lower())

                # Build a boolean mask: both first & last must appear as tokens in legal_name
                mask_name = candidates["legal_name"].fillna("").apply(
                    lambda legal: (
                        (fname_in_row in tokens_of(legal)) and
                        (lname_in_row in tokens_of(legal))
                    )
                )

                if mask_name.any():
                    candidates = candidates[mask_name]
                else:
                    # No exact‐name candidate → bail out as NOT FOUND
                    row["National_Provider_ID"] = "NOT FOUND"
                    return row
            # If either fname_in_row or lname_in_row is blank, we simply skip this strict check.

        # ─── 3) ZERO / ONE / MULTIPLE ────────────────────────────────────
        if len(candidates) == 0:
            row["National_Provider_ID"] = "NOT FOUND"
            return row

        if len(candidates) == 1:
            rec = candidates.iloc[0]
        else:
            # Prepare normalized phone/fax from the row
            rp = normalize_phone(row.get("Phone_Office"))
            rf = normalize_phone(row.get("Fax"))

            # ── a) Filter by phone/fax ───────────────────────────────────
            cand_pf = candidates.copy()
            cand_pf["phone_norm"] = cand_pf["phone"].apply(normalize_phone)
            cand_pf["fax_norm"]   = cand_pf["fax"].apply(normalize_phone)

            matched_pf = cand_pf[
                (cand_pf["phone_norm"] == rp) |
                (cand_pf["fax_norm"]   == rf)
            ]

            if len(matched_pf) == 1:
                rec = matched_pf.iloc[0]
            else:
                # ── b) Phone/fax did not yield exactly one → address refinements ──
                row_state = str(row.get("Prov_State", "")).strip().lower()
                row_city = str(row.get("City", "")).strip().lower()
                row_zip = str(row.get("Postal_Zip_Code", "")).strip()
                row_street = str(row.get("Address1", "")).strip()

                pool = candidates.copy()  # revert to the full list

                # ── c) Filter by state ─────────────────────────────────────
                if row_state:
                    filtered_state = []
                    for _, cand in pool.iterrows():
                        addr = \
                            extract_address_components(cand["practice_address"])
                        if addr["state"] == row_state:
                            filtered_state.append(cand)
                    if filtered_state:
                        pool = pd.DataFrame(filtered_state)

                # ── d) If still >1, filter by city ──────────────────────────
                if len(pool) > 1 and row_city:
                    filtered_city = []
                    for _, cand in pool.iterrows():
                        addr = \
                            extract_address_components(cand["practice_address"])
                        if addr["city"] == row_city:
                            filtered_city.append(cand)
                    if filtered_city:
                        pool = pd.DataFrame(filtered_city)

                # ── e) If still >1, filter by zip ────────────────────────────
                if len(pool) > 1 and row_zip:
                    filtered_zip = []
                    for _, cand in pool.iterrows():
                        addr = \
                            extract_address_components(cand["practice_address"])
                        if addr["zip"] == row_zip:
                            filtered_zip.append(cand)
                    if filtered_zip:
                        pool = pd.DataFrame(filtered_zip)

                # ── f) If still >1, filter by street (ignore suite/apt) ─────
                if len(pool) > 1 and row_street:
                    row_street_norm = normalize_street_for_matching(row_street)
                    filtered_street = []
                    for _, cand in pool.iterrows():
                        addr = \
                            extract_address_components(cand["practice_address"])
                        cand_street_norm = \
                            normalize_street_for_matching(addr["street"])
                        if row_street_norm == cand_street_norm:
                            filtered_street.append(cand)
                    if filtered_street:
                        pool = pd.DataFrame(filtered_street)

                # ── g) After all refinements, exactly one?
                if len(pool) == 1:
                    rec = pool.iloc[0]
                else:
                    # ── h) Still ambiguous (zero or >1) → mark and bail out
                    row["National_Provider_ID"] = "MULTIPLE VALUES"
                    return row

        # ─── 4) Unpack single “rec” into row—but only into blank columns ───
        if is_blank(row.get("National_Provider_ID")):
            row["National_Provider_ID"] = rec["npi"]

        # (b) Parse rec["practice_address"] into Address1, City, Prov_State, Postal_Zip_Code
        addr_parts = [p.strip() for p in rec["practice_address"].split(",")]
        if len(addr_parts) > 0 and is_blank(row.get("Address1")):
            row["Address1"] = addr_parts[0]
        if len(addr_parts) > 1 and is_blank(row.get("City")):
            row["City"] = addr_parts[1]
        if len(addr_parts) > 2:
            st_zip = addr_parts[2].split()
            if st_zip:
                if is_blank(row.get("Prov_State")):
                    row["Prov_State"] = st_zip[0]
            if len(st_zip) > 1:
                if is_blank(row.get("Postal_Zip_Code")):
                    row["Postal_Zip_Code"] = st_zip[1]

        # (c) Phone_Office
        if is_blank(row.get("Phone_Office")):
            row["Phone_Office"] = rec["phone"]
        # (d) Fax
        if is_blank(row.get("Fax")):
            row["Fax"] = rec["fax"]
        # (e) Taxonomy_Code
        if is_blank(row.get("Taxonomy_Code")):
            row["Taxonomy_Code"] = rec["primary_taxonomy"]
        # (f) State_License_Number
        if is_blank(row.get("State_License_Number")):
            row["State_License_Number"] = rec["license_number"]
        # (g) Medicaid_Provider_No
        if is_blank(row.get("Medicaid_Provider_No")):
            row["Medicaid_Provider_No"] = rec["medicaid_id"]

        return row

    df_import = df_import_c2.apply(lookup_and_merge_by_name, axis=1).copy()
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    # Checking the columns to meet the proper format
    df_import['Phone_Office'] = \
        df_import['Phone_Office'].apply(dmr.contain_phone_2)
    df_import['Phone_Cell'] = \
        df_import['Phone_Cell'].apply(dmr.contain_phone_2)
    df_import['Phone_Pager'] = \
        df_import['Phone_Pager'].apply(dmr.contain_phone_2)
    df_import['Phone_Other'] = \
        df_import['Phone_Other'].apply(dmr.contain_phone_2)
    df_import['Phone_Home'] = \
        df_import['Phone_Home'].apply(dmr.contain_phone_2)
    df_import['Fax'] = df_import['Fax'].apply(dmr.contain_phone_2)

    df_import['First_Name'] = df_import['First_Name'].apply(dmr.name_proper)
    df_import['Last_Name'] = df_import['Last_Name'].apply(dmr.name_proper)
    df_import['Middle_Name'] = df_import['Middle_Name'].apply(dmr.name_proper)

    df_import['City'] = df_import['City'].apply(dmr.name_proper)
    df_import['Address1'] = df_import['Address1'].apply(dmr.address_proper)
    df_import['Address2'] = df_import['Address2'].apply(dmr.address_proper)
    df_import['Address3'] = df_import['Address3'].apply(dmr.address_proper)

    def proper_zipcode(s):
        if not s:
            return ''
        else:
            s = str(s).strip()
            s_no_dashes = s.replace('-','')
            digits = [c for c in s_no_dashes if c.isdigit()]
            if len(digits) == 5:
                zipcode = ''.join(digits)
            elif len(digits) > 5:
                zipcode = ''.join(digits[:5])
            else:
                zipcode = s
            return zipcode

    df_import['Postal_Zip_Code'] = df_import['Postal_Zip_Code'].apply(proper_zipcode)

    professional_type_mapping = {
        'Advanced Practice Nurse':['367A00000X', '367H00000X','364S00000X',
                        '364SA2100X','364SA2200X','364SC2300X','364SC1501X',
                        '364SC0200X','364SE0003X','364SE1400X','364SF0001X',
                        '364SG0600X','364SH1100X','364SH0200X','364SI0800X',
                        '364SL0600X','364SM0705X','364SN0000X','364SN0800X',
                        '364SX0106X','364SX0200X','364SX0204X','364SP0200X',
                        '364SP1700X','364SP2800X','364SP0808X','364SP0809X',
                        '364SP0807X','364SP0810X','364SP0811X','364SP0812X',
                        '364SP0813X','364SR0400X','364SS0200X','364ST0500X',
                        '364SW0102X','367500000X'],
        'Alternate Physician':['PENDING'],
        'Anesthesiologist':['207L00000X','207LA0401X','207LC0200X','207LH0002X',
                            '207LP2900X', '207LP3000X', '207LP4000X'],
        'Audiologist':['231H00000X', '231HA2400X',  '231HA2500X', '237600000X'],
        'Cardiologist': ['207RA0001X', '207RI0011X', '207RC0000X', '207UN0901X',
                            '2080P0202X'],
        'Chiropractor': ['111N00000X','111NI0013X', '111NI0900X', '111NN0400X', 
                            '111NN1001X', '111NX0100X', '111NX0800X', '111NP0017X',
                            '111NR0200X', '111NR0400X', '111NS0005X', '111NT0100X'],
        'Dental Hygienist':['124Q00000X'],
        'Dentist':['122300000X', '1223D0004X', '1223D0001X', '1223E0200X',
                    '1223G0001X', '1223P0106X','1223X0008X','1223S0112X',
                    '125Q00000X','1223X2210X','1223X0400X','1223P0221X',
                    '1223P0300X', '1223P0700X'],
        'Dermatologist':['207NI0002X', '207ND0900X', '207ND0101X','207NP0225X',
                            '207NS0135X'],
        'Endocrinologist':['207RE0101X', '2080P0205X', '207VE0102X'],
        'Gastroenterologist':['207RG0100X','2080P0206X', ],
        'Geriatrician':['207QG0300X', '207RG0300X'],
        'Medical Director':['PENDING'],
        'Medical Specialist':['207RH0005X', '207RA0201X', '207RH0000X', 
                        '207RI0008X', '207RM1200X', '207RI0200X', '207RP1002X',
                        '207RS0012X', '207RS0010X', '207RA0401X', '207VG0400X',
                        '193200000X','193400000X', '103K00000X', '174400000X',
                        '208000000X', '2081P0301X','2081H0002X','2081N0008X',
                        '2081P2900X','2081P0010X','2081P0004X','2081S0010X',
                        '173000000X'],
        'Medical Student':['390200000X'],
        'Nephrologist':['207RN0300X'],
        'Neurologist':['2084N0400X', '2084N0402X', '2084N0008X', '2081N0008X'],
        'Neuropsychologist':['2251N0400X'],
        'Neurosurgeon':['207T00000X'],
        'Nurse Practitioner':['363L00000X','363LA2100X','363LA2200X','363LC1500X',
                                '363LC0200X','363LF0000X','363LG0600X','363LN0000X',
                                '363LN0005X','363LX0001X','363LX0106X','363LP0200X',
                                '363LP0222X','363LP1700X','363LP2300X','363LP0808X',
                                '363LS0200X','363LW0102X'],
        'Oncologist':['207RH0003X'],
        'Ophthalmologist':['207W00000X','207WX0120X','207WX0009X','207WX0109X',
                            '207WX0200X','207WX0110X','207WX0107X','207WX0108X'],
        'Optometrist':['152W00000X','152WC0802X','152WL0500X','152WX0102X',
                        '152WP0200X','152WS0006X','152WV0400X','156F00000X',
                        '156FC0800X','156FC0801X','156FX1700X','156FX1100X',
                        '156FX1101X','156FX1800X','156FX1201X','156FX1202X',
                        '156FX1900X'],
        'Orthopedist':['207X00000X','207XS0114X','207XX0004X','207XS0106X',
                        '207XS0117X','207XX0801X','207XP3100X','207XX0005X'],
        'Otolaryngologist':['207Y00000X','207YS0123X','207YX0602X','207YX0905X',
                            '207YX0901X','207YP0228X','207YX0007X','207YS0012X'],
        'Physician':['208D00000X'],
        'Physicians Assistant':['363A00000X', '363AM0700X','363AS0400X'],
        'Podiatrist':['213E00000X','213ES0103X','213ES0131X','213EG0000X',
                        '213EP1101X','213EP0504X','213ER0200X','213ES0000X'],
        'Primary Physician':['207Q00000X', '207QA0505X', '207QH0002X',
                                '207R00000X', '208M00000X', '207RH0002X', 
                                '208100000X'],
        'Psychiatrist':['2084A0401X','2084P0802X','2084B0040X','2084P0301X',
                        '2084P0804X','2084N0600X','2084D0003X','2084E0001X',
                        '2084F0202X','2084P0805X','2084H0002X','2084A2900X',
                        '2084P0005X','2084N0008X','2084B0002X','2084P2900X',
                        '2084P0800X','2084P0015X','2084S0012X','2084S0010X',
                        '2084V0102X'],
        'Psychologist':['103G00000X','103GC0700X', '103T00000X','103TA0400X',
                        '103TA0700X','103TC0700X','103TC2200X','103TB0200X',
                        '103TC1900X','103TE1000X','103TE1100X','103TF0000X',
                        '103TF0200X','103TP2701X','103TH0004X','103TH0100X',
                        '103TM1700X','103TM1800X','103TP0016X','103TP0814X',
                        '103TP2700X','103TR0400X','103TS0200X','103TW0100X'],
        'Pulmonologist':['207RP1001X'],
        'Rheumatologist':['207RR0500X'],
        'Surgeon':['208200000X','2082S0099X','2082S0105X', '208600000X',
                    '2086H0002X','2086S0120X','2086P0122X','2086S0122X',
                    '2086S0105X','2086S0102X','2086X0206X','2086S0127X',
                    '2086S0129X','208G00000X','204F00000X'],
        'Urologist':['208800000X','2088F0040X','2088P0231X'],
        'Wound Specialist':['163WW0000X']
        }

    code_to_prof = {
        code: prof
        for prof, codes in professional_type_mapping.items()
        for code in codes
    }

    # 3) now overwrite in your df_import
    #    only for rows where Taxonomy_Code is non-empty
    mask = (
        df_import['Taxonomy_Code'].notna() &
        (df_import['Taxonomy_Code'].astype(str).str.strip() != "")
    )

    # map each code → prof, defaulting to "NOT FOUND"
    df_import.loc[mask, 'Medical_Professional_Type'] = (
        df_import.loc[mask, 'Taxonomy_Code']
                .map(code_to_prof)
                .fillna('PRIMARY PHYSICIAN')
    )

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    template_path = resource_path("respurces", "templates", 
                                  "MEDICAL_PROFESSIONAL.xlsx")
    if not template_path.exists():
        error_message = f"Template file not found at {template_path}"
        log_fn("⚠️ " + error_message)
        raise FileNotFoundError(error_message)
    output_file_path.parent.mkdir(parents=True, exist_ok=True)
    # Copy the template file to the output_file_path (this creates the new file based on the template)
    shutil.copy(template_path, output_file_path)

    # Load the workbook from the newly created file
    wb = load_workbook(output_file_path)

    # Select the "Data" sheet; ensure the sheet name matches exactly
    ws = wb['Medical Prof Data']

    # Assume that the first row in the template is the header.
    # We will write the data starting at row 2.
    start_row = 2

    # Get the dataframe containing your final allergy data.
    # Make sure that the dataframe's columns match the header in the template.
    for r_idx, row in enumerate(df_import.itertuples(index=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for dv in ws.data_validations.dataValidation:
        # dv.ranges is a list of range strings (e.g. ["A2:A100"])
        new_ranges = []
        for rng in dv.ranges:
            # Parse the original range boundaries (start_col, start_row, end_col, end_row)
            rng_str = str(rng)  # ensure we pass a string, e.g. "A2:F100"
            min_col, min_row, max_col, max_row = range_boundaries(rng_str)

            # Update the end row to the current maximum row with data (ws.max_row)
            new_range = (
                f"{ws.cell(row=min_row, column=min_col).coordinate}:" 
                f"{ws.cell(row=ws.max_row, column=max_col).coordinate}"
            )
            new_ranges.append(new_range)
        # Replace the validation's range list with the updated ranges
        dv.ranges = new_ranges
        # Save the workbook to preserve changes
    wb.save(output_file_path)

    return output_file_path
