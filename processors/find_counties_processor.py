#VERSION: 1.0.0

import os
import re
import shutil
import sys
import threading
from openpyxl.utils import range_boundaries
from openpyxl import load_workbook
import pandas as pd
from fuzzywuzzy import fuzz
from itertools import combinations
from pathlib import Path
from utils import data_migration_resources as dmr
from utils.exceptions import AbortedByUser
from utils.file_utils import resource_path

root_abs_path = Path(__file__).resolve().parent.parent.parent

def process_find_counties_file(
    import_file_path: str,
    output_file_path: str,
    log_fn = None,
    abort_event=threading.Event()
):  
    
    xlsx = pd.ExcelFile(import_file_path)
    sheet_name = xlsx.sheet_names[0]

    if log_fn is None:
       log_fn = lambda msg: None
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    if not isinstance(output_file_path, Path):
        output_file_path = Path(output_file_path)

    df_import = pd.read_excel(import_file_path)
    df_import = df_import.fillna('')
    
    contacts_columns = ['Facility_Id', 'Client_ID_Number', 'Last_Name', 'First_Name',
                        'Middle_Name', 'Address1', 'Address2', 'County', 
                        'Postal_Zip_Code', 'City', 'Prov_State','Country']

    existing = set(df_import.columns)

    missing = [col for col in contacts_columns if col not in existing]

    if missing:
        log_fn("This option works exclusively with Contacts imports.")
        log_fn(f"⚠️ Missing required columns: {', '.join(missing)}")
        raise ValueError(
            "This file does not have the expected columns. "
            "Please ensure it is a Contacts import file."
        )
        
    try:
        zip_to_county_db = resource_path(
            "resources", "databases", "zip_code_database.xlsx"
        )
        df_zip_to_county = pd.read_excel(
            zip_to_county_db,
            usecols=['zip', 'primary_city', 'acceptable_cities', 'county',
                     'state']
        )
        df_zip_to_county['zip'] = df_zip_to_county['zip'].apply(
            lambda x: f"{int(x):05d}" if pd.notnull(x) else x
        )
    except FileNotFoundError:
        log_fn("Zip code database not found. Please ensure the file exists in the resources directory.")
        raise FileNotFoundError(
            "The zip code database file is missing. "
            "Please ensure it exists in the resources directory."
        )   
   
    log_fn("✅ Contacts file successfully loaded.")
    # -----------------------------------------------------------------------------
    # 1) Utility functions for normalization & fuzzy
    # -----------------------------------------------------------------------------
    def normalize_zip(val):
        if pd.isna(val) or (isinstance(val, str) and val.strip() == ""):
            return ""
        try:
            return str(int(val)).zfill(5)
        except (ValueError, TypeError):
            return str(val).strip()

    def normalize_string(s: str) -> str:
        return re.sub(r"[.\-\s]+", "", s.upper())

    def swap_adjacent(s, i, j):
        lst = list(s)
        lst[i], lst[j] = lst[j], lst[i]
        return "".join(lst)

    def find_corrected_zip(z, valid_zips):
        if not re.fullmatch(r"\d{5}", z):
            return None
        for i, j in combinations(range(5), 2):
            candidate = swap_adjacent(z, i, j)
            if candidate in valid_zips:
                return candidate
        return None

    # -----------------------------------------------------------------------------
    # 2) Build lookup structures
    # -----------------------------------------------------------------------------
    df_zip_to_county["zip"] = df_zip_to_county["zip"].apply(lambda x: f"{int(x):05d}" if pd.notnull(x) else x)
    df_zip_to_county["state"] = df_zip_to_county["state"].str.upper()
    df_zip_to_county["primary_city"] = df_zip_to_county["primary_city"].str.upper()

    # ZIP → county
    zip_to_county = df_zip_to_county.set_index("zip")["county"].to_dict()

    # (state, city) → set of ZIPs
    city_state_to_zips = (
        df_zip_to_county
        .groupby(["state", "primary_city"])["zip"]
        .apply(set)
        .to_dict()
    )

    # ZIP → { state, cities_set, norm_map }
    zip_to_city_state = {}
    for _, row in df_zip_to_county.iterrows():
        z = row["zip"]
        state = row["state"]
        cities = {row["primary_city"]}
        if pd.notna(row["acceptable_cities"]):
            cities |= {c.strip().upper() for c in row["acceptable_cities"].split(",")}
        zip_to_city_state[z] = {
            "state": state,
            "cities": cities,
            "norm_map": { normalize_string(c): c for c in cities }
        }

    # New: (state, city) → set of counties
    city_state_to_counties = {}
    for _, row in df_zip_to_county.iterrows():
        st = row["state"]
        cnty = row["county"]
        cities = {row["primary_city"]}
        if pd.notna(row["acceptable_cities"]):
            cities |= {c.strip().upper() for c in row["acceptable_cities"].split(",")}
        for city_upper in cities:
            city_state_to_counties.setdefault((st, city_upper), set()).add(cnty)

    # -----------------------------------------------------------------------------
    # 3) Row‐by‐row correction + record which rows change
    # -----------------------------------------------------------------------------
    corrected_city_idxs = []
    filled_state_idxs = []
    corrected_zip_idxs = []
    filled_county_idxs = []

    def lookup_and_correct(row):
        idx = row.name
        orig_zip = normalize_zip(row["Postal_Zip_Code"])
        raw_city = str(row["City"] or "").strip()
        raw_state = str(row["Prov_State"] or "").strip().upper()

        # 0) If no ZIP but city+state present & unique county → fill County only
        if not orig_zip and raw_city and raw_state:
            key = (raw_state, raw_city.upper())
            counties = city_state_to_counties.get(key, set())
            if len(counties) == 1:
                filled_county_idxs.append(idx)
                return pd.Series({
                    "Postal_Zip_Code": "",
                    "County": next(iter(counties)),
                    "City": raw_city,
                    "Prov_State": raw_state
                })

        best_city = raw_city
        best_state = raw_state
        used_zip = orig_zip

        zip_info = zip_to_city_state.get(orig_zip)
        if zip_info:
            valid_state = zip_info["state"]
            norm_map = zip_info["norm_map"]
            city_norm = normalize_string(raw_city)

            # exact match?
            if raw_city.upper() in zip_info["cities"] and raw_state == valid_state:
                pass
            else:
                # fuzzy‐score highest
                scores = [(fuzz.ratio(city_norm, n), n, p) for n,p in norm_map.items()]
                scores.sort(reverse=True, key=lambda x: x[0])
                top_score, top_norm, top_proper = scores[0]
                longest = max(len(city_norm), len(top_norm))
                one_char_thresh = (longest - 1) / longest * 100

                if top_score >= 90 or top_score >= one_char_thresh:
                    best_city = top_proper
                    corrected_city_idxs.append(idx)
                    if not raw_state or raw_state != valid_state:
                        best_state = valid_state
                        filled_state_idxs.append(idx)

            # if still no city match, try swapping ZIP digits
            if best_city == raw_city and (not raw_state or raw_state == valid_state):
                corrected = find_corrected_zip(orig_zip,
                                city_state_to_zips.get((valid_state, raw_city.upper()), set()))
                if corrected:
                    used_zip = corrected
                    corrected_zip_idxs.append(idx)

        county = zip_to_county.get(used_zip, "")
        return pd.Series({
            "Postal_Zip_Code": used_zip,
            "County": county,
            "City": best_city,
            "Prov_State": best_state
        })

    # Apply corrections
    df_import["Postal_Zip_Code"] = df_import["Postal_Zip_Code"].apply(normalize_zip)
    corr = df_import.apply(lookup_and_correct, axis=1)
    df_import.update(corr)

    # final cleanups
    df_import["Postal_Zip_Code"].replace("00000", "", inplace=True)
    df_import["County"] = df_import["County"].str.replace(r"County", "", regex=True).str.strip()

    df_import['City'] = df_import['City'].apply(dmr.name_proper)

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
  
    template_path = resource_path(
        "resources", "templates", "CLIENT_CONTACT.xlsx")
    if not template_path.exists():
        error_message = f"Template file not found at {template_path}"
        log_fn("⚠️ " + error_message)
        raise FileNotFoundError(error_message)
    output_file_path.parent.mkdir(parents=True, exist_ok=True)
    # Copy the template file to the output_file_path (this creates the new file based on the template)
    shutil.copy(template_path, output_file_path)

    # Load the workbook from the newly created file
    wb = load_workbook(output_file_path)

    # Select the "Data" sheet; ensure the sheet name matches exactly
    ws = wb[sheet_name]

    # Assume that the first row in the template is the header.
    # We will write the data starting at row 2.
    start_row = 2

    # Make sure that the dataframe's columns match the header in the template.
    for r_idx, row in enumerate(df_import.itertuples(index=False),
                                start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for dv in ws.data_validations.dataValidation:
        # dv.ranges is a list of range strings (e.g. ["A2:A100"])
        new_ranges = []
        for rng in dv.ranges:
            # Parse the original range boundaries (start_col, start_row, end_col, end_row)
            rng_str = str(rng)  # ensure we pass a string, e.g. "A2:F100"
            min_col, min_row, max_col, max_row = range_boundaries(rng_str)

            # Update the end row to the current maximum row with data (ws.max_row)
            new_range = (
                f"{ws.cell(row=min_row, column=min_col).coordinate}:" 
                f"{ws.cell(row=ws.max_row, column=max_col).coordinate}"
            )
            new_ranges.append(new_range)
        # Replace the validation's range list with the updated ranges
        dv.ranges = new_ranges
        # Save the workbook to preserve changes
    wb.save(output_file_path)
    #log_fn("✅ Counties successfully processed.")
    return output_file_path
   
        
