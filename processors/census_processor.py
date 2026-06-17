#VERSION: 1.1.1

import re
import numpy as np
import os
import pandas as pd
from pathlib import Path
import datetime
import sys
import shutil
import threading
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.cell_range import MultiCellRange
from utils import data_migration_resources as dmr
from utils.exceptions import AbortedByUser
from utils.file_utils import resource_path
from utils.file_utils import get_template_path

#root_abs_path = Path(__file__).resolve().parent.parent.parent

def process_census_file(
        facility_code,
        facility_name,
        facility_type,
        input_file_path,
        output_file_path,
        log_fn=None,
        abort_event=threading.Event()
):
    if log_fn is None:
       log_fn = lambda msg: None

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    if not isinstance(output_file_path, Path):
        output_file_path = Path(output_file_path)
    
    # Read the census file
    try:
        df_census_raw = pd.read_excel(input_file_path)
        
    except Exception as e:
        error_message = f"Error reading census file: {str(e)}"
        log_fn(error_message)
        raise Exception(error_message)

    all_cols_in = \
        dmr.import_columns_checker('Census',
                                   df_census_raw.columns.tolist())
    if not all_cols_in[0]:
        log_fn("⚠️ Not all expected columns are present in the input file.")
        log_fn(f"Missing: {all_cols_in[1]}")
        log_fn("⚠️ Cannot process this file at this time.")
        raise ValueError("This file does not have the expected columns.")
    
    log_fn("✅ Successfully read the census file")

    df_census_c1 = df_census_raw.dropna(how='all').reset_index(drop=True)
    if df_census_c1.empty:
        log_fn("⚠️ The input file is empty or has no valid data.")
        raise ValueError("The input file is empty or has no valid data.")

    def dataframe_checker(df, name):
        message =''
        if df.shape[0]:
            message = f"{name} has {df.shape[0]} rows"
        else:
            messages = f"{name} is empty"
        return message
    
    df_census_c1['Effective_Date'] = pd.to_datetime(df_census_c1['Start Date'], 
                                                    errors='coerce')
    resid_mask = df_census_c1['Effective_Date'].isna()

    df_census_c1['Resident'] = None
    df_census_c1['Client_ID_Number'] = None

    id_pattern = r"(\d+)(-\d+)?$"

    # this gives a DataFrame with one column (0) holding the digits
    id_extracted = df_census_c1.loc[resid_mask, 'Start Date'] \
                    .str.extract(id_pattern)

    # assign group 1 into Client_ID_Number
    df_census_c1.loc[resid_mask, 'Client_ID_Number'] = id_extracted[0]

    df_census_c1.loc[resid_mask, 'Resident'] = (
        df_census_c1.loc[resid_mask, 'Start Date']
        .str.replace(id_pattern, '', regex=True)
        .str.strip()
    )

    df_census_c1['Resident'] = df_census_c1['Resident'].ffill()
    df_census_c1['Client_ID_Number'] = df_census_c1['Client_ID_Number'].ffill()

    df_census_c2 = df_census_c1[~resid_mask].copy()
    if df_census_c2.empty:
        log_fn("⚠️ No valid data found in the census file after processing.")
        raise ValueError("No valid data found in the census file after processing.")    
    
    df_census_c2 = df_census_c2.sort_values(by=['Resident', 'Effective_Date'],
                                            ascending=[True, False])
    df_census_c2 = df_census_c2.reset_index(drop=True)

    #status_unique = list(df_census_c2['Status'].unique())
    if facility_type == 'SNF':

        # lookup dicts
        leaving_events = {
            'hospital leave': ('HP', 'TO'),
            'discharged':       ('D',  'DD'),
            'discharge':        ('D',  'DD'),
            'therapeutic leave':('TP', 'L'),
            'expired':          ('D',  'DE'),
        }
        admission_events = {
            'admission': ('A',  'AA'),
            'return':    ('A',  'RA'),
        }
        in_house_events = {
            'payer_change': ('A', 'PC'),
            'room_change':  ('A', 'RC'),
        }
    else:

        leaving_events = {
            'hospital leave': ('HML', 'L'),
            'discharged':       ('D',  'MO'),
            'discharge':        ('D',  'MO'),
            'therapeutic leave':('HML', 'L'),
            'expired':          ('D',  'DC'),
        }
        admission_events = {
            'admission': ('I',  'MI'),
            'return from leave':    ('I',  'RL'),
            'readmission':('I',  'MI')
        }
        in_house_events = {
            'payer_change': ('I', 'PC'),
            'room_change':  ('I', 'AC'),
        }


    # def _get_codes(status, lookup):
    #     """Return (Status_Code, Action_Code) for the first key found in status.lower()."""
    #     s = status.lower()
    #     for key, (sc, ac) in lookup.items():
    #         if key in s:
    #             return sc, ac
    #     return None, None


    def normalize_status(status):
        return re.sub(r"\s+", " ", str(status).lower().strip())


    def classify_census_status(status, facility_type):
        s = normalize_status(status)

        if facility_type == "SNF":
            rules = [
                (r"\bexpired\b", ("D", "DE")),
                (r"\bdischarg(?:e|ed)\b", ("D", "DD")),
                (r"\bhospital\s+leave\b", ("HP", "TO")),
                (r"\btherapeutic\s+leave\b", ("TP", "L")),
                (r"\breturn(?:ed)?\b", ("A", "RA")),
                (r"\b(?:admission|admitted|readmission|readmitted)\b", ("A", "AA")),
            ]
        else:
            rules = [
                (r"\bexpired\b", ("D", "DC")),
                (r"\bdischarg(?:e|ed)\b", ("D", "MO")),
                (r"\b(?:hospital|therapeutic)\s+leave\b", ("HML", "L")),
                (r"\breturn(?:ed)?\b", ("I", "RL")),
                (r"\b(?:admission|admitted|readmission|readmitted|move\s*in)\b", ("I", "MI")),
            ]

        for pattern, codes in rules:
            if re.search(pattern, s):
                return codes

        return (None, None)


    sel_idx = []
    sel_sc  = []
    sel_ac  = []

    for client_id, grp in df_census_c2.groupby('Resident'):
        # sort chronologically
        grp = grp.sort_values('Effective_Date')
        #status_lc = grp['Status'].str.lower()
        grp = grp.copy()
        grp[["Event_Status_Code", "Event_Action_Code"]] = grp["Status"].apply(
            lambda x: pd.Series(classify_census_status(x, facility_type))
        )

        # 1) find the oldest admission
        adm_mask = grp["Event_Action_Code"].isin(["AA", "RA", "MI", "RL"])
        if not adm_mask.any():
            try:
                log_fn(f"Warning: no admission found for Client_ID_Number {client_id!r}")
            except Exception as e:
                print(e)
            continue

        adm_row = grp.loc[adm_mask].iloc[0]
        idx_adm = adm_row.name
        sel_idx.append(idx_adm)
        sc = adm_row["Event_Status_Code"]
        ac = adm_row["Event_Action_Code"]
        sel_sc.append(sc); sel_ac.append(ac)

        # 2) find the most recent leaving event
        leave_mask = grp["Event_Action_Code"].isin(["DD", "DE", "TO", "L", "MO", "DC"])
        if leave_mask.any():
            last_leave = grp.loc[leave_mask].iloc[-1]
            idx_leave = last_leave.name
            sel_idx.append(idx_leave)
            sc = last_leave['Event_Status_Code']
            ac = last_leave['Event_Action_Code']
            sel_sc.append(sc); sel_ac.append(ac)

            # if that leave is the very last row, stop here
            if last_leave['Effective_Date'] == grp['Effective_Date'].max():
                continue

            cutoff = last_leave['Effective_Date']
            # 3) admission after that leaving?
            after_leave = grp[grp['Effective_Date'] > cutoff]
            adm2_mask = after_leave['Event_Action_Code'].isin(
                ['AA', 'RA', 'MI', 'RL']
            )
            if adm2_mask.any():
                adm2 = after_leave.loc[adm2_mask].iloc[-1]
                idx_adm2 = adm2.name
                sel_idx.append(idx_adm2)
                sc = adm2['Event_Status_Code']
                ac = adm2['Event_Action_Code']
                sel_sc.append(sc); sel_ac.append(ac)

        # 4) only consider the very last event for an in-house change
        events = grp.sort_values('Effective_Date')
        if len(events) > 1:
            # get the last two rows
            last_idx = events.index[-1]
            prev_idx = events.index[-2]
        
            # if we haven’t already selected this row (as admission or leave)…
            if last_idx not in sel_idx:
                curr = events.loc[last_idx]
                prev = events.loc[prev_idx]
        
                room_changed  = curr['Unit/Room/Bed'] != prev['Unit/Room/Bed']
                payer_changed = curr['Payer']         != prev['Payer']
        
                if room_changed or payer_changed:
                    # figure out whether it’s a payer_change or room_change
                    key = 'payer_change' if payer_changed else 'room_change'
                    sc, ac = in_house_events[key]
        
                    sel_idx.append(last_idx)
                    sel_sc .append(sc)
                    sel_ac .append(ac)

    # build the new df
    df_census_c3 = df_census_c2.loc[sel_idx].copy()
   
    df_census_c3['Status_Code'] = sel_sc
    df_census_c3['Action_Code'] = sel_ac

    # (optional) sort nicely by client and date
    df_census_c3 = df_census_c3.sort_values(
        ['Resident', 'Effective_Date']
    )

    # split into three new columns
    df_census_c3[['Unit','Room','Bed']] = (
       df_census_c3['Unit/Room/Bed']
        .str.split('/', expand=True)
        .apply(lambda col: col.str.strip())  # trim any extra whitespace
)
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    df_census_c3['Facility_Code'] = facility_code
    df_census_c3['Effective_Date'] = \
        df_census_c3['Effective_Date'].dt.strftime('%m/%d/%Y')
    df_census_c3['Place_Holder'] = ''
    df_census_c3['Floor'] = '1st Floor'

    columns_to_keep_snf = ['Facility_Code', 'Client_ID_Number', 'Effective_Date',
                    'Status_Code', 'Action_Code', 'Place_Holder', 'Payer',
                    'Place_Holder', 'Place_Holder', 'Place_Holder', 'Place_Holder',
                    'Place_Holder', 'Place_Holder', 'Place_Holder', 'Unit', 
                    'Floor', 'Room', 'Bed', 'Place_Holder', 'Place_Holder', 
                    'Resident'
                    ]

    columns_to_keep_alf = ['Facility_Code', 'Client_ID_Number', 'Effective_Date',
                    'Status_Code', 'Action_Code', 'Place_Holder', 'Payer',
                    'Place_Holder', 'Place_Holder', 'Place_Holder',
                    'Unit', 'Floor', 'Room', 'Bed', 'Place_Holder', 'Place_Holder',
                    'Resident'
                    ]

    if facility_type == 'SNF':
        df_census = df_census_c3[columns_to_keep_snf].copy()
    else:
        df_census = df_census_c3[columns_to_keep_alf].copy()

    df_census = df_census.fillna("")

    log_fn(f"✅ {len(df_census['Client_ID_Number'].unique())} unique residents processed.") 


    # Adjust the path as necessary.
    template_path = get_template_path("Resident_CENSUS.xlsx")
    # template_path = resource_path(
    #     "resources", "templates", "Census.xlsx")
    if not template_path.exists():
        error_message = f"Template file not found at {template_path}"
        log_fn("⚠️ " + error_message)
        raise FileNotFoundError(error_message)

    output_file_path.parent.mkdir(parents=True, exist_ok=True)
    # Copy the template file to the output_file_path (this creates the new file based on the template)
    shutil.copy(template_path, output_file_path)

    # Load the workbook from the newly created file
    wb = load_workbook(output_file_path)

    # Select the appropriated sheet; ensure the sheet name matches exactly
    if facility_type == 'SNF':
        sheetname = 'SNF Data'
    else:
        sheetname = 'ALF ILF Data'
    ws = wb[sheetname]

    # Assume that the first row in the template is the header.
    # We will write the data starting at row 2.
    start_row = 2

    # Get the dataframe containing your final allergy data.
    # Make sure that the dataframe's columns match the header in the template.
    for r_idx, row in enumerate(df_census.itertuples(index=False), 
                                start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for dv in ws.data_validations.dataValidation:
        # dv.ranges is a list of range strings (e.g. ["A2:A100"])
        new_ranges = []
        for rng in dv.ranges:
            # Parse the original range boundaries (start_col, start_row, end_col, end_row)
            rng_str = str(rng)  # ensure we pass a string, e.g. "A2:F100"
            min_col, min_row, max_col, max_row = range_boundaries(rng_str)

            # Update the end row to the current maximum row with data (ws.max_row)
            new_range = (
                f"{ws.cell(row=min_row, column=min_col).coordinate}:" 
                f"{ws.cell(row=ws.max_row, column=max_col).coordinate}"
            )
            new_ranges.append(new_range)
        # Replace the validation's range list with the updated ranges
        dv.ranges = new_ranges
        # Save the workbook to preserve changes
    wb.save(output_file_path)
    wb.close()
