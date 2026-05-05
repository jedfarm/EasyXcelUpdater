#VERSION: 1.0.0

import os
import pandas as pd
from pathlib import Path
import numpy as np
import re
import openpyxl
import threading
import shutil
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from utils import data_migration_resources as dmr
from utils.exceptions import AbortedByUser
from utils.file_utils import resource_path
from utils.file_utils import get_template_path
# new: go up three levels to land in src/easyxcel
#root_abs_path = Path(__file__).resolve().parent.parent.parent

def process_weights_file(
    facility_code,
    facility_name,
    facility_type,
    input_file_path,
    output_file_path,
    log_fn=None,
    abort_event=threading.Event()
):
    if log_fn is None:
       log_fn = lambda msg: None

    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    if not isinstance(output_file_path, Path):
        output_file_path = Path(output_file_path)
    # Read the weights file
    try:
        df_weights_raw = pd.read_excel(input_file_path)
    except Exception as e:
        raise Exception(f"Error reading weights file: {str(e)}")

    all_cols_in = \
        dmr.import_columns_checker('Weights',
                                   df_weights_raw.columns.tolist())
    if not all_cols_in[0]:
        if len(all_cols_in[1]) == 1 and 'HOUR' in all_cols_in[1]:
            log_fn("⚠️ 'Hour' column is missing, but work can still be done.")
        else:
            log_fn("⚠️ Not all expected columns are present in the weights file.")
            log_fn(f"Missing: {all_cols_in[1]}")
            log_fn("⚠️ Cannot process this file at this time.")
            raise ValueError("Weights file does not have the expected columns.")

    column_names = [col.title().strip().replace('\n', ' ')
                    for col in df_weights_raw.columns.tolist()]
    df_weights_raw.columns = column_names

    log_fn("✅ Weights file successfully loaded.")

    # Data Cleaning
    std_vitals_id = '1'

    # Droping empty rows
    df_weights_c1 = \
        df_weights_raw.dropna(axis=0, 
                              how='all').reset_index(drop=True).copy()
    if df_weights_c1.empty:
        log_fn("⚠️ No data to process in the weights file.")
        raise ValueError("Weights file is empty or has no valid data.")
        
    if "Hour" in column_names:
        df_weights_c1['Datetime'] = np.where(
            df_weights_c1['Hour'].notna(),
            df_weights_c1['Date/Time Recorded'].astype(str) + ' ' + \
            df_weights_c1['Hour'].astype(str),
            df_weights_c1['Date/Time Recorded']
        )
        
        df_weights_c1['Datetime'] = pd.to_datetime(df_weights_c1['Datetime'],
                                                   errors='coerce')
    else:
        df_weights_c1['Datetime'] = \
            pd.to_datetime(df_weights_c1['Date/Time Recorded'], errors='coerce')

    # Fixing issues with column spread of resident information
    join_mask = df_weights_c1['Resident'].notna() &\
        df_weights_c1['Date/Time Recorded'].notna() &\
        df_weights_c1['Datetime'].isna() 

    df_weights_c1.loc[join_mask, 'Resident'] = (
        df_weights_c1.loc[join_mask, 'Resident'].astype(str).str.strip() + ' ' +
        df_weights_c1.loc[join_mask, 'Date/Time Recorded'].astype(str).str.strip()
    )
  
    res_pattern = r"^([A-z'-]+, +[A-z]+ *[A-z]+?) +(\d+)"
    df_weights_c1['Resident_Name'] = None
    df_weights_c1['Client_ID_Number'] = np.nan

    df_weights_c1[['Resident_Name','Client_ID_Number']] = (
        df_weights_c1['Resident']
        .str.extract(res_pattern, expand=True)
    )

    df_weights_c1['Resident_Name'] = df_weights_c1['Resident_Name'].ffill()
    df_weights_c1['Client_ID_Number'] = df_weights_c1['Client_ID_Number'].ffill()
    
    if abort_event.is_set():
        raise AbortedByUser("⚠️ Process aborted by user.")

    mask_1 = (df_weights_c1['Date/Time Recorded'].notna()) &\
             (df_weights_c1['Weight'].notna())
            
    df_weights_c2 = df_weights_c1[mask_1].reset_index(drop=True).copy()
    
    if df_weights_c2.empty:
        log_fn("⚠️ No valid data to process in the weights file.")
        raise ValueError("Weights file has no valid data after cleaning.")

    # 1) normalize (lowercase & strip)
    w = df_weights_c2['Weight'].astype(str).str.lower().str.strip()

    # 2) extract number + optional unit (lbs?|kg)
    pattern = r'^(?P<number>\d+(?:\.\d+)?)(?:\s*(?P<unit>lbs?|kg))?$'
    ext = w.str.extract(pattern)

    # 3) cast the number to float
    ext['number'] = ext['number'].astype(float)

    # 4) set factor = 2.20462 only for kg, else 1
    ext['factor'] = np.where(ext['unit'] == 'kg', 2.20462, 1)

    # 5) compute your final Value (always in pounds)
    df_weights_c2['Value'] = ext['number'] * ext['factor']

    # (optional) round if you like
    df_weights_c2['Value'] = df_weights_c2['Value'].round(2)

    df_weights_c3 = (
        df_weights_c2
        .dropna(subset=['Value'])
        .reset_index(drop=True)
    )
    
    if abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    if df_weights_c3.empty:
        log_fn("No valid weight data to process after cleaning.")
        raise ValueError("⚠️ Weights file has no valid weight data after cleaning.")

    df_weights_c4 = \
        df_weights_c3.drop_duplicates(subset=['Client_ID_Number', 'Value',
                                              'Datetime'])

    df_weights_c4 = df_weights_c4.reset_index(drop=True)

    df_weights_c4['Datetime'] = \
        df_weights_c4['Datetime'].apply(dmr.format_date_or_datetime)

    df_weights_c4['Facility_Id'] = facility_code
    df_weights_c4['Std_Vitals_ID'] = std_vitals_id
    df_weights_c4['Type'] = ''

    ordered_columns = ['Facility_Id', 'Client_ID_Number', 'Std_Vitals_ID',
                       'Datetime', 'Value', 'Type', 'Resident_Name']

    df_weights = df_weights_c4[ordered_columns].copy()   
    num_residents = df_weights['Client_ID_Number'].nunique()
    log_fn(f"✅ Processed {num_residents} unique residents' weights.")

    # Write to an Excel Template
    template_path = get_template_path("WEIGHTS_AND_VITALS.xlsx")
    #template_path = resource_path("resources", "templates","WEIGHTS_AND_VITALS.xlsx")
    if not template_path.exists():
        error_message = f"Template file not found at {template_path}"
        log_fn("⚠️ " + error_message)
        raise FileNotFoundError(error_message)
    output_file_path.parent.mkdir(parents=True, exist_ok=True)
    # Copy the template file to the output_file_path (this creates the new file 
    # # based on the template)
    shutil.copy(template_path, output_file_path)

    # Load the workbook from the newly created file
    wb = load_workbook(output_file_path)

    # Select the "Data" sheet; ensure the sheet name matches exactly
    ws = wb['Data']

    # Assume that the first row in the template is the header.
    # We will write the data starting at row 2.
    start_row = 2

    # Get the dataframe containing your final diagnosis data.
    # In your code, this is likely stored in a variable such as df_diagnosis.
    # Make sure that the dataframe's columns match the header in the template.
    for r_idx, row in enumerate(df_weights.itertuples(index=False), 
                                start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save(output_file_path)

    wb.close()
