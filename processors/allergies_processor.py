#VERSION: 1.0.0

from pathlib import Path
import os
import re
import pandas as pd
import requests
import shutil
import threading
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.utils import range_boundaries
from utils import data_migration_resources as dmr
from utils.exceptions import AbortedByUser
from utils.file_utils import resource_path

# new: go up three levels to land in src/easyxcel
root_abs_path = Path(__file__).resolve().parent.parent.parent


def process_allergies_file(
        facility_code,
        facility_name,
        facility_type,
        input_file_path,
        output_file_path,
        log_fn=None,
        abort_event=threading.Event()
):
    if log_fn is None:
        log_fn = lambda msg: None

    #log_fn(f"root_abs_path type: {type(root_abs_path)} value: {root_abs_path}")

    if abort_event is not None and abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    if not isinstance(output_file_path, Path):
        output_file_path = Path(output_file_path)

    try:
        df_allergies_raw = pd.read_excel(input_file_path, header=0)
        
    except Exception as e:
        error_message = f"Error reading allergies file: {str(e)}"
        log_fn(error_message)
        raise Exception(error_message)
        
    all_cols_in = \
        dmr.import_columns_checker('Allergies',
                                   df_allergies_raw.columns.tolist())
    if not all_cols_in[0]:
        log_fn("⚠️ Not all expected columns are present in the allergies file.")
        log_fn(f"Missing: {all_cols_in[1]}")
        log_fn("⚠️ Cannot process this file at this time.")
        raise ValueError("This file does not have the expected columns.")
    
    log_fn("✅ Allergy file successfully loaded.")

    # read the allergen database
    try:
        allergen_data_path = resource_path(
            "resources", "databases", "CommonAllergensMiniDatabase.xlsx")
        df_allergens_small = pd.read_excel(allergen_data_path)
    except Exception as e:
        error_message = f"Error reading allergen database: {str(e)}"
        log_fn("⚠️ Error reading allergen database: " + error_message)
        raise Exception(error_message)

    # PROCESSING DATA
    df_allergies_c1 = df_allergies_raw.dropna(how='all')
    df_allergies_c1 = df_allergies_c1.copy()
    if df_allergies_c1.empty:
        log_fn("⚠️ No data found in the allergies file.")
        raise ValueError("The input file is empty or has no valid data.")   

    df_allergies_c1['MR#'] = df_allergies_c1['MR#'].ffill()
    df_allergies_c1['Name'] = df_allergies_c1['Name'].ffill()
    df_allergies_c1['Client_ID_Number'] = df_allergies_c1['MR#'].str.strip() 

    df_allergies_c2 = df_allergies_c1.copy()

    if df_allergies_c2.empty:
        log_fn("⚠️ No data found in the allergies file.")
        raise ValueError("The input file is empty or has no valid data.")
    else:
        log_fn(f"Processing {len(df_allergies_c2['Client_ID_Number'].unique())} unique residents")

    def cat_finder(s, df_allergens_database=df_allergens_small):
        """
        This function attempts to identify a category for the input string `s`
        by looking it up in `df_allergens_database`.
        
        It follows these steps:
        1. Normalize the input string (trim spaces and convert to uppercase).
        2. If the string contains parentheses, split the string into two parts:
            - The part outside the parentheses (primary term)
            - The part inside the parentheses (secondary term)
            * If a closing parenthesis is missing, everything after the opening parenthesis 
            is considered the inside term.
        3. Search using the outside term first.
        4. If no result is found and an inside term exists, search using the inside term.
        5. Return the category found, or an empty string if nothing is found.
        """
        df = df_allergens_database
        s = str(s).strip().upper()
        
        # Initialize outside and inside search terms
        outside_term = s
        inside_term = ""
        
        # Check if there's an open parenthesis
        if '(' in s:
            # Use the text before '(' as the primary (outside) term.
            start_idx = s.index('(')
            outside_term = s[:start_idx].strip()
            
            # If there's a closing parenthesis after the open parenthesis, extract the inside term.
            # Otherwise, take everything after the '(' as the inside term.
            if ')' in s[start_idx:]:
                end_idx = s.index(')', start_idx)
                inside_term = s[start_idx+1:end_idx].strip()
            else:
                inside_term = s[start_idx+1:].strip()
        
        # Define a helper function to perform the lookup with a given term.
        def search_term(term):
            # Immediate classification for certain keywords.
            if 'VACCINE' in term or 'IODINE' in term:
                return 'DRUG'
            
            # Escape regex special characters in the search term.
            term_escaped = re.escape(term)
            
            # First, search in the 'NAME' column.
            found = df[df['NAME'].str.contains(term_escaped, regex=True, na=False)]
            
            # If nothing is found, search in the 'ALT NAME' column.
            if len(found) == 0:
                found = df[df['ALT NAME'].str.contains(term_escaped, regex=True, na=False)]
            
            # If exactly one match is found, return its category.
            if len(found) == 1:
                return found['CATEGORY'].iloc[0]
            # If multiple matches exist, try to pick an entry that exactly matches
            # the length of the trimmed search term.
            elif len(found) > 1:
                for i in found.index:
                    if len(term.strip()) == len(found['NAME'].loc[i]):
                        return found['CATEGORY'].loc[i]
            return ""
        
        # First, perform the search using the outside term.
        category = search_term(outside_term)
        
        # If no category is found and an inside term exists, try searching with the inside term.
        if not category and inside_term:
            category = search_term(inside_term)
            
        return category

    if abort_event is not None and abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    # Applying the function
    df_allergies_c2['Category'] = df_allergies_c2['Allergy'].apply(cat_finder)


    def is_drug(s):
        base_endpoint = 'https://api.fda.gov/drug/drugsfda.json?'
        API_key = 'api_key=D23TxBWehyqwWqE187cMa4AUZ47FfQEqliDOtZWN'
        brand_name = 'openfda.brand_name:'
        limit = 'limit=2'
        
        pattern = r'\([^)]*\)'
        s = str(s)
        
        # If the input string is empty, return False
        if len(s) == 0:
            return False
        
        # First, remove anything inside parentheses (including the parentheses)
        s_no_paren = re.sub(pattern, '', s)
        clean_s_no_paren = ' '.join(s_no_paren.split())  # Normalize spaces
        
        # Build URL and search with the cleaned string (without parentheses)
        drug = clean_s_no_paren
        URL = f"{base_endpoint}{API_key}&search={brand_name}{drug}&{limit}"
        data = requests.get(URL).json()
        
        # If results are found using the cleaned string, return True
        if 'results' in data and len(data['results']) >= 1:
            return True
        elif 'error' in data and data['error'].get('code') == 'NOT_FOUND':
            # If there was no result and there is information within parentheses,
            # then extract the text inside the parentheses.
            paren_matches = re.findall(pattern, s)
            if paren_matches:
                # Remove the surrounding parentheses from each match and join if multiple
                drug_from_parens = ' '.join(match[1:-1].strip() for match in paren_matches)
                URL_paren = f"{base_endpoint}{API_key}&search={brand_name}{drug_from_parens}&{limit}"
                data_paren = requests.get(URL_paren).json()
                if 'results' in data_paren and len(data_paren['results']) >= 1:
                    return True
        # If no valid results found with either approach, return False
        return False
        
    def find_drugs(cat, allergen):
        output = ''
        if cat == '':
            result = is_drug(str(allergen))
            if result:
                output = 'Drug'
        else:
            output = cat
        return output

    df_allergies_c2['Category'] = df_allergies_c2.apply(lambda row: find_drugs(row['Category'], row['Allergy']), axis=1)

    df_allergies_c2 = df_allergies_c2.rename(columns={'Allergy': 'Allergen', 
                                        'Start Date': 'Onset Date'})
    df_allergies_c2['Category'] = df_allergies_c2['Category'].str.title()
    
    if abort_event is not None and abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")

    def capitalize_except_parentheses(text):
        text = str(text)
        # Splitting the text into words and parentheses groups
        parts = re.findall(r'\([^)]*\)|\b\w+\b', text)
        new_parts = []
        inside_parentheses = False

        for part in parts:
            # Check if the part is within parentheses
            if part.startswith('(') and part.endswith(')'):
                inside_parentheses = True
            else:
                inside_parentheses = False

            # Capitalize the part if not inside parentheses
            if not inside_parentheses:
                new_parts.append(part.capitalize())
            else:
                new_parts.append(part)

        return ' '.join(new_parts)

    # Applying the function to the 'Allergen' column
    df_allergies_c2['Allergen'] = \
        df_allergies_c2['Allergen'].apply(capitalize_except_parentheses)
    if 'Severity' not in df_allergies_c2.columns:
        df_allergies_c2['Severity'] = ''
        
    df_allergies_c3 = df_allergies_c2[['Client_ID_Number', 'Onset Date', 
                                       'Category', 'Allergen', 'Severity',
                                       'Name']]
    df_allergies_c3 = df_allergies_c3.copy()

    df_allergies_c4 = df_allergies_c3.drop_duplicates(
        subset=['Client_ID_Number', 'Allergen'])
    df_allergies_c4.reset_index(drop=True, inplace=True)

    df_allergies_c4['Facility_Code'] = facility_code
    df_allergies_c4['Status'] = 'Active'
    df_allergies_c4['Resolved Date'] = ''
    df_allergies_c4['Type'] = ''
    df_allergies_c4['Reaction_Type'] = ''
    df_allergies_c4['Reaction_Sub_Type'] = ''
    df_allergies_c4['Reaction_Note'] = ''
    df_allergies_c4['Severity'] = df_allergies_c4['Severity'].fillna('')

    ordered_cols = ['Facility_Code', 'Client_ID_Number', 'Status', 'Onset Date',
                    'Resolved Date', 'Category', 'Type', 'Severity', 'Reaction_Type',
                    'Reaction_Sub_Type', 'Allergen','Reaction_Note', 'Name']

    df_allergies = df_allergies_c4[ordered_cols].copy()
    num_residents = df_allergies['Client_ID_Number'].nunique()
    log_fn(f"✅ Processed {num_residents} unique residents' allergies.")    
    log_fn("⚠️ Please, review the Category column, the assignments are not always 100% right.")

    if abort_event is not None and abort_event.is_set():
        raise AbortedByUser("Process aborted by user.")
    
    template_path = resource_path(
        "resources", "templates", "ALLERGY.xlsx")
    if not template_path.exists():
        error_message = f"Template file not found at {template_path}"
        log_fn("⚠️ " + error_message)
        raise FileNotFoundError(error_message)
    
    output_file_path.parent.mkdir(parents=True, exist_ok=True)

    # Copy the template file to the output_file_path (this creates the new file based on the template)
    shutil.copy(template_path, output_file_path)

    # Load the workbook from the newly created file
    wb = load_workbook(output_file_path)

    # Select the "Data" sheet; ensure the sheet name matches exactly
    ws = wb['Data']

    # Assume that the first row in the template is the header.
    # We will write the data starting at row 2.
    start_row = 2

    # Get the dataframe containing your final allergy data.
    # Make sure that the dataframe's columns match the header in the template.
    for r_idx, row in enumerate(df_allergies.itertuples(index=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for dv in ws.data_validations.dataValidation:
        # dv.ranges is a list of range strings (e.g. ["A2:A100"])
        new_ranges = []
        for rng in dv.ranges:
            # Parse the original range boundaries (start_col, start_row, end_col, end_row)
            min_col, min_row, max_col, max_row = range_boundaries(str(rng))
            # Update the end row to the current maximum row with data (ws.max_row)
            new_range = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=ws.max_row, column=max_col).coordinate}"
            new_ranges.append(new_range)
        # Replace the validation's range list with the updated ranges
        dv.ranges = new_ranges
        # Save the workbook to preserve changes
    wb.save(output_file_path)
    wb.close()


    
