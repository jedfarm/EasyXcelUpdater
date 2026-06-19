#VERSION: 1.0.3

import os
import shutil
from openpyxl.utils import range_boundaries
from openpyxl import load_workbook
import threading
import pandas as pd
from pathlib import Path
from utils import data_migration_resources as dmr
from utils.exceptions import AbortedByUser
from utils.file_utils import resource_path
from utils.file_utils import get_template_path

def process_filter_residents_file(
    import_file_path: str,
    residents_file_path: str,
    output_file_path: str,
    log_fn=None,
    abort_event=None
):
    if log_fn is None:
        log_fn = lambda msg: None

    if abort_event is None:
        abort_event = threading.Event()

    try:
        if abort_event.is_set():
            raise AbortedByUser("Process aborted by user.")

        if not isinstance(output_file_path, Path):
            output_file_path = Path(output_file_path)

        log_fn("Reading import file...")
        xlsx = pd.ExcelFile(import_file_path)
        sheet_name = xlsx.sheet_names[0]
        df_import = pd.read_excel(import_file_path)

        log_fn("Reading residents file...")
        residents_required_cols = [
            "Client_ID_Number",
            "Last_Name",
            "First_Name",
            "Middle_Name"
        ]

        try:
            df_residents = pd.read_excel(
                residents_file_path,
                usecols=residents_required_cols
            )
        except Exception as e:
            log_fn("⚠️ Could not read the residents file with the expected columns.")
            log_fn(f"Expected columns: {', '.join(residents_required_cols)}")
            log_fn(f"Details: {e}")
            return None

        if df_import.empty:
            log_fn("⚠️ The import file is empty.")
            return None

        if df_residents.empty:
            log_fn("⚠️ The residents file is empty.")
            return None

        column_names = df_import.columns

        if "Status_Code" in column_names and "Action_Code" in column_names:
            template = "Resident_CENSUS.xlsx"
            log_fn("Census template selected.")

        elif "Reaction_Type" in column_names and "Allergen" in column_names:
            template = "ALLERGY.xlsx"
            log_fn("Allergy template selected.")

        elif "Ethnicity" in column_names and "Race_Code" in column_names:
            template = "CLIENT.xlsx"
            log_fn("Client template selected.")
            log_fn("⚠️ This import is the reference, it should not be used as a target.")

        elif "Client_Name" in column_names and "Balance_Due" in column_names:
            template = "CLIENT_BALANCE_FORWARD.xlsx"
            log_fn("Client Balance Forward template selected.")

        elif "Relation_Code" in column_names and "ContactType01" in column_names:
            template = "CLIENT_CONTACT.xlsx"
            log_fn("Client Contact template selected.")

        elif "ICD_Code" in column_names and "Diag_Classification_ID" in column_names:
            template = "CLIENT_DIAGNOSIS.xlsx"
            log_fn("Client Diagnosis template selected.")

        elif "Std_Vitals_ID" in column_names and "WV_Date" in column_names:
            template = "WEIGHTS_AND_VITALS.xlsx"
            log_fn("Weights and Vitals template selected.")

        else:
            log_fn("⚠️ EasyXcel does not support this import file.")
            log_fn(f"Columns found: {', '.join(map(str, df_import.columns))}")
            return None

        duplicated_ids = (
            df_residents[df_residents.duplicated(subset=["Client_ID_Number"], keep=False)]
            ["Client_ID_Number"]
            .dropna()
            .unique()
            .tolist()
        )

        if duplicated_ids:
            log_fn("⚠️ The CLIENT file contains duplicated Client_ID_Number values:")
            for resident_id in duplicated_ids:
                log_fn(f"- {resident_id}")
            log_fn("We suggest resolving this issue before proceeding.")

        orig_cols = df_import.columns.tolist()

        if "Client_ID_Number" not in df_import.columns:
            log_fn("⚠️ Import file has no Client_ID_Number column.")

            if "Resident" not in df_import.columns:
                log_fn("⚠️ Import file also has no Resident column. Cannot filter residents.")
                return None

            df_import["Client_ID_Number"] = pd.NA

        import_ids = (
            df_import["Client_ID_Number"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        all_empty_or_nan = import_ids.eq("").all()

        df_residents["Client_ID_Number"] = (
            df_residents["Client_ID_Number"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        if all_empty_or_nan:
            log_fn("⚠️ No Client_ID_Number found in the import file.")

            if "Resident" not in df_import.columns:
                log_fn("⚠️ No Resident column found. Cannot match by resident name.")
                return None

            df_import = df_import.drop(columns=["Client_ID_Number"], errors="ignore")

            df_residents["Last_Name"] = df_residents["Last_Name"].fillna("").astype(str)
            df_residents["First_Name"] = df_residents["First_Name"].fillna("").astype(str)
            df_residents["Middle_Name"] = df_residents["Middle_Name"].fillna("").astype(str)

            df_import["Resident"] = (
                df_import["Resident"]
                .fillna("")
                .astype(str)
                .str.upper()
                .str.strip()
            )

            has_comma_format = df_import["Resident"].str.contains(",", na=False).any()

            if has_comma_format:
                df_residents["Resident_1"] = (
                    df_residents["Last_Name"] + ", " +
                    df_residents["First_Name"] + " " +
                    df_residents["Middle_Name"]
                )

                df_residents["Resident_2"] = (
                    df_residents["Last_Name"] + ", " +
                    df_residents["First_Name"]
                )

            else:
                df_residents["Resident_1"] = (
                    df_residents["First_Name"] + " " +
                    df_residents["Middle_Name"] + " " +
                    df_residents["Last_Name"]
                )

                df_residents["Resident_2"] = (
                    df_residents["First_Name"] + " " +
                    df_residents["Last_Name"]
                )

            df_residents["Resident_1"] = (
                df_residents["Resident_1"]
                .fillna("")
                .astype(str)
                .str.upper()
                .str.replace(r"\s+", " ", regex=True)
                .str.strip()
            )

            df_residents["Resident_2"] = (
                df_residents["Resident_2"]
                .fillna("")
                .astype(str)
                .str.upper()
                .str.replace(r"\s+", " ", regex=True)
                .str.strip()
            )

            df_import["Resident"] = (
                df_import["Resident"]
                .str.replace(r"\s+", " ", regex=True)
                .str.strip()
            )

            residents_long = (
                df_residents[["Client_ID_Number", "Resident_1", "Resident_2"]]
                .assign(
                    Resident_variants=lambda df: df[["Resident_1", "Resident_2"]].values.tolist()
                )
                .explode("Resident_variants")
                .rename(columns={"Resident_variants": "Resident_cleaned"})
            )

            residents_long = residents_long[
                residents_long["Resident_cleaned"].fillna("").astype(str).str.strip() != ""
            ]

            df_import_filtered = (
                df_import
                .merge(
                    residents_long,
                    left_on="Resident",
                    right_on="Resident_cleaned",
                    how="inner"
                )
                .reset_index(drop=True)
            )

            for col in orig_cols:
                if col not in df_import_filtered.columns:
                    df_import_filtered[col] = ""

            df_import_filtered = df_import_filtered[orig_cols]

        else:
            df_import["Client_ID_Number"] = import_ids

            def ensure_same_id_format_and_merge(
                df_import,
                df_residents,
                log_fn=print,
                col="Client_ID_Number",
                threshold=0.5
            ):
                s1 = df_import[col].fillna("").astype(str).str.strip()
                s2 = df_residents[col].fillna("").astype(str).str.strip()

                nonempty1 = s1.ne("") & s1.str.lower().ne("nan")
                nonempty2 = s2.ne("") & s2.str.lower().ne("nan")

                if not nonempty1.any():
                    log_fn(f"⚠️ No non-empty {col} values found in the import file.")
                    return pd.DataFrame(columns=df_import.columns)

                if not nonempty2.any():
                    log_fn(f"⚠️ No non-empty {col} values found in the residents file.")
                    return pd.DataFrame(columns=df_import.columns)

                pct_dash_1 = s1[nonempty1].str.contains("-", regex=False).mean()
                pct_dash_2 = s2[nonempty2].str.contains("-", regex=False).mean()

                import_has_dash = pct_dash_1 > threshold
                residents_has_dash = pct_dash_2 > threshold

                if import_has_dash != residents_has_dash:
                    log_fn("⚠️ Please make sure both files use the same Client_ID_Number format.")
                    log_fn(f"Import file has '-' in {pct_dash_1:.1%} of IDs.")
                    log_fn(f"Residents file has '-' in {pct_dash_2:.1%} of IDs.")
                    return pd.DataFrame(columns=df_import.columns)

                df_import[col] = s1
                df_residents[col] = s2

                return (
                    df_import
                    .merge(df_residents[[col]], on=col, how="inner")
                    .reset_index(drop=True)
                )

            df_import_filtered = ensure_same_id_format_and_merge(
                df_import,
                df_residents,
                log_fn=log_fn
            )

        if df_import_filtered.empty:
            log_fn("⚠️ No matching residents found.")
            log_fn("Please make sure this is the right residents file.")
            return None

        residents_on_file = set(
            df_import_filtered["Client_ID_Number"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        residents_all = set(
            df_residents["Client_ID_Number"]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        residents_not_on_file = sorted(
            rid for rid in residents_all - residents_on_file if rid != ""
        )

        if residents_not_on_file:
            log_fn("Residents not found in the import file:")
            for resident_id in residents_not_on_file:
                log_fn(f"- {resident_id}")
        else:
            log_fn("All residents found in the import file.")

        if abort_event.is_set():
            raise AbortedByUser("Process aborted by user.")

        template_path = get_template_path(template)

        if not template_path.exists():
            log_fn(f"⚠️ Template file not found: {template_path}")
            return None

        output_file_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy(template_path, output_file_path)

        wb = load_workbook(output_file_path)

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        elif "Data" in wb.sheetnames:
            ws = wb["Data"]
        else:
            wb.close()
            log_fn(f"⚠️ Template does not contain sheet '{sheet_name}' or 'Data'.")
            return None

        for r_idx, row in enumerate(df_import_filtered.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        for dv in ws.data_validations.dataValidation:
            new_ranges = []

            for rng in dv.ranges:
                rng_str = str(rng)
                min_col, min_row, max_col, max_row = range_boundaries(rng_str)

                new_range = (
                    f"{ws.cell(row=min_row, column=min_col).coordinate}:"
                    f"{ws.cell(row=ws.max_row, column=max_col).coordinate}"
                )

                new_ranges.append(new_range)

            dv.ranges = new_ranges

        wb.save(output_file_path)
        wb.close()

        log_fn(f"✅ Output created: {output_file_path}")

        return output_file_path, template

    except AbortedByUser:
        log_fn("⚠️ Process aborted by user.")
        return None

    except Exception as e:
        log_fn("✘ Filter Residents failed.")
        log_fn(f"Error type: {type(e).__name__}")
        log_fn(f"Details: {e}")
        return None


# def process_filter_residents_file(
#     import_file_path: str,
#     residents_file_path: str,
#     output_file_path: str,
#     log_fn = None,
#     abort_event=threading.Event()
# ):
#     xlsx = pd.ExcelFile(import_file_path)
#     sheet_name = xlsx.sheet_names[0]
#
#     df_import = pd.read_excel(import_file_path)
#
#     if log_fn is None:
#         log_fn = lambda msg: None
#
#     if abort_event.is_set():
#         raise AbortedByUser("Process aborted by user.")
#
#     if not isinstance(output_file_path, Path):
#         output_file_path = Path(output_file_path)
#
#     # Determine which template to use (same as in your standalone script)
#     column_names = df_import.columns
#     if 'Status_Code' in column_names and 'Action_Code' in column_names:
#         template = 'Resident_CENSUS.xlsx'
#         log_fn("Census template selected.")
#     elif 'Reaction_Type' in column_names and 'Allergen' in column_names:
#         template = 'ALLERGY.xlsx'
#         log_fn("Allergy template selected.")
#     elif 'Ethnicity' in column_names and 'Race_Code' in column_names:
#         template = 'CLIENT.xlsx'
#         log_fn("Client template selected.")
#         log_fn("⚠️ This import is the reference, it should not be used as a target.")
#
#     elif 'Client_Name' in column_names and 'Balance_Due' in column_names:
#         template = 'CLIENT_BALANCE_FORWARD.xlsx'
#         log_fn("Client Balance Forward template selected.")
#     elif 'Relation_Code' in column_names and 'ContactType01' in column_names:
#         template = 'CLIENT_CONTACT.xlsx'
#         log_fn("Client Contact template selected.")
#     elif 'ICD_Code' in column_names and\
#           'Diag_Classification_ID' in column_names:
#         template = 'CLIENT_DIAGNOSIS.xlsx'
#         log_fn("Client Diagnosis template selected.")
#     elif 'Std_Vitals_ID' in column_names and 'WV_Date' in column_names:
#         template = 'WEIGHTS_AND_VITALS.xlsx'
#         log_fn("Weights and Vitals template selected.")
#     else:
#         template = None
#
#     if template is None:
#         log_fn("EasyXcel does not support this import file.")
#         raise Exception("Unable to pick a template for this import")
#
#     df_residents = pd.read_excel(residents_file_path, usecols=['Client_ID_Number',
#                                     'Last_Name', 'First_Name', 'Middle_Name'])
#
#     duplicated_ids = df_residents[df_residents.duplicated(subset=['Client_ID_Number'])]['Client_ID_Number'].unique().tolist()
#     if duplicated_ids:
#         log_fn("⚠️ The CLIENT file contains duplicates:")
#             for resident_id in duplicated_ids:
#                log_fn(f"- {resident_id}")
#
#         log_fn("We suggest to resolve this issue before proceeding")
#
#     orig_cols = df_import.columns.tolist()
#     mask = df_import['Client_ID_Number'].isna() | (df_import['Client_ID_Number'] == '')
#     all_empty_or_nan = mask.all()
#
#     if all_empty_or_nan:
#         log_fn("⚠️ No Client_ID_Number found in the import file. ")
#         if 'Resident' in df_import.columns:
#             df_import = df_import.drop('Client_ID_Number', axis=1)
#             if any(df_import['Resident'].str.contains(',')):
#                 df_import['Resident'] = df_import['Resident'].str.upper()
#                 df_residents['Resident_1'] = df_residents['Last_Name'] \
#                     +", "+ df_residents['First_Name'] + " "+ df_residents['Middle_Name']
#                 df_residents['Resident_2'] = df_residents['Last_Name'] \
#                     +", "+ df_residents['First_Name']
#                 df_residents['Resident_1'] = df_residents['Resident_1'].str.upper().str.strip()
#                 df_residents['Resident_2'] = df_residents['Resident_2'].str.upper().str.strip()
#             else:
#                 df_residents['Resident_1'] = df_residents['First_Name'] \
#                     +" "+ df_residents['Middle_Name'] + " "+ df_residents['Last_Name']
#                 df_residents['Resident_2'] = df_residents['First_Name'] \
#                     +" "+ df_residents['Last_Name']
#                 df_residents['Resident_1'] = df_residents['Resident_1'].str.upper().str.strip()
#                 df_residents['Resident_2'] = df_residents['Resident_2'].str.upper().str.strip()
#
#             columns_to_keep =['Client_ID_Number', 'Resident_1', 'Resident_2']
#             df_residents = df_residents[columns_to_keep].copy()
#             df_import['Resident'] = \
#                 df_import['Resident'].str.upper().str.strip()
#
#             # 3) turn those two columns into one “lookup” column …
#             residents_long = (
#                 df_residents
#                 .assign(Resident_variants=lambda df:
#                         df[['Resident_1','Resident_2']].values.tolist())
#                 .explode('Resident_variants')
#                 .rename(columns={'Resident_variants':'Resident_cleaned'})
#             )
#
#             # 4) now do a single merge on your cleaned key
#             df_import_f1 = (
#                 df_import
#                 .merge(residents_long,
#                         left_on='Resident',
#                         right_on='Resident_cleaned',
#                         how='inner')
#                 .reset_index(drop=True)
#             )
#
#             df_import_filtered = df_import_f1[orig_cols]
#
#         else:
#             log_fn("We cannot do the requested operation")
#             raise Exception("No Client_ID_Number or Resident column found in import file.")
#     else:
#         def ensure_same_id_format_and_merge(df_import, df_residents,
#                                             log_fn=print, col="Client_ID_Number", threshold=0.5):
#             # Clean strings (note: astype(str) turns NaN into 'nan', so we filter those out)
#             s1 = df_import[col].astype(str).str.strip()
#             s2 = df_residents[col].astype(str).str.strip()
#
#             nonempty1 = (s1 != "") & (s1.str.lower() != "nan")
#             nonempty2 = (s2 != "") & (s2.str.lower() != "nan")
#
#             if not nonempty1.any() or not nonempty2.any():
#                 msg = (f"No non-empty {col} values to compare: "
#                     f"import_n={int(nonempty1.sum())}, residents_n={int(nonempty2.sum())}.")
#                 log_fn(msg)
#                 raise ValueError(msg)
#
#             pct_dash_1 = s1[nonempty1].str.contains("-").mean()  # fraction in [0,1]
#             pct_dash_2 = s2[nonempty2].str.contains("-").mean()
#
#             # Ambiguous exactly-at-50% cases → refuse to guess
#             if pct_dash_1 == 0.5 or pct_dash_2 == 0.5:
#                 msg = (f"Ambiguous {col} format: import has '-' in {pct_dash_1:.1%}, "
#                     f"residents in {pct_dash_2:.1%}. Unable to determine a majority style.")
#                 log_fn(msg)
#                 raise ValueError(msg)
#
#             import_has_dash = pct_dash_1 > threshold
#             residents_has_dash = pct_dash_2 > threshold
#
#             if import_has_dash != residents_has_dash:
#                 msg = (f"{col} format mismatch: import has '-' in {pct_dash_1:.1%} "
#                     f"(n={int(nonempty1.sum())}), residents in {pct_dash_2:.1%} "
#                     f"(n={int(nonempty2.sum())}).")
#                 #log_fn(msg)
#                 log_fn("⚠️ Please, make sure both files use the same Client_ID_Number format.")
#                 raise ValueError(msg)
#
#             # If we get here, formats are alike (both mostly have '-' or both mostly don't)
#             df_import[col] = s1
#             df_residents[col] = s2
#             return (
#                 df_import.merge(df_residents, on=col, how="inner")
#                         .reset_index(drop=True)
#             )
#
#         # Usage:
#
#         df_import_filtered = \
#             ensure_same_id_format_and_merge(df_import,
#                                             df_residents, log_fn=log_fn)
#
#     if df_import_filtered.shape[0] == 0:
#         log_fn("No matching residents found. \n"
#         "Please, make sure is the right residents file or that Client_ID_Number is correct.")
#         raise Exception("No matching residents")
#     else:
#         residents_on_file = set(df_import_filtered['Client_ID_Number'])
#         residents_all = set(df_residents['Client_ID_Number'])
#         residents_not_on_file = residents_all - residents_on_file
#         if residents_not_on_file: # Check if the set is not empty
#             log_fn("Residents not found in the import file:")
#             for resident_id in sorted(list(residents_not_on_file)): # Sort for consistent output
#                log_fn(f"- {resident_id}")
#         else:
#             log_fn("All residents found in the import file.")
#
#     if abort_event.is_set():
#         raise AbortedByUser("Process aborted by user.")
#     # Adjust the path as necessary.
#     template_path = get_template_path(template)
#     #template_path = resource_path("resources", "templates", template)
#     if not template_path.exists():
#         error_message = f"Template file not found at {template_path}"
#         log_fn("⚠️ " + error_message)
#         raise FileNotFoundError(error_message)
#
#     output_file_path.parent.mkdir(parents=True, exist_ok=True)
#
#     # Copy the template file to the output_file_path (this creates the new file based on the template)
#     shutil.copy(template_path, output_file_path)
#
#     # Load the workbook from the newly created file
#     wb = load_workbook(output_file_path)
#
#     # Select the "Data" sheet; ensure the sheet name matches exactly
#     ws = wb[sheet_name]
#
#     # Assume that the first row in the template is the header.
#     # We will write the data starting at row 2.
#     start_row = 2
#
#     # Get the dataframe containing your final allergy data.
#     # Make sure that the dataframe's columns match the header in the template.
#     for r_idx, row in enumerate(df_import_filtered.itertuples(index=False),
#                                 start=start_row):
#         for c_idx, value in enumerate(row, start=1):
#             ws.cell(row=r_idx, column=c_idx, value=value)
#
#     for dv in ws.data_validations.dataValidation:
#         # dv.ranges is a list of range strings (e.g. ["A2:A100"])
#         new_ranges = []
#         for rng in dv.ranges:
#             # Parse the original range boundaries (start_col, start_row, end_col, end_row)
#             rng_str = str(rng)  # ensure we pass a string, e.g. "A2:F100"
#             min_col, min_row, max_col, max_row = range_boundaries(rng_str)
#
#             # Update the end row to the current maximum row with data (ws.max_row)
#             new_range = (
#                 f"{ws.cell(row=min_row, column=min_col).coordinate}:"
#                 f"{ws.cell(row=ws.max_row, column=max_col).coordinate}"
#             )
#             new_ranges.append(new_range)
#         # Replace the validation's range list with the updated ranges
#         dv.ranges = new_ranges
#         # Save the workbook to preserve changes
#     wb.save(output_file_path)
#
#     # user_downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
#     # os.makedirs(user_downloads_folder, exist_ok=True)  # Ensure the user's Downloads folder exists
#     # user_output_file_path =\
#     #     os.path.join(user_downloads_folder, os.path.basename(output_file_path))
#     # shutil.copy2(output_file_path, user_output_file_path)
#
#     # # Optionally, append debug messages to be displayed in your Log/Errors window:
#     # debug_messages.append(f"Excel file written to project downloads folder: {output_file_path}")
#     # debug_messages.append(f"Excel file also copied to user's Downloads folder: {user_output_file_path}")
#
#     return output_file_path, template


